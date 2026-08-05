"""
Retirement Income and Tax Planning Simulator v7
================================================
With accumulation phase, Monte Carlo, and Excel export.

Usage:
  pip install streamlit plotly pandas numpy openpyxl
  streamlit run retirement_v7.py

v7 enhancements (financial modeling, on top of v6's correctness fixes):
  1. Cash/MM interest is now taxed annually as ordinary income (was
     completely untaxed before -- both growth and withdrawals).
  2. Legacy pool has its own configurable return/std-dev assumption
     instead of automatically matching your own Roth account.
  3. Monte Carlo now correlates PreTax/Roth/HSA/Legacy Pool returns via a
     shared market factor (configurable strength) instead of drawing each
     bucket fully independently, which understated true sequence-of-returns
     tail risk. Cash/MM gets a smaller fraction of that correlation.
  4. Optional fat-tailed (Student-t) return distribution as an alternative
     to Normal, for more realistic crash frequency/severity.
  5. Optional "surviving spouse" scenario: filing status switches from MFJ
     to Single the year after the first spouse's death, with single-filer
     brackets/deduction/IRMAA threshold, a partial SS survivor benefit, and
     optional pension and living-expense reductions (the "widow's penalty").
  6. Legacy/estate figures now also show an after-tax-to-heirs value:
     inherited pretax/HSA money is taxed at an assumed heir rate under the
     SECURE Act's 10-year rule, while Roth (including the legacy pool)
     passes tax-free -- these are not equivalent dollars to your heirs.
  7. Optional dynamic "guardrails" spending strategy (Guyton-Klinger style)
     as an alternative to fixed real spending: a persistent spending cut/
     raise triggered by your trailing withdrawal rate drifting outside a
     band around your starting rate.
  8. Excel export gained a Monte Carlo sheet (percentile bands + success
     rate) whenever Monte Carlo is enabled.
  9. New Sensitivity (tornado) tab: varies one assumption at a time to show
     which ones actually move your outcome the most.

v6 correctness fixes (from a full audit of v5):
  1. Legacy Roth gifts were being subtracted from cash TWICE (once via
     total_exp/the draw engine, again via an explicit cash-to-Roth
     transfer) -- fixed to a single, reconciling debit/credit.
  2. In down-market years the model withdrew the legacy gift as an
     "expense" but never deposited it anywhere (a phantom
     Legacy_Inheritance figure with no real balance behind it) -- fixed
     so the money simply stays invested and isn't withdrawn at all.
  3. Account "depleted" flags were one-way and never reset, permanently
     freezing growth on any account that hit zero and later recovered
     (e.g. Roth refilled by a legacy gift) -- now evaluated live.
  4. Negative cash was floored to $0 in Cash_EOY and Total_Liquid_Assets,
     hiding real funding shortfalls and inflating Monte Carlo success
     rates -- now shown/counted honestly.
  5. Itemized deductions double-counted the non-taxable portion of JSS
     income (already excluded from taxable income, then deducted again).
  6. Oregon taxable income over-subtracted an extra 15% of Social
     Security that was never part of the taxable base to begin with.
  7. Oregon Roth-conversion tax used a flat 9% guess instead of the
     actual marginal OREGON_BRACKETS calculation used for federal.
  8. RMD table stopped at age 100 and silently reused the age-95 factor
     for older ages -- extended through the IRS floor of 2.0 at 120+.
  9. RMD was computed on the current year's already-grown balance
     instead of the prior year-end balance (overstated RMDs slightly).
 10. Social Security taxability used a flat $6,000 tier-2 add-on instead
     of the IRS-correct min($6,000, 50% of benefits) cap.
 11. The "performance_draw_only" cash cap always used the static average
     return even inside Monte Carlo runs, instead of that year's actual
     simulated return.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from typing import Dict, List, Tuple
import io
import os
from pathlib import Path

# ============================================================
# CONSTANTS
# ============================================================

RMD_TABLE = {
    72: 27.4, 73: 26.5, 74: 25.5, 75: 24.6, 76: 23.7, 77: 22.9,
    78: 22.0, 79: 21.1, 80: 20.2, 81: 19.4, 82: 18.5, 83: 17.7,
    84: 16.8, 85: 16.0, 86: 15.2, 87: 14.4, 88: 13.7, 89: 12.9,
    90: 12.2, 91: 11.5, 92: 10.8, 93: 10.1, 94: 9.5, 95: 8.9,
    96: 8.4, 97: 7.8, 98: 7.3, 99: 6.8, 100: 6.4,
    # Table previously stopped at 100 and silently fell back to the age-95
    # factor (8.9) for any older age, understating RMDs for anyone whose
    # planning horizon runs past 100. Full IRS Pub. 590-B Uniform Lifetime
    # Table (Appendix B, Table III) continues to a floor of 2.0 at 120+.
    101: 6.0, 102: 5.6, 103: 5.2, 104: 4.9, 105: 4.6, 106: 4.3,
    107: 4.1, 108: 3.9, 109: 3.7, 110: 3.5, 111: 3.4, 112: 3.3,
    113: 3.1, 114: 3.0, 115: 2.9, 116: 2.8, 117: 2.7, 118: 2.5,
    119: 2.3, 120: 2.0,
}
RMD_TABLE_FLOOR_FACTOR = 2.0  # "120 and older" per the IRS table
IRMAA_MAGI_THRESHOLD = 206_000
IRMAA_MAGI_THRESHOLD_SINGLE = 103_000  # single-filer IRMAA tier-1 threshold (~half of MFJ)
IRMAA_MONTHLY_SURCHARGE = 230.80
FEDERAL_BRACKETS = [
    (0.10, 24_800), (0.12, 100_800), (0.22, 211_400),
    (0.24, 403_550), (0.32, 512_450), (0.35, 768_700),
    (0.37, float("inf")),
]
# Single-filer brackets: ~half of MFJ through the 32% bracket (matches how the
# real brackets compare), then diverging at the top since MFJ gets
# disproportionately more room in the 35% bracket than a simple halving would give.
SINGLE_FEDERAL_BRACKETS = [
    (0.10, 12_400), (0.12, 50_400), (0.22, 105_700),
    (0.24, 201_775), (0.32, 256_225), (0.35, 609_350),
    (0.37, float("inf")),
]
OREGON_BRACKETS = [
    (0.0475, 8_824), (0.0675, 22_059),
    (0.0875, 250_000), (0.0990, 10_000_000),
]
# Oregon's single-vs-MFJ bracket differences are much smaller than federal's;
# OREGON_BRACKETS is used for both filing statuses as a documented simplification.

# ============================================================
# TAX ENGINE
# ============================================================

def calc_tax(taxable, brackets, yrs, infl):
    if taxable <= 0: return 0.0
    tax, prev = 0.0, 0.0
    for rate, ceil in brackets:
        ac = ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
        band = max(0.0, min(taxable, ac) - prev)
        tax += band * rate; prev = ac
        if taxable <= ac: break
    return tax

def bracket_ceiling(brackets, target_rate, yrs, infl):
    for rate, ceil in brackets:
        if rate >= target_rate:
            return ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
    return 0.0

def ss_taxable_portion(ss, other, single=False):
    if ss <= 0: return 0.0
    prov = other + ss * 0.5
    tier1, tier2 = (25_000, 34_000) if single else (32_000, 44_000)
    tier2_cap = 4_500 if single else 6_000  # single-filer worksheet cap is $4,500 (half of MFJ's $6,000)
    if prov < tier1: return 0.0
    elif prov < tier2: return min(ss * 0.5, (prov - tier1) * 0.5)
    else:
        # IRS worksheet caps the tier-2 add-on at min(cap, 50% of SS
        # benefits) — a flat cap overstates taxable SS whenever benefits
        # are low relative to the cap (0.5*ss < cap).
        tier2_addon = min(tier2_cap, ss * 0.5)
        return min(ss * 0.85, tier2_addon + (prov - tier2) * 0.85)

def get_rmd(bal, age, start):
    if age < start or bal <= 0: return 0.0
    if age in RMD_TABLE:
        f = RMD_TABLE[age]
    elif age > max(RMD_TABLE):
        f = RMD_TABLE_FLOOR_FACTOR  # IRS table floors at 2.0 for 120+
    else:
        f = RMD_TABLE[min(RMD_TABLE)]  # shouldn't happen given the age < start guard above
    return bal / f if f > 0 else 0.0

def _standardized_shock(n, rng, fat_tailed, t_df):
    """Mean-0, unit-variance shock series. Normal by default; Student-t gives
    fatter tails (more frequent/severe extreme years) when fat_tailed=True."""
    if fat_tailed:
        raw = rng.standard_t(t_df, size=n)
        return raw / np.sqrt(t_df / (t_df - 2))  # rescale to unit variance
    return rng.standard_normal(n)

def generate_returns(target, std, mx, n, rng, fat_tailed=False, t_df=5):
    """Independent (uncorrelated) return series for a single bucket."""
    z = _standardized_shock(n, rng, fat_tailed, t_df)
    return np.clip(target + std * z, -1.0, mx)

def generate_correlated_returns(bucket_specs, n, rng, correlation=0.85, fat_tailed=False, t_df=5):
    """
    bucket_specs: {name: (target, std, max_up)}.
    All buckets share one common market shock (weighted by `correlation`),
    plus their own idiosyncratic shock -- mirrors how real accounts holding
    similar underlying assets move together in the same year rather than
    being statistically independent of each other.
    """
    common = _standardized_shock(n, rng, fat_tailed, t_df)
    out = {}
    for name, (target, std, mx) in bucket_specs.items():
        idio = _standardized_shock(n, rng, fat_tailed, t_df)
        z = correlation * common + np.sqrt(max(0.0, 1 - correlation ** 2)) * idio
        out[name] = np.clip(target + std * z, -1.0, mx)
    return out

def build_mc_return_overrides(cfg, num_years, rng, std_dev, max_up, correlation=0.85, fat_tailed=False, t_df=5):
    """One year-by-year correlated return series per bucket for a single
    Monte Carlo path, including the legacy pool's own return/std assumption."""
    bucket_specs = {
        "pretax": (cfg["pretax_return"], std_dev, max_up),
        "roth": (cfg["roth_return"], std_dev, max_up),
        "hsa": (cfg["hsa_return"], std_dev, max_up),
        "legacy_pool": (cfg.get("legacy_pool_return", cfg["roth_return"]),
                         cfg.get("legacy_pool_std", std_dev), max_up),
    }
    ov = generate_correlated_returns(bucket_specs, num_years, rng, correlation, fat_tailed, t_df)
    # Cash/MM: much lower volatility, only lightly correlated with the
    # broader equity market (money-market rates track policy rates, not stocks).
    ov["cash"] = generate_correlated_returns(
        {"cash": (cfg["cash_return"], std_dev / 3, max_up / 2)},
        num_years, rng, correlation * 0.3, fat_tailed, t_df,
    )["cash"]
    return ov


def load_starting_balances(path="starting_balances.txt"):
    defaults = {
        "401k": None,
        "roth": None,
        "hsa": None,
        "s_plus_5": None,
        "s_plus_10": None,
        "cash": None,
    }
    if not Path(path).exists():
        return defaults, False
    aliases = {
        "401k": "401k",
        "401(k)": "401k",
        "pretax": "401k",
        "roth": "roth",
        "rothira": "roth",
        "roth_ira": "roth",
        "hsa": "hsa",
        "s+5": "s_plus_5",
        "s5": "s_plus_5",
        "s_plus_5": "s_plus_5",
        "s+10": "s_plus_10",
        "s10": "s_plus_10",
        "s_plus_10": "s_plus_10",
        "cash": "cash",
    }
    loaded = False
    try:
        for raw in Path(path).read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.replace(":", " ").split()
            if len(parts) < 2:
                continue
            key = parts[0].strip().lower().replace(" ", "").replace("-", "").replace("/", "")
            key = aliases.get(key)
            if key is None:
                continue
            try:
                defaults[key] = int(float(parts[1].replace(",", "").replace("$", "")))
                loaded = True
            except ValueError:
                pass
    except Exception:
        return defaults, False
    return defaults, loaded
# ============================================================
# ACCUMULATION PHASE
# ============================================================

def run_accumulation(cfg):
    """Grow balances from current age to retirement with contributions."""
    cur_age = cfg["current_age"]
    ret_age = cfg["retirement_age"]
    base_year = 2024
    years = ret_age - cur_age
    if years <= 0:
        return [], cfg["pretax_401k"], cfg["roth_ira"], cfg["hsa"], cfg["cash"], cfg["s_plus_5yr"], cfg["s_plus_10yr"]

    pt = float(cfg["pretax_401k"])
    ro = float(cfg["roth_ira"])
    hs = float(cfg["hsa"])
    ca = float(cfg["cash"])
    s5 = float(cfg["s_plus_5yr"])
    s10 = float(cfg["s_plus_10yr"])

    # Annual contributions
    c_401k = cfg.get("contrib_401k", 24_500)
    c_roth401k = cfg.get("contrib_roth401k", 8_000)
    c_roth_ira = cfg.get("contrib_roth_ira", 8_700)  # x2 for MFJ
    c_hsa = cfg.get("contrib_hsa", 8_700)
    c_mega_backdoor = cfg.get("contrib_mega_backdoor", 29_000)
    c_employer_match = cfg.get("contrib_employer_match", 18_000)
    c_cash_annual = cfg.get("contrib_cash_annual", 150_000)
    c_cash_final_lump = cfg.get("contrib_cash_final_lump", 120_000)

    # Growth rates (same as retirement performance assumptions)
    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]

    rows = []
    for i in range(years):
        age = cur_age + i
        yr = base_year + i

        # Growth first (skip year 0 — balances are "today")
        if i > 0:
            pt *= (1 + pr_pt)
            ro *= (1 + pr_ro)
            hs *= (1 + pr_hs)
            ca *= (1 + pr_ca)
            # S+ grows in deferred comp while employed
            s5 *= (1 + pr_pt)
            s10 *= (1 + pr_pt)

        # Contributions (end of year)
        pt += c_401k + c_employer_match
        ro += c_roth401k + c_roth_ira * 2 + c_mega_backdoor  # 2x Roth IRA for MFJ
        hs += c_hsa
        ca += c_cash_annual

        # Final year lump sum
        if i == years - 1:
            ca += c_cash_final_lump

        rows.append({
            "Phase": "Accumulation",
            "Age": age, "Year": yr,
            "PreTax_EOY": pt, "Roth_EOY": ro, "HSA_EOY": hs, "Cash_EOY": ca,
            "S_Plus_5yr": s5, "S_Plus_10yr": s10,
            "Contrib_PreTax": c_401k + c_employer_match,
            "Contrib_Roth": c_roth401k + c_roth_ira * 2 + c_mega_backdoor,
            "Contrib_HSA": c_hsa,
            "Contrib_Cash": c_cash_annual + (c_cash_final_lump if i == years - 1 else 0),
            "Total_Liquid_Assets": pt + ro + hs + ca,
        })

    return rows, pt, ro, hs, ca, s5, s10


# ============================================================
# RETIREMENT SIMULATION ENGINE
# ============================================================

def run_simulation(cfg, return_overrides=None):
    results = []
    base_year = 2024
    ret_age = cfg["retirement_age"]
    cur_age = cfg["current_age"]
    end_age = cfg["planning_end_age"]
    ret_year = base_year + (ret_age - cur_age)
    infl = cfg["inflation_rate"]
    binfl = cfg["bracket_inflation"]
    num_years = end_age - ret_age + 1

    # Run accumulation phase to get retirement-day balances
    accum_rows, pt, ro, hs, ca, s5_bal, s10_bal = run_accumulation(cfg)
    cash_basis = float(accum_rows[-1].get("Cash_Basis", cfg["cash"])) if accum_rows else float(cfg["cash"])

    # S+ payout tracking
    s5_annual = s5_bal / 5.0 if s5_bal > 0 else 0.0
    s10_annual = s10_bal / 10.0 if s10_bal > 0 else 0.0
    s5_rem, s10_rem = 5, 10
    jss_rec_rem = cfg["jss_recovery_years"]

    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]

    cum_gifts, cum_legacy_roth, cum_legacy_inheritance, cum_lump_sums = 0.0, 0.0, 0.0, 0.0
    CASH_DISTRESS_FLOOR = -50_000  # below this, we stop accruing "growth" on a cash deficit
    legacy_pool = 0.0  # money already gifted into the kids' Roth accounts -- legally
    # theirs, not the household's. Kept segregated from `ro` (the parents' own
    # Roth) so it (a) compounds on its own and can be charted, and (b) can
    # never be drawn back out by the household's own retirement withdrawals.

    # Guyton-Klinger-style dynamic guardrails (optional alternative to fixed
    # real spending). guardrail_factor is a persistent multiplier on base
    # spending -- it only changes when last year's withdrawal rate breached a
    # guardrail, and the change compounds forward (a permanent step, not a
    # one-year blip), same as the real strategy.
    use_guardrails = cfg.get("spending_strategy") == "guardrails"
    guardrail_factor = 1.0
    initial_wr = None
    prev_wr = None

    for idx in range(num_years):
        age = ret_age + idx
        yr = ret_year + idx
        yfb = yr - base_year
        yir = idx

        row = {"Phase": "Retirement", "Age": age, "Year": yr, "Years_Retired": yir}

        # Surviving-spouse ("widow's penalty") scenario: filing status switches
        # from MFJ to Single starting the year AFTER the first spouse's death
        # (the IRS allows MFJ status for the year of death itself). Off by
        # default -- when disabled, behavior is identical to before.
        is_widowed = bool(cfg.get("model_widow_scenario", False)) and age > cfg.get("first_death_age", 999)
        row["Filing_Status"] = "Single (Widowed)" if is_widowed else "MFJ"
        fed_brackets_yr = SINGLE_FEDERAL_BRACKETS if is_widowed else FEDERAL_BRACKETS
        std_ded_multiplier = 0.5 if is_widowed else 1.0  # single standard deduction is ~half of MFJ's
        irmaa_threshold_base = IRMAA_MAGI_THRESHOLD_SINGLE if is_widowed else IRMAA_MAGI_THRESHOLD

        # ── DYNAMIC GUARDRAILS (Guyton-Klinger style) ──
        # Evaluated using LAST year's ending withdrawal rate (the portfolio
        # condition known at the start of this year), before this year's
        # spending is set -- same sequencing as the real strategy.
        if use_guardrails and yir > 0 and initial_wr and prev_wr is not None:
            band = cfg.get("guardrail_band_pct", 0.20)
            adj = cfg.get("guardrail_adjustment_pct", 0.10)
            upper, lower = initial_wr * (1 + band), initial_wr * (1 - band)
            if prev_wr > upper:
                guardrail_factor *= (1 - adj)   # portfolio stressed -- cut spending
            elif 0 < prev_wr < lower:
                guardrail_factor *= (1 + adj)   # portfolio well ahead -- give a raise
        row["Guardrail_Factor"] = guardrail_factor

        # ── GROWTH ──
        # "Depleted" is now evaluated live off the current balance every year,
        # instead of a one-way flag that, once set, permanently froze growth
        # even after a balance recovered (e.g. a Roth that hit zero and was
        # later refilled by a legacy contribution or Roth conversion never
        # used to earn interest again under the old logic).
        if return_overrides:
            r_pt, r_ro = return_overrides["pretax"][yir], return_overrides["roth"][yir]
            r_hs, r_ca = return_overrides["hsa"][yir], return_overrides["cash"][yir]
            r_lp = return_overrides["legacy_pool"][yir] if "legacy_pool" in return_overrides else r_ro
        else:
            r_pt, r_ro, r_hs, r_ca = pr_pt, pr_ro, pr_hs, pr_ca
            r_lp = cfg.get("legacy_pool_return", pr_ro)

        # Snapshot the pretax balance as of the *prior* year-end, before this
        # year's growth is applied — RMDs are legally based on the 12/31
        # balance from the year before, not a balance that already includes
        # this year's market movement.
        pt_prior_year_end = pt
        ca_pre_growth = ca  # snapshot before growth, used to compute this year's taxable interest below

        if yir > 0:
            if pt > 0: pt *= (1 + r_pt)
            if ro > 0: ro *= (1 + r_ro)
            if hs > 0: hs *= (1 + r_hs)
            if ca > CASH_DISTRESS_FLOOR: ca *= (1 + r_ca)
            if legacy_pool > 0: legacy_pool *= (1 + r_lp)
            row["Return_PreTax"], row["Return_Roth"] = r_pt, r_ro
            row["Return_HSA"], row["Return_Cash"] = r_hs, r_ca
            row["Return_Legacy_Pool"] = r_lp
        else:
            row["Return_PreTax"] = row["Return_Roth"] = row["Return_HSA"] = row["Return_Cash"] = 0.0
            row["Return_Legacy_Pool"] = 0.0

        # Cash/MM interest is taxable annually as ordinary income (a money
        # market fund doesn't defer tax the way an equity account's
        # unrealized gains do) -- computed off the balance BEFORE this year's
        # growth and draws. Previously cash growth and withdrawals were never
        # taxed anywhere in the model.
        cash_interest_taxable = 0.0
        if cfg.get("tax_cash_interest", True) and yir > 0 and ca_pre_growth > 0:
            cash_interest_taxable = ca_pre_growth * r_ca
        row["Cash_Interest_Taxable"] = cash_interest_taxable

        # ── INCOME SOURCES ──
        sp_inc = 0.0
        if s5_rem > 0: sp_inc += s5_annual; s5_rem -= 1
        if s10_rem > 0: sp_inc += s10_annual; s10_rem -= 1
        row["S_Plus_Income"] = sp_inc

        jss_inc = jss_tax = 0.0
        if age >= cfg["jss_start_age"]:
            jss_inc = cfg["jss_annual_amount"] * (1 + cfg["jss_cola"]) ** (age - cfg["jss_start_age"])
            if jss_rec_rem > 0: jss_rec_rem -= 1
            else: jss_tax = jss_inc
        if is_widowed:
            # Pension survivorship: many plans pay a reduced (or zero) benefit
            # to the survivor. Default 100% (no change) unless configured.
            jss_survivor_pct = cfg.get("jss_survivor_pct", 1.0)
            jss_inc *= jss_survivor_pct
            jss_tax *= jss_survivor_pct
        row["JSS_Income"], row["JSS_Taxable"] = jss_inc, jss_tax

        ss_inc = 0.0
        if age >= cfg["ss_start_age"]:
            ss_inc = cfg["ss_annual_amount"] * (1 + cfg["ss_cola"]) ** (age - cfg["ss_start_age"])
        if is_widowed:
            # Social Security survivor benefit: the survivor keeps the HIGHER
            # of the two individual benefits, effectively losing the smaller
            # one. Modeled as a configurable % of the combined household
            # benefit retained (default 65%).
            ss_inc *= cfg.get("ss_survivor_pct", 0.65)
        row["SS_Income"] = ss_inc

        rent_inc = cfg["rental_gross"] * (1 + infl) ** yir
        rent_tax = rent_inc * cfg["rental_taxable_pct"]
        row["Rental_Income"], row["Rental_Taxable"] = rent_inc, rent_tax

        # ── EXPENSES ──
        base_exp = cfg["base_annual_expenses"] * (1 + infl) ** yir
        # Post-80 expense reduction (lifestyle slowdown)
        if age >= 80:
            base_exp *= (1 - cfg.get("expense_reduction_post80", 0.0))
        # Widowhood expense reduction (one person's living costs, on top of
        # any post-80 reduction that also applies)
        if is_widowed:
            base_exp *= (1 - cfg.get("expense_reduction_widowhood", 0.0))
        # Dynamic guardrails: a persistent step up/down based on the
        # portfolio's trailing withdrawal rate (distinct from, and stacks
        # with, the reductions above and the same-year bad-return cut below --
        # this reacts to multi-year trajectory, those react to lifestyle age
        # or a single bad year).
        if use_guardrails:
            base_exp *= guardrail_factor
        healthcare = (cfg["healthcare_pre_medicare"] if age < 65 else cfg["healthcare_post_medicare"]) * (1 + cfg["healthcare_inflation"]) ** yir
        hdhp = cfg["hdhp_annual"] * (1 + infl) ** yir if age < 65 else 0.0
        gifts = cfg["gifts_annual"] * (1 + infl) ** yir
        lump = cfg.get("lump_sums", {}).get(age, 0.0)
        legacy_target_per_child = 0.0
        legacy_target_total = 0.0
        if yir < cfg.get("legacy_years", 10) and cfg.get("num_children", 0) > 0:
            legacy_target_per_child = cfg["roth_legacy_per_child"] * (1 + cfg["legacy_inflation"]) ** yir
            legacy_target_total = cfg["num_children"] * legacy_target_per_child

        # ── DISCRETIONARY REDUCTION ON NEGATIVE RETURN ──
        # Moved ahead of total_exp (was after) so the legacy-funding decision
        # right below can use it.
        # Use pretax return as proxy for overall market return in that year
        is_negative_return = False
        if return_overrides and yir < len(return_overrides["pretax"]):
            if return_overrides["pretax"][yir] < 0:
                is_negative_return = True
        elif cfg["pretax_return"] < 0:
            is_negative_return = True

        # Only fund the legacy gift out of this year's cash flow when it will
        # actually be deposited into Roth below (i.e. not a down-market year).
        # Previously legacy_target_total was added to total_exp
        # *unconditionally*, so the draw engine withdrew it from real accounts
        # as "spending" even in years the code elsewhere skipped the actual
        # Roth gift -- the money vanished without ever landing anywhere.
        # Skipping it here means it simply stays invested and becomes
        # ordinary inheritance later (tracked below, not a separate pot).
        legacy_funding_this_year = legacy_target_total if (legacy_target_total > 0 and not is_negative_return) else 0.0

        total_exp = base_exp + healthcare + hdhp + gifts + lump + legacy_funding_this_year

        if is_negative_return:
            reduction = cfg.get("neg_ret_draw_reduction", 0.0)
            # Cannot reduce expenses below 0, but total_exp should be large enough
            reduction_applied = min(reduction, total_exp)
            total_exp -= reduction_applied
            row["Discretionary_Reduction"] = reduction_applied
        else:
            row["Discretionary_Reduction"] = 0.0

        row["Base_Expenses"], row["Healthcare_Cost"], row["HDHP"] = base_exp, healthcare, hdhp
        row["Gifts"], row["Lump_Sum"], row["Legacy_Roth"] = gifts, lump, legacy_funding_this_year
        row["Legacy_Target_Per_Child"] = legacy_target_per_child
        row["Legacy_Target_Total"] = legacy_target_total
        row["Total_Expenses"] = total_exp

        passive = sp_inc + jss_inc + ss_inc + rent_inc
        row["Passive_Income"] = passive

        rmd = get_rmd(pt_prior_year_end, age, cfg["rmd_start_age"])
        row["RMD"] = rmd

        # ── BRACKET-OPTIMIZED DRAW STRATEGY ──
        ptd = rod = hsd = cad = 0.0
        need = max(0.0, total_exp - passive)
        std_ded_est = cfg["standard_deduction"] * std_ded_multiplier * (1 + binfl) ** yfb
        base_taxable = jss_tax + rent_tax + sp_inc
        sst_est = ss_taxable_portion(ss_inc, base_taxable, single=is_widowed)
        existing_taxable = base_taxable + sst_est
        br12_gross = bracket_ceiling(fed_brackets_yr, 0.12, yfb, binfl)
        pretax_room_12 = max(0.0, br12_gross + std_ded_est - existing_taxable)
        is_rmd_phase = age >= cfg["rmd_start_age"]

        if is_rmd_phase:
            if pt > 0: ptd = min(rmd, pt); need = max(0.0, need - ptd)
            if need > 0 and ca > 0: d = min(need, ca); cad += d; need -= d
            if need > 0 and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "RMD-Dominated"
        else:
            if need > 0 and pt > 0:
                d = min(need, pretax_room_12, pt); ptd += d; need -= d
            if need > 0 and ca > 0:
                # Was always cfg's static average cash return (pr_ca), even
                # inside Monte Carlo runs where the realized draw for this
                # specific year is r_ca -- now uses the actual realized return.
                mx = ca * r_ca if cfg["performance_draw_only"] else ca
                d = min(need, max(0.0, mx)); cad += d; need -= d
            if need > 0 and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "Bracket-Optimized"

        if need > 0:
            for nm, avail in [("pretax", pt-ptd), ("cash", ca-cad),
                               ("roth", ro-rod), ("hsa", hs-hsd)]:
                if need <= 0: continue
                d = min(need, max(0.0, avail))
                if nm == "pretax": ptd += d
                elif nm == "cash": cad += d
                elif nm == "roth": rod += d
                elif nm == "hsa": hsd += d
                need -= d

        row["PreTax_Draw"], row["Roth_Draw"] = ptd, rod
        row["HSA_Draw"], row["Cash_Draw"] = hsd, cad
        row["RMD_Excess"] = max(0.0, rmd - (total_exp - passive)) if is_rmd_phase else 0.0

        # ── TAX ──
        other_taxable = ptd + jss_tax + rent_tax + sp_inc + cash_interest_taxable
        sst = ss_taxable_portion(ss_inc, other_taxable, single=is_widowed)
        gross_taxable = other_taxable + sst
        std_ded = cfg["standard_deduction"] * std_ded_multiplier * (1 + binfl) ** yfb
        med_ded = max(0.0, healthcare - 0.075 * gross_taxable)
        # NOTE: the non-taxable portion of JSS income (jss_inc - jss_tax) was
        # previously added here too. That income was never included in
        # gross_taxable to begin with (other_taxable only adds jss_tax), so
        # deducting it again created a phantom deduction that understated tax
        # every year JSS was partially or fully non-taxable.
        item_ded = med_ded + hdhp
        best_ded = max(std_ded, item_ded)
        fed_taxable = max(0.0, gross_taxable - best_ded)
        fed_tax = calc_tax(fed_taxable, fed_brackets_yr, yfb, binfl)
        # Oregon fully exempts Social Security. The only SS-related dollars
        # ever present in gross_taxable are `sst` (other_taxable has no raw
        # ss_inc term), so subtracting `sst` alone fully backs SS out of the
        # Oregon base. The old formula also subtracted an extra ss_inc*0.15
        # that was never part of gross_taxable, understating Oregon tax.
        or_taxable = max(0.0, gross_taxable - sst - best_ded)
        or_tax = calc_tax(or_taxable, OREGON_BRACKETS, yfb, binfl) if cfg["oregon_resident"] else 0.0
        total_tax = fed_tax + or_tax

        row["Gross_Taxable_Income"], row["Federal_Taxable_Income"] = gross_taxable, fed_taxable
        row["Standard_Deduction"], row["Itemized_Deduction"], row["Best_Deduction"] = std_ded, item_ded, best_ded
        row["Deduction_Type"] = "Itemized" if item_ded > std_ded else "Standard"
        row["Federal_Tax"], row["Oregon_Tax"], row["Total_Tax"] = fed_tax, or_tax, total_tax
        row["Effective_Tax_Rate"] = total_tax / gross_taxable if gross_taxable > 0 else 0.0

        magi = gross_taxable + (ss_inc - sst)
        irmaa_thr = irmaa_threshold_base * (1 + binfl) ** yfb
        irmaa_hit = age >= 65 and magi > irmaa_thr
        irmaa_people = 1 if is_widowed else 2  # only one person on Medicare once widowed
        irmaa_cost = (IRMAA_MONTHLY_SURCHARGE * 12 * irmaa_people * (1 + infl) ** yfb) if irmaa_hit else 0.0
        row["MAGI"], row["IRMAA_Threshold"] = magi, irmaa_thr
        row["IRMAA_Hit"], row["IRMAA_Cost"] = irmaa_hit, irmaa_cost
        fpl_700 = 7 * 20_440 * (1 + infl) ** yfb
        row["FPL_700"], row["Under_700_FPL"] = fpl_700, magi < fpl_700

        # ── ROTH CONVERSION ──
        roth_conv_amt = roth_conv_tax = 0.0
        if cfg["roth_conversion_enabled"] and pt - ptd > cfg["roth_conversion_margin"]:
            tgt = 0.12 if cfg["roth_conversion_target_bracket"] == "12%" else 0.22
            conv_room = max(0.0, bracket_ceiling(fed_brackets_yr, tgt, yfb, binfl) - fed_taxable)
            if cfg["irmaa_avoidance"] and age >= 63:
                conv_room = min(conv_room, max(0.0, irmaa_thr - magi))
            roth_conv_amt = min(conv_room, max(0.0, pt - ptd - cfg["roth_conversion_margin"]))
            if roth_conv_amt > 0:
                fc = calc_tax(fed_taxable + roth_conv_amt, fed_brackets_yr, yfb, binfl) - fed_tax
                if cfg["oregon_resident"]:
                    # Was a flat 9% guess, inconsistent with the marginal-bracket
                    # approach used for federal (and not matching any actual
                    # OREGON_BRACKETS rate at typical conversion income levels).
                    oc = calc_tax(or_taxable + roth_conv_amt, OREGON_BRACKETS, yfb, binfl) - or_tax
                else:
                    oc = 0.0
                roth_conv_tax = fc + oc
        row["Roth_Conversion"], row["Roth_Conversion_Tax"] = roth_conv_amt, roth_conv_tax

        # ── UPDATE BALANCES ──
        pt -= (ptd + roth_conv_amt); ro -= rod; ro += roth_conv_amt; hs -= hsd; ca -= cad
        total_income = passive + ptd + rod + hsd + cad
        surplus = total_income - total_exp - total_tax - roth_conv_tax - irmaa_cost
        ca += surplus

        legacy_roth_per_child = 0.0
        legacy_inheritance_per_child = 0.0
        legacy_roth_total = 0.0
        legacy_inheritance_total = 0.0
        legacy_actual = 0.0
        if legacy_target_total > 0 and cfg["num_children"] > 0:
            if legacy_funding_this_year > 0:
                # legacy_funding_this_year was already withdrawn from the
                # household's accounts above via total_exp/need (single
                # debit, spread across pt/ca/ro/hs by the normal draw
                # waterfall). It now moves into the segregated legacy_pool --
                # NOT back into the parents' own `ro` -- because once gifted
                # it's legally the kids' money, in the kids' own Roth
                # accounts, no longer part of the household's spendable net
                # worth and never available for the household's own future
                # withdrawals. (Previously this was credited back into `ro`,
                # which meant it was silently drawn back down years later by
                # the household's own Roth_Draw waterfall -- verified: with
                # no fix, gifted balances were being spent on the parents'
                # own retirement expenses in later years.)
                legacy_roth_total = legacy_funding_this_year
                legacy_roth_per_child = legacy_target_per_child
                legacy_pool += legacy_roth_total
                legacy_actual = legacy_roth_total
            # Any portion not funded to Roth this year (skipped in a
            # down-market year) was never withdrawn -- it's still sitting in
            # the household's own balances. This is a pure label for "will
            # pass as ordinary inheritance," not a separate pot of money.
            legacy_inheritance_per_child = max(0.0, legacy_target_per_child - legacy_roth_per_child)
            legacy_inheritance_total = legacy_inheritance_per_child * cfg["num_children"]

        cum_gifts += gifts
        cum_legacy_roth += legacy_roth_total
        cum_legacy_inheritance += legacy_inheritance_total
        cum_lump_sums += lump
        n_kids = cfg.get("num_children", 0)
        row["Surplus_Deficit"], row["Total_Income"] = surplus, total_income
        row["Bad_Return_Year"] = is_negative_return
        row["Legacy_Roth_Per_Child"] = legacy_roth_per_child
        row["Legacy_Inheritance_Per_Child"] = legacy_inheritance_per_child
        row["Legacy_Roth_Total"] = legacy_roth_total
        row["Legacy_Inheritance_Total"] = legacy_inheritance_total
        row["Legacy_Total"] = legacy_roth_total + legacy_inheritance_total  # this year's FLOW only, not a running total
        row["Legacy_Roth_Actual"] = legacy_actual
        row["Cum_Gifts"] = cum_gifts
        row["Cum_Legacy_Roth"] = cum_legacy_roth  # nominal sum of contributions made (no growth)
        row["Legacy_Pool_EOY"] = legacy_pool       # actual compounding balance of gifted money
        row["Legacy_Pool_Per_Child"] = (legacy_pool / n_kids) if n_kids else 0.0
        row["Cum_Legacy_Inheritance"] = cum_legacy_inheritance
        row["Cum_Legacy_Inheritance_Per_Child"] = (cum_legacy_inheritance / n_kids) if n_kids else 0.0
        # Running total legacy VALUE as of this point in time (grown Roth pool
        # + amounts still sitting in the household's own accounts earmarked
        # for inheritance). This is what "value at the end of the plan"
        # should read -- previously that metric read Legacy_Total (this
        # year's flow), which is 0 in nearly every year outside the
        # legacy_years gifting window, including almost always the final
        # simulated year of a multi-decade plan.
        row["Legacy_Value_To_Date"] = legacy_pool + cum_legacy_inheritance
        row["Legacy_Value_To_Date_Per_Child"] = row["Legacy_Pool_Per_Child"] + row["Cum_Legacy_Inheritance_Per_Child"]
        row["Cum_Lump_Sums"] = cum_lump_sums

        # Retirement/HSA accounts can't structurally go negative (draws are
        # already capped at the available balance upstream); this is just a
        # defensive floor for floating-point noise near zero.
        pt = max(0.0, pt)
        ro = max(0.0, ro)
        hs = max(0.0, hs)
        legacy_pool = max(0.0, legacy_pool)
        # Cash IS allowed to go negative -- it represents a genuine funding
        # shortfall (expenses the plan couldn't cover from any account).
        # Previously this was floored at $0 in both Cash_EOY and
        # Total_Liquid_Assets, which hid insolvency: a plan running a large,
        # growing cash deficit could still show a "positive" total and get
        # counted as a Monte Carlo success just because pt/ro/hs were still
        # positive. Now the deficit is shown and counted honestly.

        row["PreTax_EOY"], row["Roth_EOY"], row["HSA_EOY"] = pt, ro, hs
        row["Cash_EOY"] = ca
        total_liquid = pt + ro + hs + ca
        row["Total_Liquid_Assets"] = total_liquid  # parents' own spendable net worth (excludes gifted-away legacy_pool)
        row["Family_Net_Worth"] = total_liquid + legacy_pool  # includes money already gifted to the kids' Roths
        # "If death occurs at this age, what would each child actually
        # receive?" -- this is what Cum_Legacy_Inheritance does NOT capture
        # (that field only tracks a narrow sub-bucket: legacy-gift TARGET
        # amounts skipped in down-market years). The real answer is: your
        # own remaining accounts (Total_Liquid_Assets) pass to your heirs at
        # death, split evenly here across children, PLUS whatever's already
        # sitting in their own Roth accounts from prior gifting.
        row["Estate_At_Death"] = total_liquid  # what's left in your own accounts if you died this year
        row["Estate_At_Death_Per_Child"] = (total_liquid / n_kids) if n_kids else 0.0
        row["Total_Inheritance_At_Death_Per_Child"] = (row["Family_Net_Worth"] / n_kids) if n_kids else 0.0

        # After-tax value to heirs: not all dollars in the estate are worth
        # the same to your kids. Inherited pretax accounts (and HSAs, which
        # follow their own less-favorable rule) must be fully distributed
        # within 10 years under the SECURE Act and are taxed at the HEIR'S
        # own bracket -- unlike Roth, which passes tax-free. Cash is already
        # after-tax since its interest is taxed annually (see Cash_Interest_Taxable).
        heir_rate = cfg.get("heir_tax_rate", 0.24)
        pretax_after_tax_to_heirs = pt * (1 - heir_rate)
        hsa_after_tax_to_heirs = hs * (1 - heir_rate)  # HSAs lose tax-free status entirely for non-spouse heirs
        row["Heir_Tax_On_PreTax"] = pt * heir_rate
        row["Heir_Tax_On_HSA"] = hs * heir_rate
        after_tax_estate = pretax_after_tax_to_heirs + hsa_after_tax_to_heirs + ro + ca
        row["After_Tax_Estate_At_Death"] = after_tax_estate
        row["After_Tax_Family_Net_Worth"] = after_tax_estate + legacy_pool  # legacy pool is Roth -- already tax-free
        row["After_Tax_Estate_Per_Child"] = (after_tax_estate / n_kids) if n_kids else 0.0
        row["After_Tax_Total_Inheritance_Per_Child"] = (row["After_Tax_Family_Net_Worth"] / n_kids) if n_kids else 0.0

        row["Total_Real"] = (total_liquid / (1 + infl) ** yir) if yir > 0 else total_liquid

        total_draws = ptd + rod + hsd + cad
        row["Total_Draws"] = total_draws
        row["Withdrawal_Rate"] = total_draws / total_liquid if total_liquid > 0 else 0.0
        if use_guardrails:
            if yir == 0:
                initial_wr = row["Withdrawal_Rate"]
            prev_wr = row["Withdrawal_Rate"]

        br12 = bracket_ceiling(fed_brackets_yr, 0.12, yfb, binfl)
        br22 = bracket_ceiling(fed_brackets_yr, 0.22, yfb, binfl)
        row["Bracket_12_Ceiling"], row["Bracket_22_Ceiling"] = br12, br22

        results.append(row)

    return accum_rows, pd.DataFrame(results)


# ============================================================
# MONTE CARLO
# ============================================================

def run_monte_carlo(cfg, n_sims, std_dev, max_up, seed=None, correlation=0.85, fat_tailed=False, t_df=5):
    rng = np.random.default_rng(seed)
    num_years = cfg["planning_end_age"] - cfg["retirement_age"] + 1
    all_runs = []
    for _ in range(n_sims):
        ov = build_mc_return_overrides(cfg, num_years, rng, std_dev, max_up, correlation, fat_tailed, t_df)
        _, df = run_simulation(cfg, return_overrides=ov)
        all_runs.append(df)
    return all_runs

def compute_percentile_bands(runs, col, pcts=(5, 25, 50, 75, 95)):
    ages = runs[0]["Age"].values
    mx = np.array([r[col].values for r in runs])
    bands = {f"p{p}": np.percentile(mx, p, axis=0) for p in pcts}
    bands["mean"] = np.mean(mx, axis=0); bands["Age"] = ages
    return bands


# ============================================================
# SCENARIO OPTIMIZER
# ============================================================

def run_optimizer(base_cfg, mc_sims, mc_std, mc_max, mc_seed, correlation=0.85, fat_tailed=False, t_df=5):
    """
    Sweep retirement_age × ss_start_age × expense_reduction_post80
    and find combinations that achieve 100% MC success rate.
    Returns a DataFrame of all scenarios sorted by success rate.
    """
    rng_base = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)

    # Define search space
    ret_ages = list(range(base_cfg["retirement_age"], min(base_cfg["retirement_age"] + 4, 64)))
    ss_ages = list(range(max(62, base_cfg["ss_start_age"]), min(base_cfg["ss_start_age"] + 5, 71)))
    reductions = [0.0, 0.05, 0.10, 0.15, 0.20, 0.25]

    # Use fewer sims for optimizer (speed) — 50 is enough for ranking
    opt_sims = min(mc_sims, 50)
    num_years_max = base_cfg["planning_end_age"] - min(ret_ages) + 1

    results = []
    total_combos = len(ret_ages) * len(ss_ages) * len(reductions)

    for ret_age in ret_ages:
        for ss_age in ss_ages:
            if ss_age < ret_age:
                continue  # Can't claim SS before retirement in this model
            for red in reductions:
                test_cfg = {**base_cfg,
                            "retirement_age": ret_age,
                            "ss_start_age": ss_age,
                            "expense_reduction_post80": red}

                # Generate MC return paths (reuse same seed for fair comparison)
                rng = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)
                num_years = test_cfg["planning_end_age"] - ret_age + 1
                survived = 0
                final_assets = []
                final_real = []

                for _ in range(opt_sims):
                    ov = build_mc_return_overrides(test_cfg, num_years, rng, mc_std, mc_max, correlation, fat_tailed, t_df)
                    _, sim_df = run_simulation(test_cfg, return_overrides=ov)
                    last = sim_df.iloc[-1]
                    if last["Total_Liquid_Assets"] > 0:
                        survived += 1
                    final_assets.append(last["Total_Liquid_Assets"])
                    final_real.append(last["Total_Real"])

                success_rate = survived / opt_sims
                results.append({
                    "Retire_Age": ret_age,
                    "SS_Age": ss_age,
                    "Expense_Cut_80+": red,
                    "Success_Rate": success_rate,
                    "Median_Final_Assets": np.median(final_assets),
                    "Median_Final_Real": np.median(final_real),
                    "P10_Final_Assets": np.percentile(final_assets, 10),
                    "Worst_Case": min(final_assets),
                    "Extra_Work_Years": ret_age - base_cfg["retirement_age"],
                    "SS_Delay_Years": ss_age - base_cfg["ss_start_age"],
                })

    opt_df = pd.DataFrame(results).sort_values(
        ["Success_Rate", "Median_Final_Real"], ascending=[False, False]
    ).reset_index(drop=True)
    return opt_df


# ============================================================
# SENSITIVITY / TORNADO ANALYSIS
# ============================================================

# (label, cfg_key, mode, delta) -- mode is "additive" (rates/percentages,
# delta in absolute terms), "multiplicative" (dollar amounts, delta as a
# fraction), or "additive_int" (whole-number values like age).
SENSITIVITY_VARS = [
    ("Retirement Age", "retirement_age", "additive_int", 2),
    ("PreTax Return", "pretax_return", "additive", 0.015),
    ("Roth Return", "roth_return", "additive", 0.015),
    ("HSA Return", "hsa_return", "additive", 0.015),
    ("Cash/MM Return", "cash_return", "additive", 0.01),
    ("General Inflation", "inflation_rate", "additive", 0.01),
    ("Healthcare Inflation", "healthcare_inflation", "additive", 0.02),
    ("Base Annual Expenses", "base_annual_expenses", "multiplicative", 0.15),
    ("SS Annual Amount", "ss_annual_amount", "multiplicative", 0.15),
    ("Legacy Roth/Child/Year", "roth_legacy_per_child", "multiplicative", 0.30),
]

def run_sensitivity_analysis(cfg, variables=None):
    """
    Vary ONE assumption at a time (holding everything else at baseline),
    re-run the deterministic simulation, and record how much the final-year
    real (today's-dollar) value moves. Returns a DataFrame sorted by impact
    magnitude (smallest first, so the biggest driver plots at the top of a
    horizontal tornado chart), plus the baseline result for reference.
    """
    variables = variables or SENSITIVITY_VARS
    _, base_df = run_simulation(cfg)
    base_final = base_df["Total_Real"].iloc[-1]

    results = []
    for label, key, mode, delta in variables:
        base_val = cfg.get(key)
        if base_val is None:
            continue
        if mode == "additive":
            lo_val, hi_val = max(0.0, base_val - delta), base_val + delta
        elif mode == "additive_int":
            lo_val = max(cfg.get("current_age", 45), int(base_val) - delta)
            hi_val = min(70, int(base_val) + delta)
        else:  # multiplicative
            lo_val, hi_val = base_val * (1 - delta), base_val * (1 + delta)

        _, df_lo = run_simulation({**cfg, key: lo_val})
        _, df_hi = run_simulation({**cfg, key: hi_val})
        final_lo = df_lo["Total_Real"].iloc[-1]
        final_hi = df_hi["Total_Real"].iloc[-1]

        results.append({
            "Variable": label,
            "Base_Value": base_val, "Low_Value": lo_val, "High_Value": hi_val,
            "Low_Result": min(final_lo, final_hi), "High_Result": max(final_lo, final_hi),
            "Low_Is_High_Input": final_lo > final_hi,  # True if raising the input LOWERS the outcome
            "Base_Result": base_final,
            "Impact_Range": abs(final_hi - final_lo),
        })

    result_df = pd.DataFrame(results).sort_values("Impact_Range", ascending=True).reset_index(drop=True)
    return result_df, base_final


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_to_excel(accum_df, retire_df, cfg, mc_runs=None):
    """Create a professional Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    accum_fill = PatternFill("solid", fgColor="4472C4")
    input_font = Font(name="Arial", color="0000FF", size=10)  # Blue = inputs
    num_font = Font(name="Arial", size=10)
    pct_fmt = '0.0%'
    money_fmt = '$#,##0;($#,##0);"-"'
    money_k = '$#,##0'
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    def style_header(ws, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            mx = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 2, 10), 22)

    # ──────── Sheet 1: Assumptions ────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    assumptions = [
        ("RETIREMENT PLANNING MODEL", ""),
        ("", ""),
        ("TIMELINE", ""),
        ("Current Age", cfg["current_age"]),
        ("Retirement Age", cfg["retirement_age"]),
        ("Plan Through Age", cfg["planning_end_age"]),
        ("", ""),
        ("STARTING BALANCES (Today)", ""),
        ("PreTax 401(k)", cfg["pretax_401k"]),
        ("Roth IRA", cfg["roth_ira"]),
        ("HSA", cfg["hsa"]),
        ("S+ 5-Year", cfg["s_plus_5yr"]),
        ("S+ 10-Year", cfg["s_plus_10yr"]),
        ("Cash", cfg["cash"]),
        ("", ""),
        ("ANNUAL CONTRIBUTIONS (Working Years)", ""),
        ("401(k) Employee", cfg.get("contrib_401k", 24_500)),
        ("Roth 401(k)", cfg.get("contrib_roth401k", 8_000)),
        ("Roth IRA (per spouse)", cfg.get("contrib_roth_ira", 8_700)),
        ("HSA (family)", cfg.get("contrib_hsa", 8_700)),
        ("Mega Backdoor Roth", cfg.get("contrib_mega_backdoor", 29_000)),
        ("Employer Match", cfg.get("contrib_employer_match", 18_000)),
        ("Cash Savings", cfg.get("contrib_cash_annual", 150_000)),
        ("Final Year Cash Lump", cfg.get("contrib_cash_final_lump", 120_000)),
        ("", ""),
        ("PERFORMANCE", ""),
        ("PreTax Return", cfg["pretax_return"]),
        ("Roth Return", cfg["roth_return"]),
        ("HSA Return", cfg["hsa_return"]),
        ("Cash Return", cfg["cash_return"]),
        ("", ""),
        ("INCOME SOURCES", ""),
        ("SS Start Age", cfg["ss_start_age"]),
        ("SS Annual (future $)", cfg["ss_annual_amount"]),
        ("SS COLA", cfg["ss_cola"]),
        ("JSS Start Age", cfg["jss_start_age"]),
        ("JSS Annual", cfg["jss_annual_amount"]),
        ("JSS COLA", cfg["jss_cola"]),
        ("Rental Gross", cfg["rental_gross"]),
        ("", ""),
        ("EXPENSES & INFLATION", ""),
        ("Annual Expenses", cfg["base_annual_expenses"]),
        ("General Inflation", cfg["inflation_rate"]),
        ("Healthcare Pre-Medicare", cfg["healthcare_pre_medicare"]),
        ("Healthcare Post-Medicare", cfg["healthcare_post_medicare"]),
        ("Healthcare Inflation", cfg["healthcare_inflation"]),
        ("", ""),
        ("STRATEGY", ""),
        ("RMD Start Age", cfg["rmd_start_age"]),
        ("Roth Conversion Target", cfg["roth_conversion_target_bracket"]),
        ("IRMAA Avoidance", cfg["irmaa_avoidance"]),
        ("", ""),
        ("LEGACY POOL & HEIRS (v7)", ""),
        ("Legacy Pool Target Return", cfg.get("legacy_pool_return", cfg["roth_return"])),
        ("Legacy Pool Return Std Dev (MC)", cfg.get("legacy_pool_std", "")),
        ("Assumed Heir Tax Rate (PreTax/HSA)", cfg.get("heir_tax_rate", 0.24)),
        ("", ""),
        ("TAX (v7)", ""),
        ("Tax Cash/MM Interest Annually", cfg.get("tax_cash_interest", True)),
        ("", ""),
        ("SPENDING STRATEGY (v7)", ""),
        ("Strategy", cfg.get("spending_strategy", "fixed")),
        ("Guardrail Band (\u00b1%)", cfg.get("guardrail_band_pct", "")),
        ("Guardrail Spending Adjustment (%)", cfg.get("guardrail_adjustment_pct", "")),
        ("", ""),
        ("SURVIVING SPOUSE SCENARIO (v7)", ""),
        ("Model Surviving Spouse Scenario", cfg.get("model_widow_scenario", False)),
        ("Age of First Death", cfg.get("first_death_age", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("SS Survivor Benefit Retained", cfg.get("ss_survivor_pct", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("Pension Survivor Benefit Retained", cfg.get("jss_survivor_pct", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("Widowhood Expense Reduction", cfg.get("expense_reduction_widowhood", "") if cfg.get("model_widow_scenario") else "n/a"),
    ]
    for r, (label, val) in enumerate(assumptions, 1):
        ws_a.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_a.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1:
            cell.number_format = pct_fmt
            cell.font = input_font
        elif isinstance(val, (int, float)) and val >= 1000:
            cell.number_format = money_k
            cell.font = input_font
        else:
            cell.font = input_font
    ws_a.column_dimensions["A"].width = 30
    ws_a.column_dimensions["B"].width = 18

    # ──────── Sheet 2: Accumulation Phase ────────
    if accum_df is not None and len(accum_df) > 0:
        ws_acc = wb.create_sheet("Accumulation")
        acc_cols = ["Age", "Year", "Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash",
                    "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "S_Plus_5yr", "S_Plus_10yr",
                    "Total_Liquid_Assets"]
        for c, col in enumerate(acc_cols, 1):
            ws_acc.cell(row=1, column=c, value=col.replace("_", " "))
        style_header(ws_acc, len(acc_cols))
        for r, row in enumerate(accum_df.to_dict("records") if isinstance(accum_df, pd.DataFrame) else accum_df, 2):
            for c, col in enumerate(acc_cols, 1):
                val = row.get(col, "")
                cell = ws_acc.cell(row=r, column=c, value=val)
                cell.font = num_font
                if col not in ("Age", "Year", "Phase"):
                    cell.number_format = money_k
        auto_width(ws_acc)

    # ──────── Sheet 3: Retirement Phase ────────
    ws_ret = wb.create_sheet("Retirement")
    ret_cols = [
        "Age", "Year", "Draw_Strategy",
        "SS_Income", "JSS_Income", "Rental_Income", "S_Plus_Income", "Passive_Income",
        "PreTax_Draw", "Roth_Draw", "Cash_Draw", "HSA_Draw",
        "Total_Income", "Total_Expenses", "Total_Tax", "Surplus_Deficit",
        "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "Total_Liquid_Assets", "Total_Real",
        "Effective_Tax_Rate", "Withdrawal_Rate",
        "RMD", "Roth_Conversion", "Roth_Conversion_Tax",
        "MAGI", "IRMAA_Hit", "IRMAA_Cost",
        "Legacy_Target_Per_Child", "Legacy_Roth_Per_Child", "Legacy_Inheritance_Per_Child",
        "Legacy_Target_Total", "Legacy_Roth_Total", "Legacy_Inheritance_Total", "Legacy_Total",
        "Bad_Return_Year", "Cum_Gifts", "Cum_Legacy_Roth", "Cum_Legacy_Inheritance", "Cum_Lump_Sums",
        "Legacy_Pool_EOY", "Legacy_Pool_Per_Child", "Cum_Legacy_Inheritance_Per_Child",
        "Legacy_Value_To_Date", "Legacy_Value_To_Date_Per_Child",
        "Family_Net_Worth", "Estate_At_Death", "Estate_At_Death_Per_Child", "Total_Inheritance_At_Death_Per_Child",
    ]
    for c, col in enumerate(ret_cols, 1):
        ws_ret.cell(row=1, column=c, value=col.replace("_", " "))
    style_header(ws_ret, len(ret_cols))

    pct_columns = {"Effective_Tax_Rate", "Withdrawal_Rate"}
    bool_columns = {"IRMAA_Hit"}
    text_columns = {"Draw_Strategy"}
    neg_fill = PatternFill("solid", fgColor="FFE0E0")

    for r, (_, row) in enumerate(retire_df.iterrows(), 2):
        for c, col in enumerate(ret_cols, 1):
            val = row.get(col, "")
            cell = ws_ret.cell(row=r, column=c, value=val)
            cell.font = num_font
            cell.border = thin_border
            if col in pct_columns:
                cell.number_format = pct_fmt
            elif col in bool_columns:
                pass
            elif col in text_columns or col in ("Age", "Year"):
                pass
            else:
                cell.number_format = money_k
            # Red highlight for negative values
            if isinstance(val, (int, float)) and val < 0:
                cell.fill = neg_fill
    auto_width(ws_ret)

    # ──────── Sheet 4: Summary ────────
    ws_s = wb.create_sheet("Summary")
    first, last = retire_df.iloc[0], retire_df.iloc[-1]
    n_kids_xl = int(cfg.get("num_children", 0))
    summary = [
        ("RETIREMENT PLAN SUMMARY", ""),
        ("", ""),
        ("Starting Retirement Assets", first["Total_Liquid_Assets"]),
        ("  PreTax 401(k)", first["PreTax_EOY"]),
        ("  Roth IRA", first["Roth_EOY"]),
        ("  HSA", first["HSA_EOY"]),
        ("  Cash", first["Cash_EOY"]),
        ("", ""),
        (f"Your Own Estate at Age {cfg['planning_end_age']}", last["Total_Liquid_Assets"]),
        ("Real Value (inflation-adjusted)", last["Total_Real"]),
        ("", ""),
        ("Avg Withdrawal Rate", retire_df["Withdrawal_Rate"].mean()),
        ("Avg Effective Tax Rate", retire_df["Effective_Tax_Rate"].mean()),
        ("Total Roth Conversions", retire_df["Roth_Conversion"].sum()),
        ("Total Conversion Tax Paid", retire_df["Roth_Conversion_Tax"].sum()),
        ("", ""),
        ("LEGACY GIFTING", ""),
        ("Cumulative Annual Gifts (cash, separate from Roth legacy)", last.get("Cum_Gifts", 0.0)),
        ("Total Gifted to Kids' Roth (nominal, no growth)", last.get("Cum_Legacy_Roth", 0.0)),
        (f"Legacy Pool Today (grown, age {cfg['planning_end_age']})", last.get("Legacy_Pool_EOY", 0.0)),
        ("Cum Legacy Inheritance (bad-years-only sub-total, NOT your estate)", last.get("Cum_Legacy_Inheritance", 0.0)),
        ("Cumulative Lump Sums", last.get("Cum_Lump_Sums", 0.0)),
        ("", ""),
        ("ESTATE AT DEATH (if it occurred at the final planning age)", ""),
        ("Your Own Estate (passes to heirs)", last.get("Estate_At_Death", last["Total_Liquid_Assets"])),
        ("Already Gifted (Kids' own Roth, grown)", last.get("Legacy_Pool_EOY", 0.0)),
        ("Total to Family (both combined)", last.get("Family_Net_Worth", last["Total_Liquid_Assets"])),
    ]
    if n_kids_xl > 0:
        summary += [
            ("", ""),
            (f"PER CHILD (\u00f7 {n_kids_xl})", ""),
            ("  Estate / Child", last.get("Estate_At_Death_Per_Child", 0.0)),
            ("  Roth Pool / Child", last.get("Legacy_Pool_Per_Child", 0.0)),
            ("  Total / Child", last.get("Total_Inheritance_At_Death_Per_Child", 0.0)),
        ]
    summary += [
        ("", ""),
        ("IRMAA Years Hit", int(retire_df["IRMAA_Hit"].sum())),
    ]
    for r, (label, val) in enumerate(summary, 1):
        ws_s.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_s.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1 and val > 0:
            cell.number_format = pct_fmt
        elif isinstance(val, (int, float)) and abs(val) >= 1000:
            cell.number_format = money_k
        cell.font = num_font
    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 20

    # ──────── Sheet 5: Monte Carlo (only if MC was run) ────────
    if mc_runs:
        ws_mc = wb.create_sheet("Monte Carlo")
        surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
        mc_header = [
            ("MONTE CARLO SUMMARY", ""),
            ("Simulations", len(mc_runs)),
            ("Survived (Total_Liquid_Assets > 0 at final age)", surv),
            ("Success Rate", surv / len(mc_runs)),
            ("", ""),
        ]
        for r, (label, val) in enumerate(mc_header, 1):
            ws_mc.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
            cell = ws_mc.cell(row=r, column=2, value=val)
            if isinstance(val, float) and 0 < val < 1:
                cell.number_format = pct_fmt
            cell.font = num_font

        start_row = len(mc_header) + 2
        bands_nom = compute_percentile_bands(mc_runs, "Total_Liquid_Assets")
        bands_real = compute_percentile_bands(mc_runs, "Total_Real")
        mc_cols = ["Age", "P5 (Nominal)", "P25 (Nominal)", "P50 Median (Nominal)", "P75 (Nominal)", "P95 (Nominal)",
                   "P5 (Today's $)", "P25 (Today's $)", "P50 Median (Today's $)", "P75 (Today's $)", "P95 (Today's $)"]
        for c, col in enumerate(mc_cols, 1):
            ws_mc.cell(row=start_row, column=c, value=col)
        style_header_row = ws_mc[start_row]
        for cell in style_header_row:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ages = bands_nom["Age"]
        for i, age_val in enumerate(ages):
            row_vals = [int(age_val),
                        bands_nom["p5"][i], bands_nom["p25"][i], bands_nom["p50"][i], bands_nom["p75"][i], bands_nom["p95"][i],
                        bands_real["p5"][i], bands_real["p25"][i], bands_real["p50"][i], bands_real["p75"][i], bands_real["p95"][i]]
            for c, val in enumerate(row_vals, 1):
                cell = ws_mc.cell(row=start_row + 1 + i, column=c, value=val)
                cell.font = num_font
                cell.number_format = money_k if c > 1 else "0"
        auto_width(ws_mc)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# STREAMLIT UI
# ============================================================



def qp_int(name, default):
    try:
        raw = str(st.query_params.get(name, default)).replace(',', '').replace('$', '').strip()
        return int(float(raw))
    except (TypeError, ValueError):
        return default

def qp_float(name, default):
    try:
        return float(str(st.query_params.get(name, default)).replace('%', '').strip())
    except:
        return default

def get_balance_source(file_balances):
    qp = st.query_params
    web_keys = ["pretax", "roth", "hsa", "s5", "s10", "cash"]
    if any(str(qp.get(k, "")).strip() for k in web_keys):
        return "web fetch from Excel"
    if any(v is not None for v in file_balances.values()):
        return "using Starting_balances.txt"
    return "using defaults"

def main():
    st.set_page_config(page_title="Retirement Income Planner", page_icon="\U0001F4CA",
                       layout="wide", initial_sidebar_state="expanded")
    st.title("\U0001F3E6 Retirement Income & Tax Planning Simulator")
    st.caption("Accumulation + Retirement modeling with Monte Carlo and Excel export")

    with st.sidebar:
        st.header("\u2699\ufe0f Configuration")

        with st.expander("\U0001F464 Age & Timeline", expanded=True):
            current_age = st.number_input("Current Age", 45, 70, 55)
            retirement_age = st.slider("Retirement Age", 55, 63, 55)
            planning_end = st.slider("Plan Through Age", 85, 100, 89)

        with st.expander("\U0001F5A4 Surviving Spouse Scenario"):
            st.caption("Models the 'widow's penalty': filing status switches from MFJ to Single the year after the first spouse's death, with smaller brackets/deduction, a partial Social Security survivor benefit, and (optionally) reduced pension and living costs.")
            widow_scenario = st.checkbox("Model Surviving Spouse Scenario", value=False)
            if widow_scenario:
                first_death_age = st.slider("Age of First Spouse's Death", 55, 100, 85)
                ss_survivor_pct = st.slider("SS Survivor Benefit Retained (%)", 30, 100, 65) / 100
                jss_survivor_pct = st.slider("Pension Survivor Benefit Retained (%)", 0, 100, 100) / 100
                exp_red_widow = st.slider("Living Expense Reduction After Widowhood (%)", 0, 50, 20) / 100
            else:
                first_death_age, ss_survivor_pct, jss_survivor_pct, exp_red_widow = 999, 0.65, 1.0, 0.0

        with st.expander("\U0001F4B0 Current Balances (Today)", expanded=True):
            file_balances, file_loaded = load_starting_balances()
            balance_source = get_balance_source(file_balances)
            if balance_source == "web fetch from Excel":
                st.success("\U0001F4CA Using balances passed in via URL parameters (Excel launch)")
            elif balance_source == "using Starting_balances.txt":
                st.success("Loaded starting_balances.txt")
            else:
                st.info("No starting_balances.txt or URL parameters found; using defaults")
            balance_specs = [
                ("pretax_input", "pretax", file_balances["401k"] if file_balances["401k"] is not None else 1_475_000),
                ("roth_input", "roth", file_balances["roth"] if file_balances["roth"] is not None else 510_000),
                ("hsa_input", "hsa", file_balances["hsa"] if file_balances["hsa"] is not None else 130_000),
                ("s5_input", "s5", file_balances["s_plus_5"] if file_balances["s_plus_5"] is not None else 300_000),
                ("s10_input", "s10", file_balances["s_plus_10"] if file_balances["s_plus_10"] is not None else 400_000),
                ("cash_input", "cash", file_balances["cash"] if file_balances["cash"] is not None else 745_000),
            ]
            for widget_key, param_name, fallback in balance_specs:
                if param_name in st.query_params:
                    st.session_state[widget_key] = qp_int(param_name, fallback)
                elif widget_key not in st.session_state:
                    st.session_state[widget_key] = fallback

            pretax = st.number_input("PreTax 401(k) ($)", 0, 10_000_000, step=50_000, format="%d", key="pretax_input")
            roth_bal = st.number_input("Roth IRA ($)", 0, 5_000_000, step=25_000, format="%d", key="roth_input")
            hsa_bal = st.number_input("HSA ($)", 0, 500_000, step=10_000, format="%d", key="hsa_input")
            s5_bal = st.number_input("S+ 5-Year Payout ($)", 0, 2_000_000, step=25_000, format="%d", key="s5_input")
            s10_bal = st.number_input("S+ 10-Year Payout ($)", 0, 2_000_000, step=25_000, format="%d", key="s10_input")
            cash_bal = st.number_input("Cash ($)", 0, 1_000_000, step=10_000, format="%d", key="cash_input")

        with st.expander("\U0001F4BC Working Years Contributions"):
            st.caption(f"Annual contributions for {max(0, retirement_age - current_age)} remaining working years")
            c_401k = st.number_input("401(k) Employee ($)", 0, 50_000, 24_500, step=500, format="%d")
            c_roth401k = st.number_input("Roth 401(k) ($)", 0, 50_000, 8_000, step=500, format="%d")
            c_roth_ira = st.number_input("Roth IRA per Spouse ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_hsa = st.number_input("HSA Family ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_mega = st.number_input("Mega Backdoor Roth ($)", 0, 50_000, 29_000, step=1_000, format="%d")
            c_match = st.number_input("Employer Match ($)", 0, 50_000, 18_000, step=1_000, format="%d")
            st.divider()
            c_cash = st.number_input("Annual Cash Savings ($)", 0, 500_000, 150_000, step=5_000, format="%d")
            c_cash_lump = st.number_input("Final Year Cash Lump ($)", 0, 500_000, 120_000, step=10_000, format="%d")

        with st.expander("\U0001F4C8 Performance Assumptions"):
            pretax_ret = st.slider("PreTax Target Return %", 0.0, 12.0, 6.0, 0.5) / 100
            roth_ret = st.slider("Roth Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            hsa_ret = st.slider("HSA Target Return %", 0.0, 12.0, 5.0, 0.5) / 100
            cash_ret = st.slider("Cash/MM Target Return %", 0.0, 8.0, 4.0, 0.25) / 100

        with st.expander("\U0001F3B2 Monte Carlo Settings"):
            mc_enabled = st.checkbox("Enable Monte Carlo", value=True)
            mc_sims = st.slider("Simulations", 50, 10000, 5000, 50)
            mc_std = st.slider("Return Std Dev %", 1.0, 25.0, 12.0, 0.5) / 100
            mc_max = st.slider("Max Upside Cap %", 5.0, 25.0, 18.0, 0.5) / 100
            mc_seed = st.number_input("Seed (0=random)", 0, 99999, 0)
            st.caption("PreTax/Roth/HSA/Legacy Pool move together each year (same market factor) instead of being drawn fully independently -- more realistic since they typically hold similar underlying investments.")
            mc_correlation = st.slider("Cross-Account Correlation", 0.0, 1.0, 0.85, 0.05,
                help="1.0 = all equity-like accounts move in lockstep each year. 0.0 = fully independent (old behavior). Cash/MM always gets a smaller fraction of this (money-market rates track policy rates, not stocks).")
            mc_fat_tailed = st.checkbox("Fat-Tailed Returns (Student-t)", value=False,
                help="Real market returns crash harder and more often than a bell curve predicts. Enabling this uses a Student-t distribution instead of Normal, producing more realistic tail risk (more frequent extreme years) at the same target return and std dev.")
            mc_t_df = st.slider("Fat-Tail Intensity (lower = fatter tails)", 3, 15, 5, 1, disabled=not mc_fat_tailed) if mc_fat_tailed else 5

        with st.expander("\U0001F4E5 Income Sources"):
            ss_age = st.slider("SS Start Age", 62, 70, 65)
            ss_amount = st.number_input("SS Annual (Future $)", 0, 200_000, 78_000, step=1_000, format="%d")
            ss_cola = st.slider("SS COLA %", 0.0, 5.0, 2.0, 0.25) / 100
            st.divider()
            jss_age = st.slider("JSS Pension Start Age", 60, 70, 60)
            jss_amount = st.number_input("JSS Annual ($)", 0, 100_000, 18_000, step=1_000, format="%d")
            jss_cola_pct = st.slider("JSS COLA %", 0.0, 3.0, 1.0, 0.25) / 100
            jss_recovery = st.number_input("JSS Recovery Years", 0, 10, 4)
            st.divider()
            rental = st.number_input("Rental Gross ($)", 0, 200_000, 44_000, step=1_000, format="%d")
            rental_tax_pct = st.slider("Rental Taxable %", 0, 100, 50) / 100

        with st.expander("\U0001F4B8 Expenses"):
            base_exp = st.number_input("Annual Expenses ($)", 50_000, 500_000, 168_000, step=5_000, format="%d")
            inflation = st.slider("General Inflation %", 0.0, 6.0, 2.87, 0.01) / 100
            st.subheader("Healthcare")
            hc_pre = st.number_input("Pre-Medicare ($)", 0, 100_000, 33_000, step=1_000, format="%d")
            hc_post = st.number_input("Post-Medicare ($)", 0, 50_000, 12_000, step=1_000, format="%d")
            hc_inflation = st.slider("HC Inflation %", 0.0, 10.0, 5.0, 0.5) / 100
            hdhp_cost = st.number_input("HDHP ($)", 0, 30_000, 12_000, step=1_000, format="%d")
            st.subheader("Post-80 Adjustment")
            exp_red_80 = st.slider("Expense Reduction After 80 (%)", 0, 25, 25, 5) / 100
            st.caption("Lifestyle slowdown: reduce base expenses after age 80.")
            st.subheader("Negative Return Adjustment")
            neg_ret_reduction = st.number_input("Discretionary Draw Reduction if Year Return < 0 ($)", 0, 200_000, 20_000, step=1000, format="%d")
            st.subheader("Spending Strategy")
            spending_strategy_choice = st.radio("Strategy", ["Fixed Real Spending", "Dynamic Guardrails (Guyton-Klinger)"], index=0)
            spending_strategy = "guardrails" if spending_strategy_choice.startswith("Dynamic") else "fixed"
            if spending_strategy == "guardrails":
                st.caption("Spending steps down permanently if your withdrawal rate drifts too high above your starting rate, and steps up if it drifts too low -- instead of a fixed inflation-adjusted amount every year.")
                guardrail_band_pct = st.slider("Guardrail Band (± % of starting withdrawal rate)", 5, 40, 20, 5) / 100
                guardrail_adjustment_pct = st.slider("Spending Adjustment When Breached (%)", 5, 25, 10, 5) / 100
            else:
                guardrail_band_pct, guardrail_adjustment_pct = 0.20, 0.10
            st.subheader("Lump Sum")
            lump_age = st.number_input("Lump at Age", 55, 95, 70)
            lump_amt = st.number_input("Lump Amount ($)", 0, 1_000_000, 0, step=10_000, format="%d")
            lump_sums = {lump_age: lump_amt} if lump_amt > 0 else {}

        with st.expander("\U0001F381 Gifts & Giving"):
            gifts = st.number_input("Annual Gifts ($)", 0, 100_000, 0, step=1_000, format="%d")

        with st.expander("\U0001F3AF Strategy", expanded=True):
            st.markdown("**PreTax→12% bracket, then Cash, then Roth. RMD dominates at 75+.**")
            rmd_start = st.slider("RMD Start Age", 73, 75, 75)
            perf_only = st.checkbox("Performance Draw Only", value=False)
            st.divider()
            roth_conv = st.checkbox("Enable Roth Conversions", value=False)
            roth_bracket = st.selectbox("Conversion Target", ["12%", "22%"])
            roth_margin = st.number_input("Min PreTax Keep ($)", 0, 1_000_000, 100_000, step=25_000, format="%d")
            st.divider()
            irmaa_avoid = st.checkbox("IRMAA Avoidance", value=True)

        with st.expander("\U0001F468\u200D\U0001F469\u200D\U0001F467\u200D\U0001F466 Legacy"):
            num_kids = st.number_input("Children", 0, 10, 4)
            roth_per_child = st.number_input("Roth/Child/Year ($)", 0, 50_000, 7_500, step=500, format="%d")
            legacy_years = st.slider("Legacy Duration (years)", 1, 20, 20)
            st.caption("The kids' Roth pool can use its own return/risk profile instead of automatically matching your own Roth account.")
            legacy_pool_ret = st.slider("Legacy Pool Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            legacy_pool_std_pct = st.slider("Legacy Pool Return Std Dev % (Monte Carlo)", 1.0, 25.0, 12.0, 0.5) / 100
            st.caption("Inherited pretax/HSA accounts must be distributed within 10 years (SECURE Act) and are taxed at the heir's own bracket -- unlike Roth, which passes tax-free. Used for the 'after-tax to heirs' figures on the Legacy tab.")
            heir_tax_rate = st.slider("Assumed Heir Tax Rate on PreTax/HSA (%)", 0, 45, 24, 1) / 100

        with st.expander("\U0001F3DB Tax"):
            or_resident = st.checkbox("Oregon Resident", value=True)
            std_ded = st.number_input("Std Deduction MFJ ($)", 0, 50_000, 32_500, step=100, format="%d")
            bracket_infl = st.slider("Bracket Inflation %", 0.0, 5.0, 2.67, 0.01) / 100
            tax_cash_int = st.checkbox("Tax Cash/MM Interest Annually (realistic)", value=True,
                help="Money-market interest is taxed as ordinary income every year, unlike unrealized gains in an equity account. Uncheck to model 'cash' as literal currency instead of an invested account.")

        with st.expander("\U0001F4CA Display Options"):
            show_real = st.toggle("Show in Today's Dollars", value=True)
            st.caption("When on, all dollar amounts are inflation-adjusted to present value.")

    # ── BUILD CONFIG ──
    cfg = dict(
        current_age=current_age, retirement_age=retirement_age, planning_end_age=planning_end,
        pretax_401k=pretax, roth_ira=roth_bal, hsa=hsa_bal,
        s_plus_5yr=s5_bal, s_plus_10yr=s10_bal, cash=cash_bal,
        contrib_401k=c_401k, contrib_roth401k=c_roth401k, contrib_roth_ira=c_roth_ira,
        contrib_hsa=c_hsa, contrib_mega_backdoor=c_mega, contrib_employer_match=c_match,
        contrib_cash_annual=c_cash, contrib_cash_final_lump=c_cash_lump,
        pretax_return=pretax_ret, roth_return=roth_ret, hsa_return=hsa_ret, cash_return=cash_ret,
        ss_start_age=ss_age, ss_annual_amount=ss_amount, ss_cola=ss_cola,
        jss_start_age=jss_age, jss_annual_amount=jss_amount, jss_cola=jss_cola_pct,
        jss_recovery_years=jss_recovery, rental_gross=rental, rental_taxable_pct=rental_tax_pct,
        base_annual_expenses=base_exp, inflation_rate=inflation,
        healthcare_pre_medicare=hc_pre, healthcare_post_medicare=hc_post,
        healthcare_inflation=hc_inflation, hdhp_annual=hdhp_cost,
        expense_reduction_post80=exp_red_80,
        gifts_annual=gifts, lump_sums=lump_sums,
        standard_deduction=std_ded, bracket_inflation=bracket_infl, oregon_resident=or_resident,
        tax_cash_interest=tax_cash_int,
        rmd_start_age=rmd_start, draw_order=["pretax", "cash", "roth", "hsa"],
        roth_conversion_enabled=roth_conv, roth_conversion_target_bracket=roth_bracket,
        roth_conversion_margin=roth_margin, irmaa_avoidance=irmaa_avoid,
        performance_draw_only=perf_only,
        neg_ret_draw_reduction=neg_ret_reduction,
        num_children=num_kids, roth_legacy_per_child=roth_per_child,
        legacy_years=legacy_years, legacy_inflation=inflation,
        legacy_pool_return=legacy_pool_ret, legacy_pool_std=legacy_pool_std_pct,
        heir_tax_rate=heir_tax_rate,
        spending_strategy=spending_strategy,
        guardrail_band_pct=guardrail_band_pct, guardrail_adjustment_pct=guardrail_adjustment_pct,
        model_widow_scenario=widow_scenario, first_death_age=first_death_age,
        ss_survivor_pct=ss_survivor_pct, jss_survivor_pct=jss_survivor_pct,
        expense_reduction_widowhood=exp_red_widow,
    )

    # ── RUN ──
    accum_rows, df = run_simulation(cfg)
    accum_df = pd.DataFrame(accum_rows) if accum_rows else None
    mc_runs = None
    if mc_enabled:
        mc_runs = run_monte_carlo(cfg, mc_sims, mc_std, mc_max, seed=mc_seed if mc_seed > 0 else None,
                                   correlation=mc_correlation, fat_tailed=mc_fat_tailed, t_df=mc_t_df)

    # ── DEFLATOR (today's dollars conversion) ──
    # When show_real is True, deflate future dollar amounts back to present value
    dollar_label = "Today's $" if show_real else "Future $"
    def deflate(series_or_val, year_col_or_yir):
        """Deflate future dollars to today's dollars."""
        if not show_real:
            return series_or_val
        if isinstance(year_col_or_yir, (pd.Series, np.ndarray)):
            factors = (1 + inflation) ** year_col_or_yir
        else:
            factors = (1 + inflation) ** year_col_or_yir
        return series_or_val / factors

    # Build display-ready retirement df
    df_disp = df.copy()
    if show_real:
        # Years from today (accumulation years + retirement years)
        accum_years = max(0, retirement_age - current_age)
        df_disp["_deflate_yrs"] = accum_years + df_disp["Years_Retired"]
        money_cols = [c for c in df_disp.columns if c not in (
            "Age", "Year", "Years_Retired", "Phase", "Draw_Strategy",
            "Effective_Tax_Rate", "Withdrawal_Rate", "PreTax_WR",
            "IRMAA_Hit", "Under_700_FPL", "In_12_Bracket", "In_22_Bracket",
            "Deduction_Type", "PreTax_Depleted", "Roth_Depleted", "HSA_Depleted",
            "Return_PreTax", "Return_Roth", "Return_HSA", "Return_Cash",
            "RMD_Excess", "_deflate_yrs")]
        for c in money_cols:
            if c in df_disp.columns and df_disp[c].dtype in [np.float64, np.int64, float, int]:
                df_disp[c] = df_disp[c] / ((1 + inflation) ** df_disp["_deflate_yrs"])
        df_disp.drop(columns=["_deflate_yrs"], inplace=True)

    # ── KPI ROW ──
    c1, c2, c3, c4, c5 = st.columns(5)
    first, last = df_disp.iloc[0], df_disp.iloc[-1]
    if accum_df is not None and len(accum_df) > 0:
        c1.metric(f"Today's Assets", f"${accum_df.iloc[0]['Total_Liquid_Assets']:,.0f}")
    else:
        c1.metric("Starting Assets", f"${first['Total_Liquid_Assets']:,.0f}")
    c2.metric(f"At Retirement ({dollar_label})", f"${first['Total_Liquid_Assets']:,.0f}")
    legacy_pool_last = last.get("Legacy_Pool_EOY", 0.0)
    c3.metric(f"Age {planning_end} ({dollar_label})", f"${last['Total_Liquid_Assets']:,.0f}",
              delta=(f"+${legacy_pool_last:,.0f} already gifted (see Legacy tab)" if legacy_pool_last > 0 else None),
              delta_color="off",
              help="Your own remaining accounts (pretax/Roth/HSA/cash) -- excludes money already gifted to the kids' Roth accounts, since that's no longer part of your own net worth. See the Legacy tab for the combined family total.")

    depleted = df[df["Total_Liquid_Assets"] <= 0]
    if len(depleted) > 0:
        c4.metric("\u26a0\ufe0f Depleted", f"Age {int(depleted.iloc[0]['Age'])}", delta="RISK", delta_color="inverse")
    else:
        c4.metric("\u2705 Lasts To", f"Age {planning_end}+")

    if mc_runs:
        surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
        c5.metric("MC Success", f"{surv/len(mc_runs)*100:.0f}%", delta=f"{len(mc_runs)} sims", delta_color="off")
    else:
        c5.metric("Avg WR", f"{df['Withdrawal_Rate'].mean():.1%}")

    st.divider()

    # ── TABS ──
    tab_names = ["\U0001F4BC Accumulation", "\U0001F4CA Retirement Assets", "\U0001F4B0 Income & Expenses",
                 "\U0001F381 Legacy", "\U0001F3DB Tax", "\U0001F504 Roth Conversions"]
    if mc_runs: tab_names.append("\U0001F3B2 Monte Carlo")
    if mc_runs: tab_names.append("\U0001F3AF Optimizer")
    tab_names.append("\U0001F32A Sensitivity")
    tab_names.append("\U0001F4CB Full Data")
    tabs = st.tabs(tab_names)
    ti = 0

    # ── Accumulation Tab ──
    with tabs[ti]:
        if accum_df is not None and len(accum_df) > 0:
            yrs = retirement_age - current_age
            total_contrib = accum_df[["Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash"]].sum().sum()
            ac1, ac2, ac3 = st.columns(3)
            ac1.metric("Working Years", f"{yrs}")
            ac2.metric("Total Contributions", f"${total_contrib:,.0f}")
            ac3.metric("Retirement Day Assets", f"${first['Total_Liquid_Assets']:,.0f}")

            fig = go.Figure()
            for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"),
                                     ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                     ("Roth_EOY","Roth","rgba(142,68,173,0.6)"),
                                     ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
                fig.add_trace(go.Scatter(x=accum_df["Age"], y=accum_df[col], mode="lines", name=name,
                                         stackgroup="one", line=dict(width=0.5), fillcolor=color))
            fig.update_layout(title="Accumulation Phase: Account Growth",
                              xaxis_title="Age", yaxis_title="$", yaxis_tickformat="$,.0f", height=450)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(accum_df, use_container_width=True, height=300)
        else:
            st.info("No accumulation phase (already at retirement age).")
    ti += 1

    # ── Retirement Assets Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"), ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                 ("Roth_EOY","Roth","rgba(142,68,173,0.6)"), ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
            fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp[col], mode="lines", name=name,
                                     stackgroup="one", line=dict(width=0.5), fillcolor=color))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Real"], mode="lines", name="Real Value",
                                 line=dict(color="black", width=2, dash="dash")))
        fig.update_layout(title="Retirement Account Balances", xaxis_title="Age", yaxis_title="$",
                          yaxis_tickformat="$,.0f", hovermode="x unified", height=500)
        st.plotly_chart(fig, use_container_width=True)

        colors = ["red" if x > 0.04 else "orange" if x > 0.03 else "green" for x in df["Withdrawal_Rate"]]
        fig2 = go.Figure(go.Bar(x=df["Age"], y=df["Withdrawal_Rate"]*100, marker_color=colors))
        fig2.add_hline(y=4, line_dash="dash", line_color="red", annotation_text="4% Rule")
        fig2.update_layout(title="Withdrawal Rate", xaxis_title="Age", yaxis_title="%", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Income & Expenses Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, nm in [("SS_Income","SS"),("JSS_Income","JSS"),("Rental_Income","Rental"),
                        ("S_Plus_Income","S+"),("PreTax_Draw","PreTax Draw"),("Roth_Draw","Roth Draw")]:
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp[col], name=nm))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Expenses"], mode="lines+markers",
                                 name="Expenses", line=dict(color="red", width=3)))
        fig.update_layout(barmode="stack", title="Income vs Expenses", xaxis_title="Age",
                          yaxis_title="$", yaxis_tickformat="$,.0f", height=500)
        st.plotly_chart(fig, use_container_width=True)

        sd_c = ["green" if v >= 0 else "red" for v in df_disp["Surplus_Deficit"]]
        fig2 = go.Figure(go.Bar(x=df_disp["Age"], y=df_disp["Surplus_Deficit"], marker_color=sd_c))
        fig2.update_layout(title="Surplus / Deficit", xaxis_title="Age", yaxis_title="$",
                           yaxis_tickformat="$,.0f", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Legacy Tab ──
    with tabs[ti]:
        st.subheader("Legacy Planning")
        lg1, lg2, lg3 = st.columns(3)
        lg1.metric("Children", f"{int(cfg.get('num_children', 0))}")
        lg2.metric("Legacy Years", f"{int(cfg.get('legacy_years', 0))}")
        if mc_runs:
            avg_bad_years = np.mean([r["Bad_Return_Year"].sum() for r in mc_runs])
            lg3.metric("Avg Bad Return Years (MC)", f"{avg_bad_years:.1f}")
        else:
            bad_ct = int(df['Bad_Return_Year'].sum()) if 'Bad_Return_Year' in df.columns else 0
            lg3.metric("Bad Return Years", f"{bad_ct}")
            st.caption(
                "The baseline plan uses one constant assumed return every year, so it "
                "never has a down year on its own (0 here is expected, not a bug). "
                "Enable Monte Carlo to see how often a bad year actually skips a legacy "
                "gift across randomized market paths."
            )

        n_kids = int(cfg.get("num_children", 0))

        st.markdown("---")
        st.markdown("##### If you died at a given age, what would your kids actually receive?")
        st.caption(
            "Two very different pots make up the answer: money **already gifted** into the "
            "kids' own Roth accounts (irrevocable, compounding on its own), and whatever's "
            "**still sitting in your own accounts** at death (pretax, your own Roth, HSA, "
            "cash) -- which is what actually passes to heirs. 'Cum Legacy Inheritance' further "
            "below is neither of these; see its note."
        )
        min_age, max_age = int(df_disp["Age"].min()), int(df_disp["Age"].max())
        death_age = st.slider("Hypothetical age at death", min_age, max_age, max_age, key="legacy_death_age")
        death_row = df_disp[df_disp["Age"] == death_age].iloc[0]

        d1, d2, d3 = st.columns(3)
        d1.metric("Already Gifted (Kids' Roth, grown)", f"${death_row.get('Legacy_Pool_EOY', 0.0):,.0f}",
                   help="Money already moved into the kids' own Roth accounts by this age. Already theirs, not part of your estate.")
        d2.metric("Remaining in Your Estate", f"${death_row.get('Estate_At_Death', death_row['Total_Liquid_Assets']):,.0f}",
                   help="Whatever's left in your own pretax/Roth/HSA/cash accounts at this age -- this is what passes to heirs at death.")
        d3.metric("Total to Family", f"${death_row.get('Family_Net_Worth', 0.0):,.0f}",
                   help="The two combined.")

        if n_kids > 0:
            st.markdown(f"**Per child (÷ {n_kids}):**")
            e1, e2, e3 = st.columns(3)
            e1.metric("Roth Pool / Child", f"${death_row.get('Legacy_Pool_Per_Child', 0.0):,.0f}")
            e2.metric("Estate / Child", f"${death_row.get('Estate_At_Death_Per_Child', 0.0):,.0f}")
            e3.metric("Total / Child", f"${death_row.get('Total_Inheritance_At_Death_Per_Child', 0.0):,.0f}")
        else:
            st.caption("Set Number of Children > 0 to see a per-child split.")

        fig_estate = go.Figure()
        fig_estate.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Estate_At_Death"], name="Remaining Estate (your own accounts)",
                                        line=dict(color="teal", width=0.5), stackgroup="one", fillcolor="rgba(0,128,128,0.4)"))
        fig_estate.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_EOY"], name="Already Gifted (Kids' Roth)",
                                        line=dict(color="purple", width=0.5), stackgroup="one", fillcolor="rgba(128,0,128,0.4)"))
        fig_estate.add_vline(x=death_age, line_dash="dash", line_color="red")
        fig_estate.update_layout(title="What Would Pass to Family, by Age of Death", xaxis_title="Age",
                                  yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_estate, use_container_width=True)

        st.markdown("---")
        st.markdown("##### Legacy Gifting Detail")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Gifted (nominal)", f"${df_disp.iloc[-1]['Cum_Legacy_Roth']:,.0f}",
                   help="Sum of contribution amounts, at the time each was made -- does not include growth.")
        c2.metric("Legacy Pool Today", f"${df_disp.iloc[-1].get('Legacy_Pool_EOY', 0.0):,.0f}",
                   help="What that gifted money has actually grown to, compounding on its own in the kids' Roth accounts. Same number as 'Already Gifted' above.")
        c3.metric("Cum Legacy Inheritance (bad-years only)", f"${df_disp.iloc[-1].get('Cum_Legacy_Inheritance', 0.0):,.0f}",
                   help="NOT your general estate. This is only the sub-total of gift TARGETS that were skipped in down-market years and left invested rather than gifted to Roth. It's $0 whenever the plan never hits a down year (see note above) -- for what actually passes to your heirs at any age, use the 'Remaining in Your Estate' section above instead.")
        c4.metric("Total Legacy Value (Today)", f"${df_disp.iloc[-1].get('Legacy_Value_To_Date', 0.0):,.0f}",
                   help="Legacy Pool + Cum Legacy Inheritance (the bad-years sub-total, not your full estate).")

        # ── Legacy Pool balance over time (the actual compounding balance,
        # not just the annual flow into it) ──
        fig_leg_bal = go.Figure()
        fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_EOY"],
                                          name="Legacy Pool (grown)", line=dict(color="purple", width=3),
                                          fill="tozeroy", fillcolor="rgba(128,0,128,0.1)"))
        fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Cum_Legacy_Roth"],
                                          name="Cumulative Gifted (nominal, no growth)",
                                          line=dict(color="gray", width=2, dash="dot")))
        if 'Cum_Legacy_Inheritance' in df_disp.columns and df_disp['Cum_Legacy_Inheritance'].iloc[-1] > 0:
            fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Value_To_Date"],
                                              name="Total Legacy Value (Pool + Inheritance)",
                                              line=dict(color="darkgoldenrod", width=2, dash="dash")))
        fig_leg_bal.update_layout(title="Legacy Pool: Balance vs. Nominal Contributions", xaxis_title="Age",
                                  yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_bal, use_container_width=True)
        st.caption(
            "The gap between the purple and gray lines is investment growth on money already "
            "gifted. This balance is segregated from the household's own accounts -- it "
            "compounds on its own and is never drawn back down to cover the household's "
            "living expenses."
        )

        if n_kids > 0:
            fig_leg_pc_bal = go.Figure()
            fig_leg_pc_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_Per_Child"],
                                                 name="Roth Pool / Child (grown)", line=dict(color="purple", width=2.5)))
            fig_leg_pc_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Cum_Legacy_Inheritance_Per_Child"],
                                                 name="Inheritance / Child", line=dict(color="goldenrod", width=2.5)))
            fig_leg_pc_bal.update_layout(title="Per-Child Legacy Value Over Time", xaxis_title="Age",
                                         yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=380)
            st.plotly_chart(fig_leg_pc_bal, use_container_width=True)

        fig_leg_pc = go.Figure()
        fig_leg_pc.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Roth_Per_Child"], name="Roth per Child", marker_color="purple"))
        fig_leg_pc.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Inheritance_Per_Child"], name="Inheritance per Child", marker_color="goldenrod"))
        fig_leg_pc.update_layout(barmode="group", title="Annual Legacy Gift per Child by Age", xaxis_title="Age",
                                 yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_pc, use_container_width=True)

        fig_leg_total = go.Figure()
        fig_leg_total.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Roth_Total"], name="Roth Total", marker_color="mediumpurple"))
        fig_leg_total.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Inheritance_Total"], name="Inheritance Total", marker_color="darkgoldenrod"))
        fig_leg_total.update_layout(barmode="stack", title="Annual Legacy Gift Total by Age", xaxis_title="Age",
                                    yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_total, use_container_width=True)

        legacy_cols = [c for c in [
            "Age", "Year", "Bad_Return_Year",
            "Estate_At_Death", "Estate_At_Death_Per_Child",
            "Legacy_Pool_EOY", "Legacy_Pool_Per_Child",
            "Family_Net_Worth", "Total_Inheritance_At_Death_Per_Child",
            "Legacy_Target_Per_Child", "Legacy_Roth_Per_Child", "Legacy_Inheritance_Per_Child",
            "Legacy_Target_Total", "Legacy_Roth_Total", "Legacy_Inheritance_Total", "Legacy_Total",
            "Cum_Legacy_Roth", "Cum_Legacy_Inheritance", "Cum_Legacy_Inheritance_Per_Child",
            "Legacy_Value_To_Date", "Legacy_Value_To_Date_Per_Child",
        ] if c in df_disp.columns]
        st.dataframe(df_disp[legacy_cols], use_container_width=True, hide_index=True)
    ti += 1

    # ── Tax Tab ──
    with tabs[ti]:
        ca2, cb2 = st.columns(2)
        with ca2:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Federal_Tax"], name="Federal"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Oregon_Tax"], name="Oregon"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["IRMAA_Cost"], name="IRMAA"))
            fig.update_layout(barmode="stack", title="Tax Burden", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
        with cb2:
            fig = go.Figure(go.Scatter(x=df["Age"], y=df["Effective_Tax_Rate"]*100,
                                       mode="lines+markers", line=dict(color="darkblue", width=2)))
            fig.update_layout(title="Effective Tax Rate", yaxis_title="%", height=400)
            st.plotly_chart(fig, use_container_width=True)

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Federal_Taxable_Income"], name="Taxable", line=dict(color="blue", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_12_Ceiling"], name="12% Ceil", line=dict(color="green", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_22_Ceiling"], name="22% Ceil", line=dict(color="orange", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["MAGI"], name="MAGI", line=dict(color="red", width=1)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["IRMAA_Threshold"], name="IRMAA Thr", line=dict(color="darkred", dash="dot")))
        fig.update_layout(title="Income vs Brackets & IRMAA", yaxis_tickformat="$,.0f", height=450)
        st.plotly_chart(fig, use_container_width=True)
    ti += 1

    # ── Roth Conversions Tab ──
    with tabs[ti]:
        cdf = df[df["Roth_Conversion"] > 0]
        if len(cdf) > 0:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion"], name="Converted", marker_color="purple"))
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion_Tax"], name="Tax", marker_color="red"))
            fig.update_layout(barmode="group", title="Roth Conversions", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
            tc, tt = cdf["Roth_Conversion"].sum(), cdf["Roth_Conversion_Tax"].sum()
            if tc > 0: st.info(f"**Total:** ${tc:,.0f} | **Tax:** ${tt:,.0f} | **Rate:** {tt/tc:.1%}")
        else:
            st.info("No conversions.")
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["PreTax_EOY"], name="PreTax", line=dict(color="red", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Roth_EOY"], name="Roth", line=dict(color="purple", width=2)))
        fig.update_layout(title="PreTax vs Roth", yaxis_tickformat="$,.0f", height=400)
        st.plotly_chart(fig, use_container_width=True)
    ti += 1

    # ── Monte Carlo Tab ──
    if mc_runs:
        with tabs[ti]:
            surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
            m1, m2, m3 = st.columns(3)
            m1.metric("Sims", len(mc_runs)); m2.metric("Survived", surv); m3.metric("Rate", f"{surv/len(mc_runs)*100:.1f}%")

            bands = compute_percentile_bands(mc_runs, "Total_Liquid_Assets")
            ages = bands["Age"]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p95"],bands["p5"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.1)", line=dict(color="rgba(255,255,255,0)"), name="5-95%"))
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p75"],bands["p25"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.25)", line=dict(color="rgba(255,255,255,0)"), name="25-75%"))
            fig.add_trace(go.Scatter(x=ages, y=bands["p50"], name="Median", line=dict(color="blue", width=2)))
            fig.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                     line=dict(color="black", width=2, dash="dash")))
            fig.update_layout(title=f"MC Total Assets ({len(mc_runs)} runs)", yaxis_tickformat="$,.0f", height=500)
            st.plotly_chart(fig, use_container_width=True)

            n_show = min(20, len(mc_runs))
            fig3 = go.Figure()
            for i in range(n_show):
                fig3.add_trace(go.Scatter(x=mc_runs[i]["Age"], y=mc_runs[i]["Total_Liquid_Assets"],
                    mode="lines", line=dict(width=0.7, color="rgba(99,110,250,0.3)"), showlegend=False))
            fig3.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                      line=dict(color="black", width=2, dash="dash")))
            fig3.update_layout(title=f"{n_show} Sample Paths", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig3, use_container_width=True)
        ti += 1

    # ── Optimizer Tab ──
    if mc_runs:
        with tabs[ti]:
            st.subheader("\U0001F3AF Scenario Optimizer — Path to 100% Success")
            st.caption(
                "Sweeps retirement age delay, SS start age, and post-80 expense reduction "
                "to find combinations that maximize Monte Carlo success rate."
            )

            if st.button("Run Optimizer", type="primary"):
                with st.spinner("Running scenario sweep (this takes a moment)..."):
                    opt_df = run_optimizer(cfg, mc_sims, mc_std, mc_max, mc_seed,
                                            correlation=mc_correlation, fat_tailed=mc_fat_tailed, t_df=mc_t_df)

                # ── Current scenario baseline ──
                surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
                base_rate = surv / len(mc_runs) * 100
                st.info(f"**Current settings:** Retire {cfg['retirement_age']}, "
                        f"SS @ {cfg['ss_start_age']}, "
                        f"No post-80 cut → **{base_rate:.0f}% success**")

                # ── Top recommendations ──
                perfect = opt_df[opt_df["Success_Rate"] >= 1.0]
                if len(perfect) > 0:
                    st.success(f"\u2705 Found **{len(perfect)}** scenarios with 100% success rate!")
                    st.markdown("**Top 5 — least disruption to reach 100%:**")
                    # Sort by least change needed
                    perfect_sorted = perfect.sort_values(
                        ["Extra_Work_Years", "SS_Delay_Years", "Expense_Cut_80+"]
                    ).head(5)
                    for _, row in perfect_sorted.iterrows():
                        changes = []
                        if row["Extra_Work_Years"] > 0:
                            changes.append(f"retire at **{int(row['Retire_Age'])}** (+{int(row['Extra_Work_Years'])}yr)")
                        if row["SS_Delay_Years"] > 0:
                            changes.append(f"SS at **{int(row['SS_Age'])}** (+{int(row['SS_Delay_Years'])}yr)")
                        if row["Expense_Cut_80+"] > 0:
                            changes.append(f"**{row['Expense_Cut_80+']*100:.0f}%** expense cut after 80")
                        med_val = row["Median_Final_Real"] if show_real else row["Median_Final_Assets"]
                        label = "today's $" if show_real else "future $"
                        desc = " + ".join(changes) if changes else "No changes needed!"
                        st.markdown(f"- {desc} → Median final: **${med_val:,.0f}** ({label})")
                else:
                    st.warning("No 100% success scenarios found. Best options below.")

                # ── Full results table ──
                st.subheader("All Scenarios")
                display_opt = opt_df.copy()
                display_opt["Success_Rate"] = display_opt["Success_Rate"].apply(lambda x: f"{x:.0%}")
                display_opt["Expense_Cut_80+"] = display_opt["Expense_Cut_80+"].apply(lambda x: f"{x:.0%}")
                money_opt_cols = ["Median_Final_Assets", "Median_Final_Real", "P10_Final_Assets", "Worst_Case"]
                for c in money_opt_cols:
                    display_opt[c] = display_opt[c].apply(lambda x: f"${x:,.0f}")
                st.dataframe(display_opt, use_container_width=True, height=400)
            else:
                st.markdown(
                    "Click **Run Optimizer** to sweep combinations of:\n"
                    "- Retirement age: current → +3 years\n"
                    "- SS start age: current → +4 years\n"
                    "- Post-80 expense reduction: 0% to 25%\n\n"
                    "Uses 50 MC sims per scenario for speed."
                )
        ti += 1

    # ── Sensitivity Tab ──
    with tabs[ti]:
        st.subheader("\U0001F32A Sensitivity (Tornado) Analysis")
        st.caption(
            "Varies ONE assumption at a time -- holding everything else at your current "
            "settings -- and shows how much your final-year value (today's $) moves. Sorted "
            "so the biggest driver of your outcome is at the top."
        )
        if st.button("Run Sensitivity Analysis", type="primary"):
            with st.spinner("Running sensitivity sweep..."):
                sens_df, base_final = run_sensitivity_analysis(cfg)

            fig = go.Figure()
            fig.add_trace(go.Bar(
                y=sens_df["Variable"], x=sens_df["High_Result"] - sens_df["Low_Result"],
                base=sens_df["Low_Result"], orientation="h",
                marker_color="rgba(99,110,250,0.6)",
                text=[f"${lo:,.0f} \u2192 ${hi:,.0f}" for lo, hi in zip(sens_df["Low_Result"], sens_df["High_Result"])],
                textposition="inside", insidetextanchor="middle",
            ))
            fig.add_vline(x=base_final, line_dash="dash", line_color="black")
            fig.update_layout(
                title=f"Impact on Final-Year Value ({dollar_label}) \u2014 Baseline: ${base_final:,.0f}",
                xaxis_title=dollar_label, xaxis_tickformat="$,.0f",
                height=120 + 45 * len(sens_df), margin=dict(l=10),
            )
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("**Range tested per variable** (\u00b1 around your current setting):")
            disp_sens = sens_df[["Variable", "Low_Value", "Base_Value", "High_Value",
                                  "Low_Result", "High_Result", "Impact_Range"]].copy()
            for c in ["Low_Result", "High_Result", "Impact_Range"]:
                disp_sens[c] = disp_sens[c].apply(lambda x: f"${x:,.0f}")
            for c in ["Low_Value", "Base_Value", "High_Value"]:
                disp_sens[c] = disp_sens[c].apply(
                    lambda x: f"{x:.1%}" if isinstance(x, float) and abs(x) < 1 else f"{x:,.0f}"
                )
            st.dataframe(disp_sens.iloc[::-1], use_container_width=True, hide_index=True)
            st.caption(
                "Note: Retirement Age and Base Annual Expenses change the shape of the whole "
                "plan (years worked, years drawing down), so their bars aren't perfectly "
                "apples-to-apples with pure return/rate assumptions -- but the ranking still "
                "tells you where a planning decision matters most."
            )
        else:
            st.info("Click **Run Sensitivity Analysis** to see which assumptions matter most to your outcome.")
        ti += 1

    # ── Full Data Tab ──
    with tabs[ti]:
        st.caption(f"Displaying in **{dollar_label}**")
        dcols = ["Age","Year","Draw_Strategy",
                 "SS_Income","JSS_Income","Rental_Income","S_Plus_Income","Passive_Income",
                 "PreTax_Draw","Roth_Draw","Cash_Draw","HSA_Draw",
                 "PreTax_EOY","Roth_EOY","HSA_EOY","Cash_EOY",
                 "Total_Liquid_Assets","Family_Net_Worth","Total_Income","Total_Expenses","Total_Tax",
                 "Effective_Tax_Rate","Withdrawal_Rate","Surplus_Deficit",
                 "Roth_Conversion","RMD","RMD_Excess","IRMAA_Hit","Bad_Return_Year",
                 "Legacy_Target_Per_Child","Legacy_Roth_Per_Child","Legacy_Inheritance_Per_Child",
                 "Legacy_Target_Total","Legacy_Roth_Total","Legacy_Inheritance_Total","Legacy_Total",
                 "Cum_Gifts","Cum_Legacy_Roth","Legacy_Pool_EOY","Cum_Legacy_Inheritance","Cum_Lump_Sums",
                 "Estate_At_Death"]
        avail = [c for c in dcols if c in df_disp.columns]
        show = st.multiselect("Columns", df_disp.columns.tolist(), default=avail)
        disp = df_disp[show].copy()
        no_fmt = {"Age","Year","Years_Retired","Effective_Tax_Rate","Withdrawal_Rate","PreTax_WR",
                  "IRMAA_Hit","Bad_Return_Year","Under_700_FPL","In_12_Bracket","In_22_Bracket","Draw_Strategy",
                  "PreTax_Depleted","Roth_Depleted","HSA_Depleted","Return_PreTax","Return_Roth","Return_HSA","Return_Cash","Phase"}
        pct = {"Effective_Tax_Rate","Withdrawal_Rate","PreTax_WR","Return_PreTax","Return_Roth","Return_HSA","Return_Cash"}
        fmt = {}
        for c in show:
            if c in pct: fmt[c] = "{:.1%}"
            elif c not in no_fmt: fmt[c] = "${:,.0f}"
        st.dataframe(disp.style.format(fmt, na_rep="").apply(
            lambda x: ["background-color:#ffcccc" if isinstance(v,(int,float)) and v<0 else "" for v in x], axis=1),
            use_container_width=True, height=600)

        # ── DOWNLOADS ──
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("\U0001F4E5 Download CSV", df.to_csv(index=False), "retirement_plan.csv", "text/csv")
        with col_dl2:
            xlsx_buf = export_to_excel(accum_df, df, cfg, mc_runs=mc_runs)
            st.download_button("\U0001F4E5 Download Excel Report", xlsx_buf,
                               "retirement_plan.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Notes ──
    st.divider()
    with st.expander("Model Notes"):
        st.markdown("""
**Accumulation Phase:** Balances grow from current age to retirement using configured returns. Contributions include 401(k), Roth 401(k), Roth IRA (x2 MFJ), HSA, Mega Backdoor Roth, employer match, cash savings, and a final-year lump sum. S+ deferred comp also grows during employment.

**Draw Strategy:** Pre-RMD: PreTax up to 12% bracket → Cash → Roth → HSA. RMD Phase (75+): mandatory RMD first → Cash → Roth → HSA. Roth conversions fill remaining bracket space.

**Monte Carlo:** Returns ~ N(target, std_dev) by default, or Student-t if Fat-Tailed is enabled, capped at max upside. PreTax/Roth/HSA/Legacy Pool share a common market factor (Cross-Account Correlation slider); Cash/MM gets a smaller fraction of it and 1/3 the std dev.

**Legacy Gifting:** In a year with positive/neutral returns, the target Roth gift is withdrawn once (via the normal draw waterfall) and credited to a segregated Legacy Pool balance that compounds on its own return/std-dev assumption and is never drawn back down for household spending. In a down-market year, the gift is skipped entirely and the target amount simply stays invested — it isn't withdrawn, and "Legacy Inheritance" is a label for that still-invested amount, not a separate pot of money.

**After-Tax to Heirs:** Inherited pretax/HSA balances must be distributed within 10 years (SECURE Act) and are taxed at an assumed heir rate; Roth (including the Legacy Pool) passes tax-free. The Legacy tab shows both nominal and after-tax figures since a dollar in each bucket is not worth the same to your heirs.

**Surviving Spouse Scenario:** When enabled, filing status switches from MFJ to Single starting the year after the configured death age (the IRS allows MFJ in the year of death itself), pulling in single-filer federal brackets, a halved standard deduction, a halved IRMAA threshold, single Social Security taxability thresholds, and a configurable partial SS survivor benefit, pension survivor benefit, and living-expense reduction.

**Spending Strategy:** Fixed Real Spending (default) inflates your base expenses every year. Dynamic Guardrails (Guyton-Klinger style) instead compares last year's withdrawal rate to a band around your starting rate and applies a permanent step up/down in spending when breached.

**Cash Shortfalls:** Cash is allowed to go negative to represent a genuine funding gap; it is not floored to $0, so Total_Liquid_Assets and Monte Carlo success rates reflect real shortfalls rather than hiding them.

**Excel Export:** 4-sheet workbook: Assumptions, Accumulation, Retirement, Summary -- plus a 5th Monte Carlo sheet (percentile bands + success rate) whenever Monte Carlo is enabled.
        """)


if __name__ == "__main__":
    main()
