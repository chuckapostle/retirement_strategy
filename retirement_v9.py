"""
Retirement Income and Tax Planning Simulator v9
================================================
With accumulation phase, Monte Carlo, and Excel export.

Usage:
  pip install streamlit plotly pandas numpy openpyxl
  streamlit run retirement_v9.py

v15 cleanup (removed two unused features):
1. Removed the "Dynamic Guardrails (Guyton-Klinger)" spending strategy --
   guardrail_factor, its withdrawal-rate band/adjustment/floor/ceiling
   config, the Inflation Rule, and the Guardrail_Factor/Inflation_Frozen
   output columns are gone. Spending Strategy is now a two-way choice:
   Fixed Real Spending or Variable Percentage Withdrawal (VPW, default).
2. Removed the "Performance Draw Only" cash-draw option
   (performance_draw_only cfg key and its sidebar checkbox) -- the cash
   bucket's draw cap is now always the plain balance, same as every other
   account.

v14 fix (VPW was force-spending money nobody would actually consume):
1. Traced why late-plan spending under VPW spikes while the estate
collapses: VPW's withdrawal % is designed to amortize the portfolio to
(near) zero by planning_end_age, so the % necessarily accelerates
(clamped to 100% in the final year) as the horizon shrinks. But the
model had no concept of "withdrew more than was actually spent" -- the
full VPW-calculated amount always flowed into Base_Expenses/total_exp
and was drawn from the accounts as though 100% of it were consumed,
even in a large portfolio's final years where that could exceed a
realistic budget entirely (verified: a large enough portfolio could
even go temporarily insolvent once healthcare/HDHP/etc. stacked on top
of a 100%-of-portfolio VPW draw).
2. Added max_annual_expenses (today's $, UI default = your Annual
Expenses figure): a hard ceiling on base spending, applied BEFORE
total_exp/the draw waterfall are built, so anything above it simply
stays invested instead of being withdrawn at all -- not
drawn-then-unaccounted-for. Mirrors the existing
absolute_min_annual_expenses floor structurally. Also caps Guardrails'
Prosperity Rule upside; a no-op for Fixed Real Spending, which never
exceeds its own planned amount. Off (0) by default at the engine level
for direct API/test use.
3. Bumped the default Plan Through Age (planning_end_age) from 89 to 95
-- at 89 it was fighting against VPW's now-default selection, forcing
an aggressive drawdown-to-zero on a much shorter horizon than most
plans should assume. Still user-adjustable.

v13 enhancement (Variable Percentage Withdrawal -- new default strategy):
  Added VPW as a third Spending Strategy option alongside Fixed Real
  Spending and Dynamic Guardrails (Guyton-Klinger), and made it the
  default. Each year, base spending is a percentage of the CURRENT total
  portfolio balance, recalculated fresh via vpw_percentage() -- the
  standard amortization formula (same math as an RMD divisor: draw a
  balance earning an assumed real return to exactly zero over the
  remaining years), with the percentage rising as the remaining horizon
  shrinks. Popularized by the Bogleheads community specifically as an
  alternative to fixed or step-adjusted withdrawal rules. Unlike
  Guyton-Klinger's guardrail_factor, VPW has no persistent multiplier and
  therefore no "stuck" state: a bad year changes next year's dollar
  amount only because the portfolio shrank, and a full rebound restores
  spending to parity immediately (verified directly: an offsetting
  rebound the year after a -20% shock closes >98% of the gap by the
  following year, vs. Guyton-Klinger's cuts which -- per the v11 finding
  -- rarely reverse at all). The final simulated year is clamped to 100%
  (the raw ordinary-annuity formula gives slightly over 100% at n=1,
  which is meaningless as a share of a balance you actually have). Still
  respects the Absolute Minimum Annual Expenses floor as a hard backstop,
  same as the other two strategies. Purely additive: Fixed/Guardrails
  behavior and the regression pin are unchanged.

v12 default tuning (UI starting values, no engine/behavior changes):
  Annual Cash Savings 150k->40k, Final Year Cash Lump 120k->0, Brokerage
  Annual Contribution 0->200k, Fat-Tailed Returns off->on, Absolute
  Minimum Annual Expenses 0->145k, Spending Strategy Fixed->Dynamic
  Guardrails, Cumulative Spending Bound 90-150%->60-150%, Enable Roth
  Conversions off->on. Every one of these remains user-adjustable in the
  sidebar; only the value shown on first load changed.

v11 fix (guardrail floor was not actually a floor):
1. The v10 Inflation Rule applied via a separate, unbounded freeze on the
spending base, independent of Guardrail_Factor's floor/ceiling clamp.
Verified empirically: with the (then-)default 50% floor, among Monte
Carlo paths that "succeeded" (portfolio > $0 at the final age), median
real spending fell to ~40% of plan and as low as ~29% -- well past the
floor -- while Guardrail_Factor itself still reported a compliant 50%,
and MC success rate was inflated by roughly 30 percentage points versus
the same run with the Inflation Rule off. Fixed by folding the Inflation
Rule into Guardrail_Factor itself (a division by (1+infl) the year after
a down year, same permanent-step mechanism as the withdrawal-rate rule)
so BOTH rules share the same bounded number and the floor/ceiling clamp
applies to their combined effect. Confirmed via direct re-run: real
spending now bottoms out exactly at the configured floor, never below.
2. Added an absolute_min_annual_expenses floor (today's $, config default
0/off): a hard minimum on base spending, independent of what % the
guardrail floor happens to allow -- 50% of a padded plan may still be
generous, the same 50% of a tight one may not be livable. Applied after
all other reductions (post-80, widowhood, guardrail cuts), so nothing
can push spending below it once set.
3. Added a spending-percentile chart and a "worst real spending among
successful paths" metric to the Monte Carlo tab. Success rate alone
(portfolio > $0) says nothing about how much spending was cut to get
there; this makes the lifestyle experience behind a "success" visible
instead of only the asset-survival outcome.
4. Bumped the default Cumulative Spending Bound floor from 50% to 90% of
plan (guardrail_floor_pct). Deliberately tight -- in a 14%-std-dev
fat-tailed stress test this drops MC success rate from ~54% (at the
old, now-honestly-bounded 50% floor) to ~7%, because a 90% floor leaves
room for barely one adjustment step before it engages. That's an honest
number, not a bug: it reflects how little the strategy can now deviate
from Fixed Real Spending. Loosen it if you want more room for the
strategy to actually flex.

v10 enhancements (perf/caching + financial modeling, on top of v9):
  1. Monte Carlo and the scenario optimizer no longer recompute the
     (deterministic) accumulation phase on every single simulation --
     hoisted out and computed once (or once per unique retirement_age for
     the optimizer), a substantial speedup at high sim counts.
  2. run_simulation/run_monte_carlo/run_optimizer are now cached across
     Streamlit reruns (@st.cache_data), so moving a widget that doesn't
     change the underlying scenario (e.g. the "Today's Dollars" display
     toggle) no longer re-triggers a full Monte Carlo run. A seed of 0
     ("random") is deliberately left uncached so it keeps drawing fresh
     paths on every rerun, same as before.
  3. New taxable brokerage account bucket: sits between Cash and Roth in
     the draw order, tracks a running cost basis, realizes long-term
     capital gains on withdrawal taxed via the IRS "stacked on top of
     ordinary income" method (federal preferential rates; Oregon taxes it
     as ordinary income, no LTCG preference), and gets a full basis
     step-up at death like Roth/cash (unlike PreTax/HSA). Wired into
     Monte Carlo correlation, sensitivity analysis, and Excel export.
  4. Guyton-Klinger guardrails gained the "Inflation Rule": spending can
     now also skip a year's COLA following a negative portfolio return,
     alongside the existing capital-preservation/prosperity rules and the
     v8 cumulative floor/ceiling.
  5. Removed dead scaffolding that was computed but never used anywhere
     (FPL_700/Under_700_FPL and several other never-populated columns
     left over from an earlier version).
  6. Fixed a stale hardcoded base_year=2024 (now CURRENT_YEAR, computed
     live) -- cosmetic only (displayed Year column); tracing the math
     confirms no tax/dollar calculation ever depended on the specific
     value, since it always cancels out of the year-offset (yfb/yir) math.
  7. Added TAX_YEAR + verify_tax_constants(): a runtime banner that warns
     when nobody has re-verified the hardcoded federal/Oregon/LTCG
     bracket and IRMAA figures against actual current-year IRS/SSA
     numbers -- these are annually-adjusted figures baked in as of
     TAX_YEAR and will silently go stale otherwise.
  8. Added a pytest regression suite (tests/test_retirement_v9.py)
     covering the RMD table, Social Security taxability tiers, the tax
     engine, guardrail floor/ceiling, the new brokerage/LTCG math, the
     Inflation Rule, and the tax-year verifier.

v9 fix:
1. "Discretionary Draw Reduction if Year Return < 0" was a fixed nominal
dollar amount for the entire plan, while base_annual_expenses (and every
other spending line) inflates every year. That meant the same-looking
cut had a shrinking REAL bite over time -- 11.9% of base expenses in
year 1 vs. 5.1% by year 30 in a typical scenario, worth less than half
its original purchasing power. Now indexed to inflation like everything
else: entered in today's dollars, scaled by (1+infl)**yir each year, so
a bad year in year 1 and a bad year in year 30 get an equivalent real
cut. UI label updated to "(today's $)" to make this explicit.

v7 enhancements (financial modeling, on top of v6's correctness fixes):
  1. Cash/MM interest is now taxed annually as ordinary income (was
     completely untaxed before -- both growth and withdrawals).
  2. Legacy pool has its own configurable return/std-dev assumption
     instead of automatically matching your own Roth account.
  3. Monte Carlo now correlates PreTax/Roth/HSA/Legacy Pool returns via a
     shared market factor (configurable strength) instead of drawing each
     bucket fully independently, which understated true sequence-of-returns
     tail risk. Cash/MM gets a smaller fraction of that correlation.
  4. Optional fat-tailed (Student-t) return distribution as an alternative
     to Normal, for more realistic crash frequency/severity.
  5. Optional "surviving spouse" scenario: filing status switches from MFJ
     to Single the year after the first spouse's death, with single-filer
     brackets/deduction/IRMAA threshold, a partial SS survivor benefit, and
     optional pension and living-expense reductions (the "widow's penalty").
  6. Legacy/estate figures now also show an after-tax-to-heirs value:
     inherited pretax/HSA money is taxed at an assumed heir rate under the
     SECURE Act's 10-year rule, while Roth (including the legacy pool)
     passes tax-free -- these are not equivalent dollars to your heirs.
  7. Optional dynamic "guardrails" spending strategy (Guyton-Klinger style)
     as an alternative to fixed real spending: a persistent spending cut/
     raise triggered by your trailing withdrawal rate drifting outside a
     band around your starting rate.
  8. Excel export gained a Monte Carlo sheet (percentile bands + success
     rate) whenever Monte Carlo is enabled.
  9. New Sensitivity (tornado) tab: varies one assumption at a time to show
     which ones actually move your outcome the most.

v8 fixes (Guyton-Klinger guardrail realism):
1. guardrail_factor previously had no floor and no cap on cumulative
drift: a persistent multiplicative decay ((1-adj) compounded every
consecutive "stressed" year, with no reset) could compound base
spending down to an unrealistically low fraction of the original plan
during a prolonged downturn/Monte Carlo tail path. Now floored at a
configurable guardrail_floor_pct (default 50% of original spending)
and capped at guardrail_ceiling_pct (default 150%) on the upside, so
the model can never imply a household is living on, e.g., 20% of its
intended budget indefinitely.
2. Verified (no fix needed, confirmed correct): base_exp and all other
expense lines in a "bad return year" are computed from the SAME
inflation-adjusted formula used every other year (base_annual_expenses
* (1+infl)**yir, healthcare at its own inflation rate, etc.) before
any negative-return discretionary reduction or guardrail multiplier is
applied. The neg_ret_draw_reduction is capped at total_exp so it can
subtract but never invert sign the already-correctly-inflated total.
3. Legacy-contribution inflation growth verified correct (Legacy_Target_Total
compounds at the General Inflation rate every year in both the model and
the Excel export) -- confirmed by direct cell inspection. A flat $30,000
every year is only mathematically exact when General Inflation is set to
0%; any positive inflation rate produces genuine year-over-year growth.
4. Excel "Retirement" sheet was missing ~35 fields that exist in the model,
including Guardrail_Factor itself (needed to audit fix #1 above),
Filing_Status, the realized per-bucket returns (Return_PreTax/Roth/HSA/
Cash/Legacy_Pool -- needed to audit Monte Carlo correlation and fat
tails), the full tax/deduction breakdown, and the after-tax-to-heirs
figures. All added, with correct percent/boolean/text formatting
(previously everything not explicitly listed defaulted to money format).

v6 correctness fixes (from a full audit of v5):
  1. Legacy Roth gifts were being subtracted from cash TWICE (once via
     total_exp/the draw engine, again via an explicit cash-to-Roth
     transfer) -- fixed to a single, reconciling debit/credit.
  2. In down-market years the model withdrew the legacy gift as an
     "expense" but never deposited it anywhere (a phantom
     Legacy_Inheritance figure with no real balance behind it) -- fixed
     so the money simply stays invested and isn't withdrawn at all.
  3. Account "depleted" flags were one-way and never reset, permanently
     freezing growth on any account that hit zero and later recovered
     (e.g. Roth refilled by a legacy gift) -- now evaluated live.
  4. Negative cash was floored to $0 in Cash_EOY and Total_Liquid_Assets,
     hiding real funding shortfalls and inflating Monte Carlo success
     rates -- now shown/counted honestly.
  5. Itemized deductions double-counted the non-taxable portion of JSS
     income (already excluded from taxable income, then deducted again).
  6. Oregon taxable income over-subtracted an extra 15% of Social
     Security that was never part of the taxable base to begin with.
  7. Oregon Roth-conversion tax used a flat 9% guess instead of the
     actual marginal OREGON_BRACKETS calculation used for federal.
  8. RMD table stopped at age 100 and silently reused the age-95 factor
     for older ages -- extended through the IRS floor of 2.0 at 120+.
  9. RMD was computed on the current year's already-grown balance
     instead of the prior year-end balance (overstated RMDs slightly).
 10. Social Security taxability used a flat $6,000 tier-2 add-on instead
     of the IRS-correct min($6,000, 50% of benefits) cap.
 11. The "performance_draw_only" cash cap always used the static average
     return even inside Monte Carlo runs, instead of that year's actual
     simulated return.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import date
from pathlib import Path

# ============================================================
# CONSTANTS
# ============================================================

# Anchors the displayed "Year" column to the real current year. Purely
# cosmetic -- all tax-timing math uses year-offsets (yfb/yir) from this
# anchor, which cancel it out, so this never affects any dollar figure.
CURRENT_YEAR = date.today().year

# The year FEDERAL_BRACKETS, SINGLE_FEDERAL_BRACKETS, OREGON_BRACKETS,
# LTCG_BRACKETS/SINGLE_LTCG_BRACKETS, IRMAA_MAGI_THRESHOLD*,
# IRMAA_MONTHLY_SURCHARGE, and the UI's default standard_deduction were last
# checked against actual IRS Rev. Proc. / SSA COLA figures. Unlike
# CURRENT_YEAR, this does NOT auto-update -- it's a manual marker, bumped
# only when someone has actually re-verified the numbers below. Every
# *future* simulated year is already correctly inflated forward from
# whatever's here via bracket_inflation/inflation_rate regardless of how
# stale TAX_YEAR gets; what goes stale is the CURRENT-YEAR baseline every
# projection is anchored to. See verify_tax_constants() below.
TAX_YEAR = 2026

RMD_TABLE = {
    72: 27.4, 73: 26.5, 74: 25.5, 75: 24.6, 76: 23.7, 77: 22.9,
    78: 22.0, 79: 21.1, 80: 20.2, 81: 19.4, 82: 18.5, 83: 17.7,
    84: 16.8, 85: 16.0, 86: 15.2, 87: 14.4, 88: 13.7, 89: 12.9,
    90: 12.2, 91: 11.5, 92: 10.8, 93: 10.1, 94: 9.5, 95: 8.9,
    96: 8.4, 97: 7.8, 98: 7.3, 99: 6.8, 100: 6.4,
    # Table previously stopped at 100 and silently fell back to the age-95
    # factor (8.9) for any older age, understating RMDs for anyone whose
    # planning horizon runs past 100. Full IRS Pub. 590-B Uniform Lifetime
    # Table (Appendix B, Table III) continues to a floor of 2.0 at 120+.
    101: 6.0, 102: 5.6, 103: 5.2, 104: 4.9, 105: 4.6, 106: 4.3,
    107: 4.1, 108: 3.9, 109: 3.7, 110: 3.5, 111: 3.4, 112: 3.3,
    113: 3.1, 114: 3.0, 115: 2.9, 116: 2.8, 117: 2.7, 118: 2.5,
    119: 2.3, 120: 2.0,
}
RMD_TABLE_FLOOR_FACTOR = 2.0  # "120 and older" per the IRS table
IRMAA_MAGI_THRESHOLD = 206_000
IRMAA_MAGI_THRESHOLD_SINGLE = 103_000  # single-filer IRMAA tier-1 threshold (~half of MFJ)
IRMAA_MONTHLY_SURCHARGE = 230.80  # tier-1 (lowest) monthly Part B/D surcharge, per person
# Real Medicare IRMAA is a 5-tier staircase, not one on/off cliff -- surcharge
# roughly sextuples from the lowest tier to the highest. Expressed as
# multiples of IRMAA_MAGI_THRESHOLD*/IRMAA_MONTHLY_SURCHARGE above rather
# than as separate hardcoded dollar figures, so the whole staircase moves
# together when those two verified numbers are updated. Multiples mirror the
# recent-year CMS Part B schedule's relative spacing (thresholds roughly
# 1.00x/1.25x/1.56x/1.87x/3.64x of the tier-1 threshold; surcharges roughly
# 1.0x/2.5x/4.0x/5.5x/6.0x of the tier-1 surcharge) -- like the brackets
# above, re-verify against the current year's actual CMS tables, not just
# the tier-1 anchor figures (see verify_tax_constants).
IRMAA_TIERS = [
    (1.0000, 1.0), (1.2524, 2.5), (1.5631, 4.0), (1.8738, 5.5), (3.6408, 6.0),
]

# Net Investment Income Tax (IRC section 1411): 3.8% Medicare surtax on the
# LESSER of net investment income or the excess of MAGI over the threshold
# below. Fixed by statute since 2013 -- like OREGON_ESTATE_BRACKETS, NOT
# inflation-indexed, so it (deliberately) doesn't get a bracket_inflation
# scaling anywhere it's used.
NIIT_RATE = 0.038
NIIT_MAGI_THRESHOLD = 250_000  # MFJ
NIIT_MAGI_THRESHOLD_SINGLE = 200_000

# Gompertz-law mortality approximation, used only by the optional Monte
# Carlo mortality-based planning horizon (see draw_household_death_age
# below) -- NOT a transcription of the published SSA Period Life Table,
# which this file doesn't carry verified age-by-age figures for. Gompertz's
# law -- mortality hazard rising exponentially with age -- is the standard
# actuarial model real tables are smoothed against; calibrated here so the
# resulting annual death probability lands in the right neighborhood at a
# few reference ages (~1.4% at 65, ~9% at 85, ~20% at 95), doubling roughly
# every 7.5 years. Treat this as a reasonable SHAPE for old-age mortality,
# not a certified actuarial table -- don't use it for anything beyond this
# app's own what-if planning-horizon exploration.
MORTALITY_GOMPERTZ_A = 0.00889
MORTALITY_GOMPERTZ_B = 0.0924  # ln(2)/7.5
MORTALITY_GOMPERTZ_REF_AGE = 60

FEDERAL_BRACKETS = [
    (0.10, 24_800), (0.12, 100_800), (0.22, 211_400),
    (0.24, 403_550), (0.32, 512_450), (0.35, 768_700),
    (0.37, float("inf")),
]
# Single-filer brackets: ~half of MFJ through the 32% bracket (matches how the
# real brackets compare), then diverging at the top since MFJ gets
# disproportionately more room in the 35% bracket than a simple halving would give.
SINGLE_FEDERAL_BRACKETS = [
    (0.10, 12_400), (0.12, 50_400), (0.22, 105_700),
    (0.24, 201_775), (0.32, 256_225), (0.35, 609_350),
    (0.37, float("inf")),
]
OREGON_BRACKETS = [
    (0.0475, 8_824), (0.0675, 22_059),
    (0.0875, 250_000), (0.0990, 10_000_000),
]
# Oregon's single-vs-MFJ bracket differences are much smaller than federal's;
# OREGON_BRACKETS is used for both filing statuses as a documented simplification.

# Oregon Estate Transfer Tax (ORS 118.010). One of the lowest state estate-tax
# exemptions in the country -- $1M, unlike the ~$15M federal exemption, which
# is not binding for most plans this model would ever flag. Unlike the income
# brackets above, this exemption/these brackets are NOT indexed for inflation
# by statute, so -- deliberately -- calc_tax is called on this table with
# yrs=0/infl=0 (no bracket_inflation growth) wherever it's used: an $8M estate
# is taxed the same whether it's reached in year 1 or year 30. First tier
# below is the 0%-taxed exemption amount itself, expressed as a bracket so
# calc_tax's existing marginal-band logic handles it with no special-casing.
OREGON_ESTATE_BRACKETS = [
    (0.00, 1_000_000),
    (0.10, 1_500_000), (0.1025, 2_500_000), (0.105, 3_500_000),
    (0.11, 4_500_000), (0.115, 5_500_000), (0.12, 6_500_000),
    (0.13, 7_500_000), (0.14, 8_500_000), (0.15, 9_500_000),
    (0.16, float("inf")),
]

# Long-term capital gains brackets. Federal only -- Oregon has no
# preferential capital-gains rate (gains are taxed as ordinary income under
# OREGON_BRACKETS, same as everything else). Approximate current-law
# figures; like the ordinary brackets above, these need an annual refresh
# (see TAX_YEAR / verify_tax_constants below).
LTCG_BRACKETS = [
    (0.00, 96_700), (0.15, 600_050), (0.20, float("inf")),
]
SINGLE_LTCG_BRACKETS = [
    (0.00, 48_350), (0.15, 533_400), (0.20, float("inf")),
]

def verify_tax_constants():
    """Returns a warning string if nobody has bumped TAX_YEAR since the real
    calendar rolled past it -- i.e. the CURRENT-year tax-law baseline this
    whole projection is anchored to hasn't been checked against actual IRS/
    SSA figures for the year we're actually in. Returns None when current.
    A runtime nag, not a correctness guarantee: bumping TAX_YEAR without
    actually updating the bracket constants below silences this without
    fixing anything -- it only proves someone looked, not that the numbers
    are right."""
    current_year = date.today().year
    if current_year <= TAX_YEAR:
        return None
    years_stale = current_year - TAX_YEAR
    return (
        f"Tax-law constants were last verified for {TAX_YEAR}, but it's now {current_year} "
        f"({years_stale} year{'s' if years_stale != 1 else ''} stale). FEDERAL_BRACKETS, "
        f"SINGLE_FEDERAL_BRACKETS, OREGON_BRACKETS, LTCG_BRACKETS/SINGLE_LTCG_BRACKETS, "
        f"IRMAA_MAGI_THRESHOLD*, IRMAA_MONTHLY_SURCHARGE, and the default standard deduction "
        f"are all annually-adjusted figures -- check them against the current year's IRS Rev. "
        f"Proc. and SSA COLA announcement, then bump TAX_YEAR. OREGON_ESTATE_BRACKETS and "
        f"NIIT_MAGI_THRESHOLD*/NIIT_RATE aren't annually adjusted (fixed by statute), but "
        f"re-check them too in case the legislature/Congress has changed the underlying law "
        f"since {TAX_YEAR}."
    )

# ============================================================
# TAX ENGINE
# ============================================================

def calc_tax(taxable, brackets, yrs, infl):
    if taxable <= 0: return 0.0
    tax, prev = 0.0, 0.0
    for rate, ceil in brackets:
        ac = ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
        band = max(0.0, min(taxable, ac) - prev)
        tax += band * rate; prev = ac
        if taxable <= ac: break
    return tax

def calc_ltcg_tax(gain, ordinary_taxable, ltcg_brackets, yrs, infl):
    """Long-term capital gains tax under the IRS 'stacked on top of ordinary
    income' method (the Qualified Dividends and Capital Gain Tax Worksheet):
    the gain occupies bracket space starting where ordinary_taxable already
    fills up to, using the LTCG bracket thresholds -- so other ordinary
    income (e.g. a Roth conversion) stacked below it can push the gain into
    a higher LTCG bracket, same as in real life."""
    if gain <= 0: return 0.0
    tax, prev = 0.0, 0.0
    stack_top = ordinary_taxable + gain
    for rate, ceil in ltcg_brackets:
        ac = ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
        band = max(0.0, min(stack_top, ac) - max(prev, ordinary_taxable))
        tax += band * rate; prev = ac
        if stack_top <= ac: break
    return tax

def bracket_ceiling(brackets, target_rate, yrs, infl):
    for rate, ceil in brackets:
        if rate >= target_rate:
            return ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
    return 0.0

def ss_taxable_portion(ss, other, single=False):
    if ss <= 0: return 0.0
    prov = other + ss * 0.5
    tier1, tier2 = (25_000, 34_000) if single else (32_000, 44_000)
    tier2_cap = 4_500 if single else 6_000  # single-filer worksheet cap is $4,500 (half of MFJ's $6,000)
    if prov < tier1: return 0.0
    elif prov < tier2: return min(ss * 0.5, (prov - tier1) * 0.5)
    else:
        # IRS worksheet caps the tier-2 add-on at min(cap, 50% of SS
        # benefits) — a flat cap overstates taxable SS whenever benefits
        # are low relative to the cap (0.5*ss < cap).
        tier2_addon = min(tier2_cap, ss * 0.5)
        return min(ss * 0.85, tier2_addon + (prov - tier2) * 0.85)

def irmaa_tier_multiplier(magi, threshold_base, yrs, infl):
    """Which IRMAA_TIERS bracket `magi` falls into against this year's
    inflated threshold schedule -- returns the surcharge MULTIPLE (0.0 if
    under the lowest tier). Iterates tiers in ascending order and keeps
    overwriting, so the highest tier `magi` clears wins, mirroring the same
    marginal-band walk calc_tax uses for income brackets (except IRMAA is a
    true cliff per tier, not a marginal blend -- crossing a threshold by $1
    jumps the WHOLE premium to that tier's surcharge, unlike an income
    bracket where only the dollars above the line get the higher rate)."""
    mult = 0.0
    for thr_mult, sur_mult in IRMAA_TIERS:
        thr = threshold_base * thr_mult * (1 + infl) ** yrs
        if magi > thr:
            mult = sur_mult
    return mult

def get_rmd(bal, age, start):
    if age < start or bal <= 0: return 0.0
    if age in RMD_TABLE:
        f = RMD_TABLE[age]
    elif age > max(RMD_TABLE):
        f = RMD_TABLE_FLOOR_FACTOR  # IRS table floors at 2.0 for 120+
    else:
        f = RMD_TABLE[min(RMD_TABLE)]  # shouldn't happen given the age < start guard above
    return bal / f if f > 0 else 0.0

def mortality_annual_hazard(age):
    """Approximate probability of death within this year of age, per the
    Gompertz calibration above (MORTALITY_GOMPERTZ_*) -- see that comment for
    what this is and isn't."""
    h = MORTALITY_GOMPERTZ_A * np.exp(MORTALITY_GOMPERTZ_B * (age - MORTALITY_GOMPERTZ_REF_AGE))
    return 1 - np.exp(-h)

def draw_mortality_age(start_age, max_age, rng):
    """One random age-at-death for a single life, walking year by year from
    start_age using mortality_annual_hazard. Capped at max_age (the plan's
    own horizon) -- the rare tail draw that reaches it is treated as
    'survived to the edge of what this plan was simulated for', not an
    assertion that death is impossible past that age."""
    age = start_age
    while age < max_age:
        if rng.random() < mortality_annual_hazard(age):
            return age
        age += 1
    return max_age

def draw_household_death_age(start_age, max_age, rng):
    """Two independent single-life draws from the same curve, household
    horizon = the LATER of the two -- a couple's plan has to survive until
    the SURVIVING spouse's death, not either individual's alone, which is
    why joint life expectancy runs longer than either person's own.
    Simplifying assumption: the two lifespans are independent draws (no
    modeling of correlated mortality from shared lifestyle/socioeconomic
    factors, which in reality would narrow this gap somewhat)."""
    return max(draw_mortality_age(start_age, max_age, rng), draw_mortality_age(start_age, max_age, rng))

def vpw_percentage(age, planning_end_age, real_return):
    """Variable Percentage Withdrawal: the fraction of the CURRENT portfolio
    balance to spend this year, per the standard amortization formula (the
    same math behind a mortgage payment or an RMD divisor) -- draw down a
    balance earning `real_return` annually to exactly zero over the
    remaining n years. Popularized by the Bogleheads community as an
    alternative to fixed real spending: self-correcting every year by
    construction (this year's % applies to whatever the portfolio actually
    is, good or bad), so a bad year never leaves a permanent "stuck"
    reduction. The percentage rises with age as the remaining horizon n
    shrinks --
    that's intentional, not a bug: less time left means a larger share of
    what's left gets drawn each year."""
    n = max(1, planning_end_age - age + 1)  # years remaining, inclusive of this one
    if abs(real_return) < 1e-9:
        pct = 1.0 / n
    else:
        pct = real_return / (1 - (1 + real_return) ** (-n))
    # The ordinary-annuity formula above gives slightly MORE than 100% in the
    # final year (n=1) for any positive real_return -- mathematically
    # correct under an end-of-year-payment convention, but nonsensical as a
    # withdrawal rate against a CURRENT balance (you can't spend more than
    # you have). Clamp to 100%: the last year simply spends what's left.
    return min(pct, 1.0)

def _standardized_shock(n, rng, fat_tailed, t_df):
    """Mean-0, unit-variance shock series. Normal by default; Student-t gives
    fatter tails (more frequent/severe extreme years) when fat_tailed=True."""
    if fat_tailed:
        raw = rng.standard_t(t_df, size=n)
        return raw / np.sqrt(t_df / (t_df - 2))  # rescale to unit variance
    return rng.standard_normal(n)

def generate_correlated_returns(bucket_specs, n, rng, correlation=0.85, fat_tailed=False, t_df=5, common=None):
    """
    bucket_specs: {name: (target, std, max_up)}.
    All buckets share one common market shock (weighted by `correlation`),
    plus their own idiosyncratic shock -- mirrors how real accounts holding
    similar underlying assets move together in the same year rather than
    being statistically independent of each other.
    `common`: pass in a pre-drawn shock (e.g. from another call to this
    function) to correlate this bucket set against THAT SAME realized market
    factor instead of drawing an independent one -- what build_mc_return_overrides
    uses to correlate stochastic inflation to the actual equity common factor
    rather than to a fresh, unrelated draw.
    """
    if common is None:
        common = _standardized_shock(n, rng, fat_tailed, t_df)
    out = {}
    for name, (target, std, mx) in bucket_specs.items():
        idio = _standardized_shock(n, rng, fat_tailed, t_df)
        z = correlation * common + np.sqrt(max(0.0, 1 - correlation ** 2)) * idio
        out[name] = np.clip(target + std * z, -1.0, mx)
    return out

def build_mc_return_overrides(cfg, num_years, rng, std_dev, max_up, correlation=0.85, fat_tailed=False, t_df=5):
    """One year-by-year correlated return series per bucket for a single
    Monte Carlo path, including the legacy pool's own return/std assumption."""
    # Drawn once and shared with the equity buckets AND inflation below, so
    # inflation is correlated to the SAME realized market factor driving
    # this path's actual returns -- not an unrelated fresh draw. (Cash/MM
    # below still draws its own independent common factor, unchanged from
    # before -- money-market rates track policy rates, not the equity market.)
    common = _standardized_shock(num_years, rng, fat_tailed, t_df)
    bucket_specs = {
        "pretax": (cfg["pretax_return"], std_dev, max_up),
        "roth": (cfg["roth_return"], std_dev, max_up),
        "hsa": (cfg["hsa_return"], std_dev, max_up),
        "legacy_pool": (cfg.get("legacy_pool_return", cfg["roth_return"]),
                         cfg.get("legacy_pool_std", std_dev), max_up),
        "brokerage": (cfg.get("brokerage_return", cfg["roth_return"]),
                       cfg.get("brokerage_std", std_dev), max_up),
    }
    ov = generate_correlated_returns(bucket_specs, num_years, rng, correlation, fat_tailed, t_df, common=common)
    # Cash/MM: much lower volatility, only lightly correlated with the
    # broader equity market (money-market rates track policy rates, not stocks).
    ov["cash"] = generate_correlated_returns(
        {"cash": (cfg["cash_return"], std_dev / 3, max_up / 2)},
        num_years, rng, correlation * 0.3, fat_tailed, t_df,
    )["cash"]
    # Stochastic inflation: previously a single fixed rate applied to every
    # MC path, which understates real risk in a plan whose returns are
    # nominal and expenses inflate deterministically -- inflation surprise is
    # arguably the dominant risk over a 40-year horizon, and a 1970s-style
    # path was structurally impossible to draw before this. Correlated
    # (negatively, by default) to the same common equity factor above: a
    # market-crash year is more likely to also be a high-inflation surprise
    # (echoing 2022), though the real-world relationship is noisy and this is
    # a simplifying assumption, not a calibrated historical estimate --
    # configurable via inflation_correlation, including 0 (independent) or
    # positive if you'd rather model it the other way.
    infl_std = cfg.get("inflation_std", 0.0)
    if infl_std > 0:
        ov["inflation"] = generate_correlated_returns(
            {"inflation": (cfg["inflation_rate"], infl_std, 0.15)},
            num_years, rng, cfg.get("inflation_correlation", -0.15), fat_tailed, t_df, common=common,
        )["inflation"]
        # Floor well above the -1.0 (-100%) floor generate_correlated_returns
        # applies for asset returns -- sustained deflation is a real but rare
        # and structurally different regime this simple shock isn't meant to
        # capture; -2% is generous room for a single bad deflationary year
        # without letting the random walk collapse to something absurd.
        ov["inflation"] = np.clip(ov["inflation"], -0.02, None)
    return ov


def load_starting_balances(path="starting_balances.txt"):
    defaults = {
        "401k": None,
        "roth": None,
        "hsa": None,
        "s_plus_5": None,
        "s_plus_10": None,
        "cash": None,
        "brokerage": None,
    }
    if not Path(path).exists():
        return defaults, False
    aliases = {
        "401k": "401k",
        "401(k)": "401k",
        "pretax": "401k",
        "roth": "roth",
        "rothira": "roth",
        "roth_ira": "roth",
        "hsa": "hsa",
        "s+5": "s_plus_5",
        "s5": "s_plus_5",
        "s_plus_5": "s_plus_5",
        "s+10": "s_plus_10",
        "s10": "s_plus_10",
        "s_plus_10": "s_plus_10",
        "cash": "cash",
        "brokerage": "brokerage",
        "taxable": "brokerage",
        "taxablebrokerage": "brokerage",
    }
    loaded = False
    try:
        for raw in Path(path).read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.replace(":", " ").split()
            if len(parts) < 2:
                continue
            key = parts[0].strip().lower().replace(" ", "").replace("-", "").replace("/", "")
            key = aliases.get(key)
            if key is None:
                continue
            try:
                defaults[key] = int(float(parts[1].replace(",", "").replace("$", "")))
                loaded = True
            except ValueError:
                pass
    except Exception:
        return defaults, False
    return defaults, loaded
# ============================================================
# ACCUMULATION PHASE
# ============================================================

def run_accumulation(cfg):
    """Grow balances from current age to retirement with contributions."""
    cur_age = cfg["current_age"]
    ret_age = cfg["retirement_age"]
    # base_year: the calendar year Current Age is measured AS OF (not
    # necessarily today's literal date -- see the "Ages As Of Year" input).
    # The full (retirement_age - current_age) gap is real, uninterrupted
    # accumulation time starting from THIS year -- e.g. if today is most of
    # the way through 2026 with no more contributions coming, set this to
    # 2027 and enter Current Age as your age AS OF 2027, not today; the gap
    # to Retirement Age then plays out entirely in 2027-onward, with no
    # separate "dead year" bookkeeping needed here.
    base_year = cfg.get("plan_base_year", CURRENT_YEAR)
    years = ret_age - cur_age
    bk0 = float(cfg.get("brokerage", 0.0))
    # Starting cost basis defaults to the starting balance (no embedded gain
    # assumed at t=0) but is overridable for an account that already has
    # unrealized gains today.
    bk_basis0 = float(cfg.get("brokerage_basis", bk0))
    if years <= 0:
        return ([], cfg["pretax_401k"], cfg["roth_ira"], cfg["hsa"], cfg["cash"],
                cfg["s_plus_5yr"], cfg["s_plus_10yr"], bk0, bk_basis0)

    pt = float(cfg["pretax_401k"])
    ro = float(cfg["roth_ira"])
    hs = float(cfg["hsa"])
    ca = float(cfg["cash"])
    s5 = float(cfg["s_plus_5yr"])
    s10 = float(cfg["s_plus_10yr"])
    bk = bk0
    bk_basis = bk_basis0

    # Annual contributions
    c_401k = cfg.get("contrib_401k", 24_500)
    c_roth401k = cfg.get("contrib_roth401k", 8_000)
    c_roth_ira = cfg.get("contrib_roth_ira", 8_700)  # x2 for MFJ
    c_hsa = cfg.get("contrib_hsa", 8_700)
    c_mega_backdoor = cfg.get("contrib_mega_backdoor", 29_000)
    c_employer_match = cfg.get("contrib_employer_match", 18_000)
    c_cash_annual = cfg.get("contrib_cash_annual", 150_000)
    c_cash_final_lump = cfg.get("contrib_cash_final_lump", 120_000)
    c_brokerage_annual = cfg.get("contrib_brokerage_annual", 0)

    # Growth rates (same as retirement performance assumptions)
    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]
    pr_bk = cfg.get("brokerage_return", pr_ro)

    rows = []
    for i in range(years):
        age = cur_age + i
        yr = base_year + i

        # Growth first (skip year 0 — balances are "today")
        if i > 0:
            pt *= (1 + pr_pt)
            ro *= (1 + pr_ro)
            hs *= (1 + pr_hs)
            ca *= (1 + pr_ca)
            bk *= (1 + pr_bk)  # basis is untouched by growth
            # S+ grows in deferred comp while employed
            s5 *= (1 + pr_pt)
            s10 *= (1 + pr_pt)

        # Contributions (end of year) -- brokerage contributions add 1:1 to
        # both balance and basis (no gain on money that was just contributed)
        pt += c_401k + c_employer_match
        ro += c_roth401k + c_roth_ira * 2 + c_mega_backdoor  # 2x Roth IRA for MFJ
        hs += c_hsa
        ca += c_cash_annual
        bk += c_brokerage_annual
        bk_basis += c_brokerage_annual

        # Final year lump sum
        if i == years - 1:
            ca += c_cash_final_lump

        rows.append({
            "Phase": "Accumulation",
            "Age": age, "Year": yr,
            "PreTax_EOY": pt, "Roth_EOY": ro, "HSA_EOY": hs, "Cash_EOY": ca,
            "Brokerage_EOY": bk, "Brokerage_Basis": bk_basis,
            "S_Plus_5yr": s5, "S_Plus_10yr": s10,
            "Contrib_PreTax": c_401k + c_employer_match,
            "Contrib_Roth": c_roth401k + c_roth_ira * 2 + c_mega_backdoor,
            "Contrib_HSA": c_hsa,
            "Contrib_Cash": c_cash_annual + (c_cash_final_lump if i == years - 1 else 0),
            "Contrib_Brokerage": c_brokerage_annual,
            "Total_Liquid_Assets": pt + ro + hs + ca + bk,
        })

    return rows, pt, ro, hs, ca, s5, s10, bk, bk_basis


# ============================================================
# RETIREMENT SIMULATION ENGINE
# ============================================================

def run_simulation(cfg, return_overrides=None, accum_result=None):
    results = []
    base_year = cfg.get("plan_base_year", CURRENT_YEAR)
    ret_age = cfg["retirement_age"]
    cur_age = cfg["current_age"]
    end_age = cfg["planning_end_age"]
    # Must match run_accumulation: same base_year, same uninterrupted
    # (ret_age - cur_age)-year gap starting from it.
    ret_year = base_year + (ret_age - cur_age)
    infl = cfg["inflation_rate"]
    binfl = cfg["bracket_inflation"]
    num_years = end_age - ret_age + 1

    # Stochastic inflation (Monte Carlo only, via build_mc_return_overrides):
    # a per-year realized-inflation path instead of the single flat rate
    # every other simulation uses. cum_infl[yir] = cumulative compounding
    # from year 0 (retirement start, the "today's dollars" baseline) through
    # year yir -- built once here, replacing every (1 + infl) ** yir below
    # with cum_infl[yir]. Deterministic runs (baseline, sensitivity,
    # optimizer, or an MC path with inflation_std=0) have no "inflation" key
    # in return_overrides and fall through to the flat rate exactly as before.
    infl_path = return_overrides.get("inflation") if return_overrides else None
    if infl_path is not None:
        cum_infl = np.ones(num_years)
        if num_years > 1:
            cum_infl[1:] = np.cumprod(1 + infl_path[:-1])
    else:
        cum_infl = None

    # Accumulation phase is deterministic given cfg -- it never reads
    # return_overrides -- so it's identical across every Monte Carlo path
    # and across every optimizer combo that shares the same retirement_age.
    # Callers that run many simulations against the same cfg can compute it
    # once and pass it in via accum_result instead of recomputing it on
    # every single call (run_monte_carlo, run_optimizer both do this).
    if accum_result is None:
        accum_result = run_accumulation(cfg)
    accum_rows, pt, ro, hs, ca, s5_bal, s10_bal, bk, bk_basis = accum_result
    cash_basis = float(accum_rows[-1].get("Cash_Basis", cfg["cash"])) if accum_rows else float(cfg["cash"])

    # S+ payout tracking
    s5_annual = s5_bal / 5.0 if s5_bal > 0 else 0.0
    s10_annual = s10_bal / 10.0 if s10_bal > 0 else 0.0
    s5_rem, s10_rem = 5, 10
    jss_rec_rem = cfg["jss_recovery_years"]

    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]

    cum_gifts, cum_legacy_roth, cum_legacy_inheritance, cum_lump_sums = 0.0, 0.0, 0.0, 0.0
    CASH_DISTRESS_FLOOR = -50_000  # below this, we stop accruing "growth" on a cash deficit
    legacy_pool = 0.0  # money already gifted into the kids' Roth accounts -- legally
    # theirs, not the household's. Kept segregated from `ro` (the parents' own
    # Roth) so it (a) compounds on its own and can be charted, and (b) can
    # never be drawn back out by the household's own retirement withdrawals.

    use_vpw = cfg.get("spending_strategy") == "vpw"

    for idx in range(num_years):
        age = ret_age + idx
        yr = ret_year + idx
        yfb = yr - base_year
        yir = idx

        row = {"Phase": "Retirement", "Age": age, "Year": yr, "Years_Retired": yir}

        # Cumulative general-inflation factor for this year -- stochastic
        # per-path if infl_path was supplied, otherwise the flat rate's usual
        # (1 + infl) ** yir. Used everywhere general inflation compounds a
        # today's-dollars figure forward (expenses, gifts, HDHP, etc.);
        # bracket_inflation (binfl/yfb) and healthcare_inflation are separate,
        # still-deterministic assumptions, unaffected by this.
        infl_factor_yir = cum_infl[yir] if cum_infl is not None else (1 + infl) ** yir
        row["Inflation_Rate_Realized"] = infl_path[yir] if infl_path is not None else infl

        # Surviving-spouse ("widow's penalty") scenario: filing status switches
        # from MFJ to Single starting the year AFTER the first spouse's death
        # (the IRS allows MFJ status for the year of death itself). Off by
        # default -- when disabled, behavior is identical to before.
        is_widowed = bool(cfg.get("model_widow_scenario", False)) and age > cfg.get("first_death_age", 999)
        row["Filing_Status"] = "Single (Widowed)" if is_widowed else "MFJ"
        fed_brackets_yr = SINGLE_FEDERAL_BRACKETS if is_widowed else FEDERAL_BRACKETS
        ltcg_brackets_yr = SINGLE_LTCG_BRACKETS if is_widowed else LTCG_BRACKETS
        std_ded_multiplier = 0.5 if is_widowed else 1.0  # single standard deduction is ~half of MFJ's
        irmaa_threshold_base = IRMAA_MAGI_THRESHOLD_SINGLE if is_widowed else IRMAA_MAGI_THRESHOLD

        # ── GROWTH ──
        # "Depleted" is now evaluated live off the current balance every year,
        # instead of a one-way flag that, once set, permanently froze growth
        # even after a balance recovered (e.g. a Roth that hit zero and was
        # later refilled by a legacy contribution or Roth conversion never
        # used to earn interest again under the old logic).
        if return_overrides:
            r_pt, r_ro = return_overrides["pretax"][yir], return_overrides["roth"][yir]
            r_hs, r_ca = return_overrides["hsa"][yir], return_overrides["cash"][yir]
            r_lp = return_overrides["legacy_pool"][yir] if "legacy_pool" in return_overrides else r_ro
            r_bk = return_overrides["brokerage"][yir] if "brokerage" in return_overrides else r_ro
        else:
            r_pt, r_ro, r_hs, r_ca = pr_pt, pr_ro, pr_hs, pr_ca
            r_lp = cfg.get("legacy_pool_return", pr_ro)
            r_bk = cfg.get("brokerage_return", pr_ro)

        # Snapshot the pretax balance as of the *prior* year-end, before this
        # year's growth is applied — RMDs are legally based on the 12/31
        # balance from the year before, not a balance that already includes
        # this year's market movement.
        pt_prior_year_end = pt
        ca_pre_growth = ca  # snapshot before growth, used to compute this year's taxable interest below
        bk_pre_growth = bk  # snapshot before growth, used to compute this year's taxable dividend below

        if yir > 0:
            if pt > 0: pt *= (1 + r_pt)
            if ro > 0: ro *= (1 + r_ro)
            if hs > 0: hs *= (1 + r_hs)
            if ca > CASH_DISTRESS_FLOOR: ca *= (1 + r_ca)
            if legacy_pool > 0: legacy_pool *= (1 + r_lp)
            if bk > 0: bk *= (1 + r_bk)  # basis is untouched by growth
            row["Return_PreTax"], row["Return_Roth"] = r_pt, r_ro
            row["Return_HSA"], row["Return_Cash"] = r_hs, r_ca
            row["Return_Legacy_Pool"], row["Return_Brokerage"] = r_lp, r_bk
        else:
            row["Return_PreTax"] = row["Return_Roth"] = row["Return_HSA"] = row["Return_Cash"] = 0.0
            row["Return_Legacy_Pool"] = row["Return_Brokerage"] = 0.0

        # Cash/MM interest is taxable annually as ordinary income (a money
        # market fund doesn't defer tax the way an equity account's
        # unrealized gains do) -- computed off the balance BEFORE this year's
        # growth and draws. Previously cash growth and withdrawals were never
        # taxed anywhere in the model.
        cash_interest_taxable = 0.0
        if cfg.get("tax_cash_interest", True) and yir > 0 and ca_pre_growth > 0:
            cash_interest_taxable = ca_pre_growth * r_ca
        row["Cash_Interest_Taxable"] = cash_interest_taxable

        # Brokerage dividends: unlike price appreciation (deferred until you
        # sell) and PreTax/HSA (fully deferred), a taxable brokerage
        # account's dividend distribution is taxed the year it's paid,
        # whether or not you spend it -- a real, ongoing tax drag that pure
        # buy-and-hold price return ignores. Modeled as a fixed fraction of
        # the pre-growth balance (independent of r_bk -- funds keep paying
        # dividends in a down year too), taxed as a qualified dividend at
        # the same preferential rate as LTCG (added into brokerage_ltcg_gain
        # below for tax purposes, tracked separately here for transparency).
        # Assumed reinvested rather than paid to cash, so it raises the
        # account's cost basis by the same amount -- real IRS treatment: a
        # reinvested dividend is a taxable distribution immediately followed
        # by a purchase of new shares at that price.
        brokerage_dividend_taxable = 0.0
        if yir > 0 and bk_pre_growth > 0:
            brokerage_dividend_taxable = bk_pre_growth * cfg.get("brokerage_dividend_yield", 0.0)
            bk_basis += brokerage_dividend_taxable
        row["Brokerage_Dividend_Taxable"] = brokerage_dividend_taxable

        # ── INCOME SOURCES ──
        sp_inc = 0.0
        if s5_rem > 0: sp_inc += s5_annual; s5_rem -= 1
        if s10_rem > 0: sp_inc += s10_annual; s10_rem -= 1
        row["S_Plus_Income"] = sp_inc

        jss_inc = jss_tax = 0.0
        if age >= cfg["jss_start_age"]:
            jss_inc = cfg["jss_annual_amount"] * (1 + cfg["jss_cola"]) ** (age - cfg["jss_start_age"])
            if jss_rec_rem > 0: jss_rec_rem -= 1
            else: jss_tax = jss_inc
        if is_widowed:
            # Pension survivorship: many plans pay a reduced (or zero) benefit
            # to the survivor. Default 100% (no change) unless configured.
            jss_survivor_pct = cfg.get("jss_survivor_pct", 1.0)
            jss_inc *= jss_survivor_pct
            jss_tax *= jss_survivor_pct
        row["JSS_Income"], row["JSS_Taxable"] = jss_inc, jss_tax

        ss_inc = 0.0
        if age >= cfg["ss_start_age"]:
            ss_inc = cfg["ss_annual_amount"] * (1 + cfg["ss_cola"]) ** (age - cfg["ss_start_age"])
        if is_widowed:
            # Social Security survivor benefit: the survivor keeps the HIGHER
            # of the two individual benefits, effectively losing the smaller
            # one. Modeled as a configurable % of the combined household
            # benefit retained (default 65%).
            ss_inc *= cfg.get("ss_survivor_pct", 0.65)
        row["SS_Income"] = ss_inc

        rent_inc = cfg["rental_gross"] * infl_factor_yir
        rent_tax = rent_inc * cfg["rental_taxable_pct"]
        row["Rental_Income"], row["Rental_Taxable"] = rent_inc, rent_tax

        # ── EXPENSES ──
        # Pure inflation-adjusted "Annual Expenses" figure, before ANY
        # strategy-driven adjustment (post-80, widowhood, negative-return
        # cut). Recorded for audit -- confirms every scenario, including
        # bad-return years, starts from the correctly inflated baseline
        # rather than some stale/un-inflated number.
        if use_vpw:
            # Bogleheads-style VPW: base spending is a percentage of the
            # CURRENT total portfolio (post-growth, pre-draw balances
            # already in scope here), recomputed fresh every year -- no
            # inflation projection or persistent multiplier involved, since
            # this year's nominal portfolio value already reflects
            # everything (growth, inflation-driven asset prices) up to now.
            total_liquid_pre_draw = pt + ro + hs + ca + bk
            vpw_pct = vpw_percentage(age, cfg["planning_end_age"], cfg.get("vpw_real_return_pct", 0.04))
            base_exp_inflated_only = total_liquid_pre_draw * vpw_pct
            row["VPW_Percentage"] = vpw_pct
        else:
            base_exp_inflated_only = cfg["base_annual_expenses"] * infl_factor_yir
            row["VPW_Percentage"] = 0.0
        base_exp = base_exp_inflated_only
        # Post-80 expense reduction (lifestyle slowdown)
        if age >= 80:
            base_exp *= (1 - cfg.get("expense_reduction_post80", 0.0))
        # Widowhood expense reduction (one person's living costs, on top of
        # any post-80 reduction that also applies)
        if is_widowed:
            base_exp *= (1 - cfg.get("expense_reduction_widowhood", 0.0))
        # Absolute survival floor: a hard minimum on base (discretionary/
        # lifestyle) spending, in TODAY'S dollars, that no combination of
        # the reductions above (post-80, widowhood) can push spending below.
        # Off (0) by default; opt in with your own bare-minimum number.
        abs_min = cfg.get("absolute_min_annual_expenses", 0.0)
        abs_min_inflated = abs_min * infl_factor_yir if abs_min > 0 else 0.0
        row["Absolute_Min_Inflated"] = abs_min_inflated
        row["Absolute_Min_Bound_Hit"] = bool(abs_min > 0 and base_exp < abs_min_inflated)
        if abs_min > 0:
            base_exp = max(base_exp, abs_min_inflated)
        # Spending ceiling: a hard maximum on base spending, in TODAY'S
        # dollars. Matters most for VPW, whose withdrawal percentage
        # accelerates toward the plan's end (see vpw_percentage) and would
        # otherwise force this year's ENTIRE calculated draw to be treated
        # as spent -- the model has no separate concept of "withdrew more
        # than I actually spent," so an unbounded VPW percentage in the
        # final years both draws down the portfolio and inflates
        # Base_Expenses to match, whether or not that much would really be
        # consumed. Applied here, BEFORE total_exp/the draw waterfall are
        # built below, so anything above the ceiling simply stays invested
        # instead of being drawn at all -- not drawn-then-unaccounted-for.
        # A no-op for Fixed Real Spending, which never exceeds its own
        # planned amount. Off (0) by default at the engine level; the UI
        # defaults it to your Annual Expenses figure.
        max_exp = cfg.get("max_annual_expenses", 0.0)
        max_exp_inflated = max_exp * infl_factor_yir if max_exp > 0 else 0.0
        row["Max_Annual_Expenses_Inflated"] = max_exp_inflated
        row["Max_Expenses_Bound_Hit"] = bool(max_exp > 0 and base_exp > max_exp_inflated)
        if max_exp > 0:
            base_exp = min(base_exp, max_exp_inflated)
        healthcare = (cfg["healthcare_pre_medicare"] if age < 65 else cfg["healthcare_post_medicare"]) * (1 + cfg["healthcare_inflation"]) ** yir
        hdhp = cfg["hdhp_annual"] * infl_factor_yir if age < 65 else 0.0
        gifts = cfg["gifts_annual"] * infl_factor_yir
        lump = cfg.get("lump_sums", {}).get(age, 0.0)
        legacy_target_per_child = 0.0
        legacy_target_total = 0.0
        if yir < cfg.get("legacy_years", 10) and cfg.get("num_children", 0) > 0:
            legacy_target_per_child = cfg["roth_legacy_per_child"] * (1 + cfg["legacy_inflation"]) ** yir
            legacy_target_total = cfg["num_children"] * legacy_target_per_child

        # ── DISCRETIONARY REDUCTION ON NEGATIVE RETURN ──
        # Moved ahead of total_exp (was after) so the legacy-funding decision
        # right below can use it.
        # Use pretax return as proxy for overall market return in that year
        is_negative_return = False
        if return_overrides and yir < len(return_overrides["pretax"]):
            if return_overrides["pretax"][yir] < 0:
                is_negative_return = True
        elif cfg["pretax_return"] < 0:
            is_negative_return = True

        # Only fund the legacy gift out of this year's cash flow when it will
        # actually be deposited into Roth below (i.e. not a down-market year).
        # Previously legacy_target_total was added to total_exp
        # *unconditionally*, so the draw engine withdrew it from real accounts
        # as "spending" even in years the code elsewhere skipped the actual
        # Roth gift -- the money vanished without ever landing anywhere.
        # Skipping it here means it simply stays invested and becomes
        # ordinary inheritance later (tracked below, not a separate pot).
        legacy_funding_this_year = legacy_target_total if (legacy_target_total > 0 and not is_negative_return) else 0.0

        total_exp = base_exp + healthcare + hdhp + gifts + lump + legacy_funding_this_year

        if is_negative_return:
            # v9: was a fixed nominal dollar amount for the whole plan while
            # base_exp inflates every year, so the same-looking cut had a
            # shrinking real bite over time (11.9% of base expenses in year 1
            # down to 5.1% by year 30 in a typical scenario). The input is
            # entered in today's dollars and now inflates alongside
            # everything else it's meant to offset.
            reduction = cfg.get("neg_ret_draw_reduction", 0.0) * infl_factor_yir
            # Cannot reduce expenses below 0, but total_exp should be large enough
            reduction_applied = min(reduction, total_exp)
            total_exp -= reduction_applied
            row["Discretionary_Reduction"] = reduction_applied
        else:
            row["Discretionary_Reduction"] = 0.0

        row["Base_Expenses"], row["Healthcare_Cost"], row["HDHP"] = base_exp, healthcare, hdhp
        row["Gifts"], row["Lump_Sum"], row["Legacy_Roth"] = gifts, lump, legacy_funding_this_year
        row["Legacy_Target_Per_Child"] = legacy_target_per_child
        row["Legacy_Target_Total"] = legacy_target_total
        row["Total_Expenses"] = total_exp

        passive = sp_inc + jss_inc + ss_inc + rent_inc
        row["Passive_Income"] = passive

        rmd = get_rmd(pt_prior_year_end, age, cfg["rmd_start_age"])
        row["RMD"] = rmd

        # ── BRACKET-OPTIMIZED DRAW STRATEGY ──
        ptd = rod = hsd = cad = bkd = 0.0
        need = max(0.0, total_exp - passive)
        std_ded_est = cfg["standard_deduction"] * std_ded_multiplier * (1 + binfl) ** yfb
        base_taxable = jss_tax + rent_tax + sp_inc
        sst_est = ss_taxable_portion(ss_inc, base_taxable, single=is_widowed)
        existing_taxable = base_taxable + sst_est
        br12_gross = bracket_ceiling(fed_brackets_yr, 0.12, yfb, binfl)
        pretax_room_12 = max(0.0, br12_gross + std_ded_est - existing_taxable)
        is_rmd_phase = age >= cfg["rmd_start_age"]

        if is_rmd_phase:
            if pt > 0: ptd = min(rmd, pt); need = max(0.0, need - ptd)
            if need > 0 and ca > 0: d = min(need, ca); cad += d; need -= d
            if need > 0 and bk > 0: d = min(need, bk); bkd += d; need -= d
            if need > 0 and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "RMD-Dominated"
        elif cfg.get("brokerage_before_pretax", False):
            # Brokerage-first: spending draws from Cash/Brokerage before
            # PreTax. Brokerage draws are only taxed on the LTCG portion (see
            # brokerage_pref_income below), so -- unlike a PreTax draw --
            # they don't consume any of this year's ordinary-income bracket
            # room. That leaves the FULL bracket target free for the Roth
            # Conversion section further down, instead of the spending draw
            # and the conversion competing for the same limited room (which
            # is what happens in the default order below: PreTax drawn for
            # spending fills the bracket first, often leaving ~$0 of
            # conversion room in the same year). Empirically verified
            # (scratch Monte Carlo comparison, matched seeds): raises success
            # rate by several points in a typical scenario, entirely because
            # more ends up converted to Roth, not because of the account
            # balances' own growth.
            if need > 0 and ca > 0:
                d = min(need, ca); cad += d; need -= d
            if need > 0 and bk > 0: d = min(need, bk); bkd += d; need -= d
            if need > 0 and pt > 0:
                d = min(need, pretax_room_12, pt); ptd += d; need -= d
            if need > 0 and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "Bracket-Optimized (Brokerage-First)"
        else:
            if need > 0 and pt > 0:
                d = min(need, pretax_room_12, pt); ptd += d; need -= d
            if need > 0 and ca > 0:
                d = min(need, ca); cad += d; need -= d
            if need > 0 and bk > 0: d = min(need, bk); bkd += d; need -= d
            if need > 0 and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "Bracket-Optimized"

        if need > 0:
            for nm, avail in [("pretax", pt-ptd), ("cash", ca-cad), ("brokerage", bk-bkd),
                               ("roth", ro-rod), ("hsa", hs-hsd)]:
                if need <= 0: continue
                d = min(need, max(0.0, avail))
                if nm == "pretax": ptd += d
                elif nm == "cash": cad += d
                elif nm == "brokerage": bkd += d
                elif nm == "roth": rod += d
                elif nm == "hsa": hsd += d
                need -= d

        # Brokerage draws realize a pro-rated long-term capital gain off the
        # account's running cost basis -- a simplification (no lot-level
        # detail), but it captures the real economics: contributed dollars
        # come back tax-free, growth comes back as a taxed gain. basis_removed
        # is applied to bk_basis in UPDATE BALANCES below. (bk_basis here
        # already reflects this year's dividend reinvestment, computed above.)
        if bkd > 0 and bk > 0:
            basis_fraction = min(1.0, bk_basis / bk)
            brokerage_ltcg_gain = bkd * (1 - basis_fraction)
            basis_removed = bkd - brokerage_ltcg_gain
        else:
            brokerage_ltcg_gain = 0.0
            basis_removed = 0.0
        row["Brokerage_LTCG_Gain"] = brokerage_ltcg_gain
        # Qualified dividends and net LTCG are taxed together at the same
        # preferential rates (same IRS worksheet) -- combine them into one
        # pool for the stacking math below, while keeping the two reported
        # separately above/below for transparency about where each came from.
        brokerage_pref_income = brokerage_ltcg_gain + brokerage_dividend_taxable

        row["PreTax_Draw"], row["Roth_Draw"] = ptd, rod
        row["HSA_Draw"], row["Cash_Draw"], row["Brokerage_Draw"] = hsd, cad, bkd
        row["RMD_Excess"] = max(0.0, rmd - (total_exp - passive)) if is_rmd_phase else 0.0

        # ── TAX ──
        other_taxable = ptd + jss_tax + rent_tax + sp_inc + cash_interest_taxable + brokerage_pref_income
        sst = ss_taxable_portion(ss_inc, other_taxable, single=is_widowed)
        gross_taxable = other_taxable + sst
        std_ded = cfg["standard_deduction"] * std_ded_multiplier * (1 + binfl) ** yfb
        med_ded = max(0.0, healthcare - 0.075 * gross_taxable)
        # NOTE: the non-taxable portion of JSS income (jss_inc - jss_tax) was
        # previously added here too. That income was never included in
        # gross_taxable to begin with (other_taxable only adds jss_tax), so
        # deducting it again created a phantom deduction that understated tax
        # every year JSS was partially or fully non-taxable.
        item_ded = med_ded + hdhp
        best_ded = max(std_ded, item_ded)
        fed_taxable = max(0.0, gross_taxable - best_ded)
        # Federal taxes long-term capital gains at preferential rates,
        # "stacked on top of" ordinary income (IRS Qualified Dividends and
        # Capital Gain Tax Worksheet method): ordinary_taxable is fed_taxable
        # with the LTCG gain backed out, taxed at ordinary rates; the gain
        # itself is taxed via ltcg_brackets_yr starting from that floor, so
        # it can be pushed into a higher LTCG bracket by other income (e.g.
        # a large Roth conversion), same as in real life. When there's no
        # brokerage gain this reduces to the old calc_tax(fed_taxable, ...).
        # brokerage_pref_income = LTCG gain realized on this year's draw PLUS
        # this year's dividend distribution -- both preferential-rate income.
        ordinary_taxable = max(0.0, fed_taxable - brokerage_pref_income)
        fed_tax_ordinary = calc_tax(ordinary_taxable, fed_brackets_yr, yfb, binfl)
        fed_ltcg_tax = calc_ltcg_tax(brokerage_pref_income, ordinary_taxable, ltcg_brackets_yr, yfb, binfl)
        fed_tax = fed_tax_ordinary + fed_ltcg_tax
        # Oregon fully exempts Social Security. The only SS-related dollars
        # ever present in gross_taxable are `sst` (other_taxable has no raw
        # ss_inc term), so subtracting `sst` alone fully backs SS out of the
        # Oregon base. The old formula also subtracted an extra ss_inc*0.15
        # that was never part of gross_taxable, understating Oregon tax.
        or_taxable = max(0.0, gross_taxable - sst - best_ded)
        or_tax = calc_tax(or_taxable, OREGON_BRACKETS, yfb, binfl) if cfg["oregon_resident"] else 0.0
        total_tax = fed_tax + or_tax

        row["Gross_Taxable_Income"], row["Federal_Taxable_Income"] = gross_taxable, fed_taxable
        row["Standard_Deduction"], row["Itemized_Deduction"], row["Best_Deduction"] = std_ded, item_ded, best_ded
        row["Deduction_Type"] = "Itemized" if item_ded > std_ded else "Standard"
        row["Federal_Tax"], row["Oregon_Tax"], row["Total_Tax"] = fed_tax, or_tax, total_tax
        row["Effective_Tax_Rate"] = total_tax / gross_taxable if gross_taxable > 0 else 0.0

        # magi_pre_conversion excludes this year's not-yet-decided Roth
        # conversion -- used below only to size how much conversion room is
        # left before IRMAA avoidance kicks in. Final MAGI (which the
        # conversion itself counts toward, since it's ordinary taxable
        # income) is computed after the conversion amount is known, right
        # after the Roth Conversion section below.
        magi_pre_conversion = gross_taxable + (ss_inc - sst)
        irmaa_thr = irmaa_threshold_base * (1 + binfl) ** yfb

        # ── ROTH CONVERSION ──
        roth_conv_amt = roth_conv_tax = 0.0
        if cfg["roth_conversion_enabled"] and pt - ptd > cfg["roth_conversion_margin"]:
            tgt = {"12%": 0.12, "22%": 0.22, "24%": 0.24}.get(cfg["roth_conversion_target_bracket"], 0.22)
            conv_room = max(0.0, bracket_ceiling(fed_brackets_yr, tgt, yfb, binfl) - fed_taxable)
            if cfg["irmaa_avoidance"] and age >= 63:
                conv_room = min(conv_room, max(0.0, irmaa_thr - magi_pre_conversion))
            roth_conv_amt = min(conv_room, max(0.0, pt - ptd - cfg["roth_conversion_margin"]))
            if roth_conv_amt > 0:
                # Conversion income is ordinary and stacks BELOW the LTCG
                # gain (a bigger conversion can push existing gains into a
                # higher LTCG bracket) -- recompute the full ordinary+LTCG
                # stack with the conversion added rather than taxing it at a
                # flat marginal ordinary rate, to capture that knock-on
                # effect. Reduces to the old formula when there's no gain.
                ord_with_conv = ordinary_taxable + roth_conv_amt
                fed_tax_with_conv = (calc_tax(ord_with_conv, fed_brackets_yr, yfb, binfl)
                                      + calc_ltcg_tax(brokerage_pref_income, ord_with_conv, ltcg_brackets_yr, yfb, binfl))
                fc = fed_tax_with_conv - fed_tax
                if cfg["oregon_resident"]:
                    # Was a flat 9% guess, inconsistent with the marginal-bracket
                    # approach used for federal (and not matching any actual
                    # OREGON_BRACKETS rate at typical conversion income levels).
                    oc = calc_tax(or_taxable + roth_conv_amt, OREGON_BRACKETS, yfb, binfl) - or_tax
                else:
                    oc = 0.0
                roth_conv_tax = fc + oc
        row["Roth_Conversion"], row["Roth_Conversion_Tax"] = roth_conv_amt, roth_conv_tax

        # ── MAGI / IRMAA ──
        # Final MAGI includes this year's Roth conversion -- it's ordinary
        # taxable income like any other dollar in gross_taxable, and
        # evaluating IRMAA off magi_pre_conversion instead would silently
        # hide the IRMAA impact of any conversion run with IRMAA avoidance
        # turned off (avoidance sizes the conversion to stay under the
        # threshold; without it, nothing stops a large conversion from
        # blowing straight through).
        magi = magi_pre_conversion + roth_conv_amt
        # Real Medicare IRMAA is assessed off MAGI from 2 TAX YEARS PRIOR
        # (the SSA/CMS lookback), compared against the current year's
        # threshold/premium schedule -- not this year's own MAGI. A big Roth
        # conversion this year raises IRMAA 2 years from now, not
        # immediately; conversely this year's premium was already locked in
        # by a decision made 2 years ago. `results` already holds every
        # completed prior year's row by the time we get here, so the
        # lookback is just an index, not separate bookkeeping. Falls back to
        # this year's own MAGI for the first two retirement years, since the
        # model doesn't track MAGI before retirement (the accumulation phase
        # isn't tax-modeled at all -- a broader simplification, not specific
        # to IRMAA).
        lookback_magi = results[idx - 2]["MAGI"] if idx >= 2 else magi
        irmaa_tier_mult = irmaa_tier_multiplier(lookback_magi, irmaa_threshold_base, yfb, binfl) if age >= 65 else 0.0
        irmaa_hit = irmaa_tier_mult > 0
        irmaa_people = 1 if is_widowed else 2  # only one person on Medicare once widowed
        # Deliberately still the flat rate here, not infl_factor_yir: this
        # scales off yfb (years from CURRENT_YEAR, the same basis as the
        # bracket-inflated threshold above it), not yir (years from
        # retirement, what infl_factor_yir/cum_infl are indexed by) -- the
        # two counters diverge whenever retirement is years away, and
        # stochastic inflation before retirement isn't modeled at all
        # (accumulation isn't tax-modeled, so there's no path to draw from
        # for those years anyway).
        irmaa_cost = IRMAA_MONTHLY_SURCHARGE * irmaa_tier_mult * 12 * irmaa_people * (1 + infl) ** yfb
        row["MAGI"], row["IRMAA_Threshold"] = magi, irmaa_thr
        row["IRMAA_Lookback_MAGI"], row["IRMAA_Tier_Multiplier"] = lookback_magi, irmaa_tier_mult
        row["IRMAA_Hit"], row["IRMAA_Cost"] = irmaa_hit, irmaa_cost

        # ── NET INVESTMENT INCOME TAX (NIIT) ──
        # 3.8% surtax on the LESSER of net investment income or the excess of
        # (final, post-conversion) MAGI over the threshold. Net investment
        # income here is interest + brokerage LTCG/dividends + rental income
        # -- retirement account distributions (PreTax_Draw, RMD) and Roth
        # conversions are excluded from NII by statute, even though the
        # conversion DOES count toward the MAGI threshold test (a large
        # conversion can trigger or enlarge NIIT exposure without itself
        # being part of what's taxed by it).
        niit_threshold = NIIT_MAGI_THRESHOLD_SINGLE if is_widowed else NIIT_MAGI_THRESHOLD
        net_investment_income = cash_interest_taxable + brokerage_pref_income + rent_tax
        niit = max(0.0, min(net_investment_income, magi - niit_threshold)) * NIIT_RATE
        row["Net_Investment_Income"], row["NIIT"] = net_investment_income, niit

        # ── UPDATE BALANCES ──
        pt -= (ptd + roth_conv_amt); ro -= rod; ro += roth_conv_amt; hs -= hsd; ca -= cad
        bk -= bkd; bk_basis -= basis_removed
        total_income = passive + ptd + rod + hsd + cad + bkd
        surplus = total_income - total_exp - total_tax - roth_conv_tax - irmaa_cost - niit
        ca += surplus

        legacy_roth_per_child = 0.0
        legacy_inheritance_per_child = 0.0
        legacy_roth_total = 0.0
        legacy_inheritance_total = 0.0
        legacy_actual = 0.0
        if legacy_target_total > 0 and cfg["num_children"] > 0:
            if legacy_funding_this_year > 0:
                # legacy_funding_this_year was already withdrawn from the
                # household's accounts above via total_exp/need (single
                # debit, spread across pt/ca/ro/hs by the normal draw
                # waterfall). It now moves into the segregated legacy_pool --
                # NOT back into the parents' own `ro` -- because once gifted
                # it's legally the kids' money, in the kids' own Roth
                # accounts, no longer part of the household's spendable net
                # worth and never available for the household's own future
                # withdrawals. (Previously this was credited back into `ro`,
                # which meant it was silently drawn back down years later by
                # the household's own Roth_Draw waterfall -- verified: with
                # no fix, gifted balances were being spent on the parents'
                # own retirement expenses in later years.)
                legacy_roth_total = legacy_funding_this_year
                legacy_roth_per_child = legacy_target_per_child
                legacy_pool += legacy_roth_total
                legacy_actual = legacy_roth_total
            # Any portion not funded to Roth this year (skipped in a
            # down-market year) was never withdrawn -- it's still sitting in
            # the household's own balances. This is a pure label for "will
            # pass as ordinary inheritance," not a separate pot of money.
            legacy_inheritance_per_child = max(0.0, legacy_target_per_child - legacy_roth_per_child)
            legacy_inheritance_total = legacy_inheritance_per_child * cfg["num_children"]

        cum_gifts += gifts
        cum_legacy_roth += legacy_roth_total
        cum_legacy_inheritance += legacy_inheritance_total
        cum_lump_sums += lump
        n_kids = cfg.get("num_children", 0)
        row["Surplus_Deficit"], row["Total_Income"] = surplus, total_income
        row["Bad_Return_Year"] = is_negative_return
        row["Legacy_Roth_Per_Child"] = legacy_roth_per_child
        row["Legacy_Inheritance_Per_Child"] = legacy_inheritance_per_child
        row["Legacy_Roth_Total"] = legacy_roth_total
        row["Legacy_Inheritance_Total"] = legacy_inheritance_total
        row["Legacy_Total"] = legacy_roth_total + legacy_inheritance_total  # this year's FLOW only, not a running total
        row["Legacy_Roth_Actual"] = legacy_actual
        row["Cum_Gifts"] = cum_gifts
        row["Cum_Legacy_Roth"] = cum_legacy_roth  # nominal sum of contributions made (no growth)
        row["Legacy_Pool_EOY"] = legacy_pool       # actual compounding balance of gifted money
        row["Legacy_Pool_Per_Child"] = (legacy_pool / n_kids) if n_kids else 0.0
        row["Cum_Legacy_Inheritance"] = cum_legacy_inheritance
        row["Cum_Legacy_Inheritance_Per_Child"] = (cum_legacy_inheritance / n_kids) if n_kids else 0.0
        # Running total legacy VALUE as of this point in time (grown Roth pool
        # + amounts still sitting in the household's own accounts earmarked
        # for inheritance). This is what "value at the end of the plan"
        # should read -- previously that metric read Legacy_Total (this
        # year's flow), which is 0 in nearly every year outside the
        # legacy_years gifting window, including almost always the final
        # simulated year of a multi-decade plan.
        row["Legacy_Value_To_Date"] = legacy_pool + cum_legacy_inheritance
        row["Legacy_Value_To_Date_Per_Child"] = row["Legacy_Pool_Per_Child"] + row["Cum_Legacy_Inheritance_Per_Child"]
        row["Cum_Lump_Sums"] = cum_lump_sums

        # Retirement/HSA accounts can't structurally go negative (draws are
        # already capped at the available balance upstream); this is just a
        # defensive floor for floating-point noise near zero.
        pt = max(0.0, pt)
        ro = max(0.0, ro)
        hs = max(0.0, hs)
        bk = max(0.0, bk)
        bk_basis = max(0.0, min(bk_basis, bk))  # basis can never exceed balance
        legacy_pool = max(0.0, legacy_pool)
        # Cash IS allowed to go negative -- it represents a genuine funding
        # shortfall (expenses the plan couldn't cover from any account).
        # Previously this was floored at $0 in both Cash_EOY and
        # Total_Liquid_Assets, which hid insolvency: a plan running a large,
        # growing cash deficit could still show a "positive" total and get
        # counted as a Monte Carlo success just because pt/ro/hs were still
        # positive. Now the deficit is shown and counted honestly.

        row["PreTax_EOY"], row["Roth_EOY"], row["HSA_EOY"] = pt, ro, hs
        row["Cash_EOY"] = ca
        row["Brokerage_EOY"], row["Brokerage_Basis"] = bk, bk_basis
        total_liquid = pt + ro + hs + ca + bk
        row["Total_Liquid_Assets"] = total_liquid  # parents' own spendable net worth (excludes gifted-away legacy_pool)
        row["Family_Net_Worth"] = total_liquid + legacy_pool  # includes money already gifted to the kids' Roths
        # "If death occurs at this age, what would each child actually
        # receive?" -- this is what Cum_Legacy_Inheritance does NOT capture
        # (that field only tracks a narrow sub-bucket: legacy-gift TARGET
        # amounts skipped in down-market years). The real answer is: your
        # own remaining accounts (Total_Liquid_Assets) pass to your heirs at
        # death, split evenly here across children, PLUS whatever's already
        # sitting in their own Roth accounts from prior gifting.
        row["Estate_At_Death"] = total_liquid  # what's left in your own accounts if you died this year
        row["Estate_At_Death_Per_Child"] = (total_liquid / n_kids) if n_kids else 0.0
        row["Total_Inheritance_At_Death_Per_Child"] = (row["Family_Net_Worth"] / n_kids) if n_kids else 0.0

        # Oregon Estate Transfer Tax: a $1M exemption (vs. federal's ~$15M,
        # not binding for anything this model would flag) makes this the
        # actual estate-tax exposure for an Oregon resident, and it's easy to
        # cross given the account balances/contributions this model deals in.
        # Applied to total_liquid -- the full date-of-death value of the
        # household's own accounts, BEFORE the heir's own income tax on
        # inherited pretax/HSA below; estate tax and the heir's income tax
        # are separate, both apply, and neither offsets the other in real
        # law. Paid by the estate before distribution, modeled as a lump
        # deduction off the top rather than pro-rated across account types.
        # Simplification: real law exempts transfers to a surviving spouse
        # entirely (unlimited marital deduction), so this overstates exposure
        # at a first death for a still-married couple -- treat this as the
        # exposure at the SECOND (surviving) death, which is what actually
        # reaches the kids either way.
        oregon_estate_tax = calc_tax(total_liquid, OREGON_ESTATE_BRACKETS, 0, 0.0) if cfg["oregon_resident"] else 0.0
        row["Oregon_Estate_Tax"] = oregon_estate_tax

        # After-tax value to heirs: not all dollars in the estate are worth
        # the same to your kids. Inherited pretax accounts (and HSAs, which
        # follow their own less-favorable rule) must be fully distributed
        # within 10 years under the SECURE Act and are taxed at the HEIR'S
        # own bracket -- unlike Roth, which passes tax-free. Cash is already
        # after-tax since its interest is taxed annually (see Cash_Interest_Taxable).
        heir_rate = cfg.get("heir_tax_rate", 0.24)
        pretax_after_tax_to_heirs = pt * (1 - heir_rate)
        hsa_after_tax_to_heirs = hs * (1 - heir_rate)  # HSAs lose tax-free status entirely for non-spouse heirs
        row["Heir_Tax_On_PreTax"] = pt * heir_rate
        row["Heir_Tax_On_HSA"] = hs * heir_rate
        # Brokerage gets a full step-up in cost basis at death (IRC S1014):
        # heirs owe no tax on gains accrued before death, so -- like Roth
        # and cash -- it passes at full value, not haircut by heir_rate.
        after_tax_estate = pretax_after_tax_to_heirs + hsa_after_tax_to_heirs + ro + ca + bk - oregon_estate_tax
        row["After_Tax_Estate_At_Death"] = after_tax_estate
        row["After_Tax_Family_Net_Worth"] = after_tax_estate + legacy_pool  # legacy pool is Roth -- already tax-free
        row["After_Tax_Estate_Per_Child"] = (after_tax_estate / n_kids) if n_kids else 0.0
        row["After_Tax_Total_Inheritance_Per_Child"] = (row["After_Tax_Family_Net_Worth"] / n_kids) if n_kids else 0.0

        row["Total_Real"] = (total_liquid / infl_factor_yir) if yir > 0 else total_liquid
        # Today's-dollar view of actual lifestyle spending -- lets the
        # Monte Carlo tab show what a "successful" path's spending
        # experience actually looked like, not just whether the portfolio
        # survived.
        row["Base_Expenses_Real"] = (base_exp / infl_factor_yir) if yir > 0 else base_exp

        total_draws = ptd + rod + hsd + cad + bkd
        row["Total_Draws"] = total_draws
        row["Withdrawal_Rate"] = total_draws / total_liquid if total_liquid > 0 else 0.0

        # br12_gross already computed above (same fed_brackets_yr/yfb/binfl
        # this year) -- reused here instead of recomputing.
        br22 = bracket_ceiling(fed_brackets_yr, 0.22, yfb, binfl)
        row["Bracket_12_Ceiling"], row["Bracket_22_Ceiling"] = br12_gross, br22

        results.append(row)

    return accum_rows, pd.DataFrame(results)


# ============================================================
# MONTE CARLO
# ============================================================

def run_monte_carlo(cfg, n_sims, std_dev, max_up, seed=None, correlation=0.85, fat_tailed=False, t_df=5):
    rng = np.random.default_rng(seed)
    num_years = cfg["planning_end_age"] - cfg["retirement_age"] + 1
    # Accumulation is identical across every path (see run_simulation) --
    # compute it once instead of n_sims times.
    accum_result = run_accumulation(cfg)
    mortality_on = cfg.get("mortality_based_horizon", False)
    all_runs = []
    for _ in range(n_sims):
        ov = build_mc_return_overrides(cfg, num_years, rng, std_dev, max_up, correlation, fat_tailed, t_df)
        _, df = run_simulation(cfg, return_overrides=ov, accum_result=accum_result)
        if mortality_on:
            # "Plan Through Age" doubles as the mortality ceiling here -- the
            # portfolio is still simulated all the way out to it (so the
            # percentile-band charts keep showing the full trajectory), but
            # success/outcome metrics (see mc_outcome_row) now read the row
            # at THIS path's own randomly drawn death age instead of always
            # the final row, since most paths won't actually need every
            # simulated year.
            df["Death_Age"] = draw_household_death_age(cfg["current_age"], cfg["planning_end_age"], rng)
        all_runs.append(df)
    return all_runs

def mc_outcome_row(run_df):
    """The row that determines a Monte Carlo path's outcome: the final
    simulated row normally, or -- when mortality_based_horizon drew a
    Death_Age for this path -- the row at that (possibly earlier) age
    instead, since the plan didn't need to survive any further than that."""
    if "Death_Age" in run_df.columns:
        matches = run_df[run_df["Age"] == run_df["Death_Age"].iloc[0]]
        if len(matches) > 0:
            return matches.iloc[0]
    return run_df.iloc[-1]

def compute_percentile_bands(runs, col, pcts=(5, 25, 50, 75, 95)):
    ages = runs[0]["Age"].values
    mx = np.array([r[col].values for r in runs])
    bands = {f"p{p}": np.percentile(mx, p, axis=0) for p in pcts}
    bands["mean"] = np.mean(mx, axis=0); bands["Age"] = ages
    return bands


# ============================================================
# SCENARIO OPTIMIZER
# ============================================================

def run_optimizer(base_cfg, mc_sims, mc_std, mc_max, mc_seed, correlation=0.85, fat_tailed=False, t_df=5):
    """
    Sweep retirement_age × ss_start_age × expense_reduction_post80
    and find combinations that achieve 100% MC success rate.
    Returns a DataFrame of all scenarios sorted by success rate.
    """
    rng_base = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)

    # Define search space
    ret_ages = list(range(base_cfg["retirement_age"], min(base_cfg["retirement_age"] + 4, 64)))
    ss_ages = list(range(max(62, base_cfg["ss_start_age"]), min(base_cfg["ss_start_age"] + 5, 71)))
    reductions = [0.0, 0.05, 0.10, 0.15, 0.20, 0.25]

    # Use fewer sims for optimizer (speed) — 50 is enough for ranking
    opt_sims = min(mc_sims, 50)
    num_years_max = base_cfg["planning_end_age"] - min(ret_ages) + 1

    results = []
    total_combos = len(ret_ages) * len(ss_ages) * len(reductions)

    # Accumulation only depends on retirement_age among the swept variables
    # (ss_start_age/expense_reduction_post80 never touch it) -- compute it
    # once per unique retirement_age instead of once per combo per sim
    # (len(ret_ages) calls instead of up to total_combos * opt_sims).
    accum_cache = {
        ret_age: run_accumulation({**base_cfg, "retirement_age": ret_age})
        for ret_age in ret_ages
    }

    for ret_age in ret_ages:
        for ss_age in ss_ages:
            if ss_age < ret_age:
                continue  # Can't claim SS before retirement in this model
            for red in reductions:
                test_cfg = {**base_cfg,
                            "retirement_age": ret_age,
                            "ss_start_age": ss_age,
                            "expense_reduction_post80": red}

                # Generate MC return paths (reuse same seed for fair comparison)
                rng = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)
                num_years = test_cfg["planning_end_age"] - ret_age + 1
                survived = 0
                final_assets = []
                final_real = []

                for _ in range(opt_sims):
                    ov = build_mc_return_overrides(test_cfg, num_years, rng, mc_std, mc_max, correlation, fat_tailed, t_df)
                    _, sim_df = run_simulation(test_cfg, return_overrides=ov, accum_result=accum_cache[ret_age])
                    last = sim_df.iloc[-1]
                    if last["Total_Liquid_Assets"] > 0:
                        survived += 1
                    final_assets.append(last["Total_Liquid_Assets"])
                    final_real.append(last["Total_Real"])

                success_rate = survived / opt_sims
                results.append({
                    "Retire_Age": ret_age,
                    "SS_Age": ss_age,
                    "Expense_Cut_80+": red,
                    "Success_Rate": success_rate,
                    "Median_Final_Assets": np.median(final_assets),
                    "Median_Final_Real": np.median(final_real),
                    "P10_Final_Assets": np.percentile(final_assets, 10),
                    "Worst_Case": min(final_assets),
                    "Extra_Work_Years": ret_age - base_cfg["retirement_age"],
                    "SS_Delay_Years": ss_age - base_cfg["ss_start_age"],
                })

    opt_df = pd.DataFrame(results).sort_values(
        ["Success_Rate", "Median_Final_Real"], ascending=[False, False]
    ).reset_index(drop=True)
    return opt_df


# ============================================================
# SENSITIVITY / TORNADO ANALYSIS
# ============================================================

# (label, cfg_key, mode, delta) -- mode is "additive" (rates/percentages,
# delta in absolute terms), "multiplicative" (dollar amounts, delta as a
# fraction), or "additive_int" (whole-number values like age).
SENSITIVITY_VARS = [
    ("Retirement Age", "retirement_age", "additive_int", 2),
    ("PreTax Return", "pretax_return", "additive", 0.015),
    ("Roth Return", "roth_return", "additive", 0.015),
    ("HSA Return", "hsa_return", "additive", 0.015),
    ("Cash/MM Return", "cash_return", "additive", 0.01),
    ("Brokerage Return", "brokerage_return", "additive", 0.015),
    ("Brokerage Dividend Yield", "brokerage_dividend_yield", "additive", 0.01),
    ("General Inflation", "inflation_rate", "additive", 0.01),
    ("Healthcare Inflation", "healthcare_inflation", "additive", 0.02),
    ("Base Annual Expenses", "base_annual_expenses", "multiplicative", 0.15),
    ("SS Annual Amount", "ss_annual_amount", "multiplicative", 0.15),
    ("Legacy Roth/Child/Year", "roth_legacy_per_child", "multiplicative", 0.30),
]

def run_sensitivity_analysis(cfg, variables=None):
    """
    Vary ONE assumption at a time (holding everything else at baseline),
    re-run the deterministic simulation, and record how much the final-year
    real (today's-dollar) value moves. Returns a DataFrame sorted by impact
    magnitude (smallest first, so the biggest driver plots at the top of a
    horizontal tornado chart), plus the baseline result for reference.
    """
    variables = variables or SENSITIVITY_VARS
    _, base_df = run_simulation(cfg)
    base_final = base_df["Total_Real"].iloc[-1]

    results = []
    for label, key, mode, delta in variables:
        base_val = cfg.get(key)
        if base_val is None:
            continue
        if mode == "additive":
            lo_val, hi_val = max(0.0, base_val - delta), base_val + delta
        elif mode == "additive_int":
            lo_val = max(cfg.get("current_age", 45), int(base_val) - delta)
            hi_val = min(70, int(base_val) + delta)
        else:  # multiplicative
            lo_val, hi_val = base_val * (1 - delta), base_val * (1 + delta)

        _, df_lo = run_simulation({**cfg, key: lo_val})
        _, df_hi = run_simulation({**cfg, key: hi_val})
        final_lo = df_lo["Total_Real"].iloc[-1]
        final_hi = df_hi["Total_Real"].iloc[-1]

        results.append({
            "Variable": label,
            "Base_Value": base_val, "Low_Value": lo_val, "High_Value": hi_val,
            "Low_Result": min(final_lo, final_hi), "High_Result": max(final_lo, final_hi),
            "Low_Is_High_Input": final_lo > final_hi,  # True if raising the input LOWERS the outcome
            "Base_Result": base_final,
            "Impact_Range": abs(final_hi - final_lo),
        })

    result_df = pd.DataFrame(results).sort_values("Impact_Range", ascending=True).reset_index(drop=True)
    return result_df, base_final


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_to_excel(accum_df, retire_df, cfg, mc_runs=None):
    """Create a professional Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    accum_fill = PatternFill("solid", fgColor="4472C4")
    input_font = Font(name="Arial", color="0000FF", size=10)  # Blue = inputs
    num_font = Font(name="Arial", size=10)
    pct_fmt = '0.0%'
    money_fmt = '$#,##0;($#,##0);"-"'
    money_k = '$#,##0'
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    def style_header(ws, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            mx = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 2, 10), 22)

    # ──────── Sheet 1: Assumptions ────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    assumptions = [
        ("RETIREMENT PLANNING MODEL", ""),
        ("", ""),
        ("TIMELINE", ""),
        ("Current Age", cfg["current_age"]),
        ("Retirement Age", cfg["retirement_age"]),
        ("Plan Through Age", cfg["planning_end_age"]),
        ("", ""),
        ("STARTING BALANCES (Today)", ""),
        ("PreTax 401(k)", cfg["pretax_401k"]),
        ("Roth IRA", cfg["roth_ira"]),
        ("HSA", cfg["hsa"]),
        ("S+ 5-Year", cfg["s_plus_5yr"]),
        ("S+ 10-Year", cfg["s_plus_10yr"]),
        ("Cash", cfg["cash"]),
        ("Taxable Brokerage", cfg.get("brokerage", 0)),
        ("Taxable Brokerage Cost Basis", cfg.get("brokerage_basis", cfg.get("brokerage", 0))),
        ("", ""),
        ("ANNUAL CONTRIBUTIONS (Working Years)", ""),
        ("401(k) Employee", cfg.get("contrib_401k", 24_500)),
        ("Roth 401(k)", cfg.get("contrib_roth401k", 8_000)),
        ("Roth IRA (per spouse)", cfg.get("contrib_roth_ira", 8_700)),
        ("HSA (family)", cfg.get("contrib_hsa", 8_700)),
        ("Mega Backdoor Roth", cfg.get("contrib_mega_backdoor", 29_000)),
        ("Employer Match", cfg.get("contrib_employer_match", 18_000)),
        ("Cash Savings", cfg.get("contrib_cash_annual", 150_000)),
        ("Final Year Cash Lump", cfg.get("contrib_cash_final_lump", 120_000)),
        ("Taxable Brokerage Savings", cfg.get("contrib_brokerage_annual", 0)),
        ("", ""),
        ("PERFORMANCE", ""),
        ("PreTax Return", cfg["pretax_return"]),
        ("Roth Return", cfg["roth_return"]),
        ("HSA Return", cfg["hsa_return"]),
        ("Cash Return", cfg["cash_return"]),
        ("Taxable Brokerage Return", cfg.get("brokerage_return", cfg["roth_return"])),
        ("", ""),
        ("INCOME SOURCES", ""),
        ("SS Start Age", cfg["ss_start_age"]),
        ("SS Annual (future $)", cfg["ss_annual_amount"]),
        ("SS COLA", cfg["ss_cola"]),
        ("JSS Start Age", cfg["jss_start_age"]),
        ("JSS Annual", cfg["jss_annual_amount"]),
        ("JSS COLA", cfg["jss_cola"]),
        ("Rental Gross", cfg["rental_gross"]),
        ("", ""),
        ("EXPENSES & INFLATION", ""),
        ("Annual Expenses", cfg["base_annual_expenses"]),
        ("General Inflation", cfg["inflation_rate"]),
        ("Healthcare Pre-Medicare", cfg["healthcare_pre_medicare"]),
        ("Healthcare Post-Medicare", cfg["healthcare_post_medicare"]),
        ("Healthcare Inflation", cfg["healthcare_inflation"]),
        ("", ""),
        ("STRATEGY", ""),
        ("RMD Start Age", cfg["rmd_start_age"]),
        ("Roth Conversion Target", cfg["roth_conversion_target_bracket"]),
        ("IRMAA Avoidance", cfg["irmaa_avoidance"]),
        ("", ""),
        ("LEGACY POOL & HEIRS (v7)", ""),
        ("Legacy Pool Target Return", cfg.get("legacy_pool_return", cfg["roth_return"])),
        ("Legacy Pool Return Std Dev (MC)", cfg.get("legacy_pool_std", "")),
        ("Assumed Heir Tax Rate (PreTax/HSA)", cfg.get("heir_tax_rate", 0.24)),
        ("", ""),
        ("TAX (v7)", ""),
        ("Tax Cash/MM Interest Annually", cfg.get("tax_cash_interest", True)),
        ("", ""),
        ("SPENDING STRATEGY (v7)", ""),
        ("Strategy", cfg.get("spending_strategy", "fixed")),
        ("VPW Assumed Real Return", cfg.get("vpw_real_return_pct", "")),
        ("Absolute Minimum Annual Expenses (today's $)", cfg.get("absolute_min_annual_expenses", 0)),
        ("Max Annual Expenses (today's $)", cfg.get("max_annual_expenses", 0)),
        ("", ""),
        ("SURVIVING SPOUSE SCENARIO (v7)", ""),
        ("Model Surviving Spouse Scenario", cfg.get("model_widow_scenario", False)),
        ("Age of First Death", cfg.get("first_death_age", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("SS Survivor Benefit Retained", cfg.get("ss_survivor_pct", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("Pension Survivor Benefit Retained", cfg.get("jss_survivor_pct", "") if cfg.get("model_widow_scenario") else "n/a"),
        ("Widowhood Expense Reduction", cfg.get("expense_reduction_widowhood", "") if cfg.get("model_widow_scenario") else "n/a"),
    ]
    for r, (label, val) in enumerate(assumptions, 1):
        ws_a.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_a.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1:
            cell.number_format = pct_fmt
            cell.font = input_font
        elif isinstance(val, (int, float)) and val >= 1000:
            cell.number_format = money_k
            cell.font = input_font
        else:
            cell.font = input_font
    ws_a.column_dimensions["A"].width = 30
    ws_a.column_dimensions["B"].width = 18

    # ──────── Sheet 2: Accumulation Phase ────────
    if accum_df is not None and len(accum_df) > 0:
        ws_acc = wb.create_sheet("Accumulation")
        acc_cols = ["Age", "Year", "Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash",
                    "Contrib_Brokerage",
                    "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "Brokerage_EOY", "Brokerage_Basis",
                    "S_Plus_5yr", "S_Plus_10yr",
                    "Total_Liquid_Assets"]
        for c, col in enumerate(acc_cols, 1):
            ws_acc.cell(row=1, column=c, value=col.replace("_", " "))
        style_header(ws_acc, len(acc_cols))
        for r, row in enumerate(accum_df.to_dict("records") if isinstance(accum_df, pd.DataFrame) else accum_df, 2):
            for c, col in enumerate(acc_cols, 1):
                val = row.get(col, "")
                cell = ws_acc.cell(row=r, column=c, value=val)
                cell.font = num_font
                if col not in ("Age", "Year", "Phase"):
                    cell.number_format = money_k
        auto_width(ws_acc)

    # ──────── Sheet 3: Retirement Phase ────────
    ws_ret = wb.create_sheet("Retirement")
    ret_cols = [
        "Age", "Year", "Draw_Strategy", "Filing_Status",
        "SS_Income", "JSS_Income", "JSS_Taxable", "Rental_Income", "Rental_Taxable",
        "S_Plus_Income", "Passive_Income",
        "Base_Expenses", "Base_Expenses_Real", "Healthcare_Cost", "HDHP", "Gifts", "Lump_Sum",
        "Discretionary_Reduction", "VPW_Percentage",
        "Absolute_Min_Inflated", "Absolute_Min_Bound_Hit",
        "Max_Annual_Expenses_Inflated", "Max_Expenses_Bound_Hit",
        "PreTax_Draw", "Roth_Draw", "Cash_Draw", "HSA_Draw", "Brokerage_Draw", "Total_Draws",
        "Total_Income", "Total_Expenses", "Total_Tax", "Surplus_Deficit",
        "Return_PreTax", "Return_Roth", "Return_HSA", "Return_Cash", "Return_Legacy_Pool", "Return_Brokerage",
        "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "Brokerage_EOY", "Brokerage_Basis",
        "Total_Liquid_Assets", "Total_Real",
        "Cash_Interest_Taxable", "Brokerage_LTCG_Gain", "Brokerage_Dividend_Taxable",
        "Gross_Taxable_Income", "Federal_Taxable_Income",
        "Standard_Deduction", "Itemized_Deduction", "Best_Deduction", "Deduction_Type",
        "Federal_Tax", "Oregon_Tax",
        "Effective_Tax_Rate", "Withdrawal_Rate",
        "RMD", "RMD_Excess", "Roth_Conversion", "Roth_Conversion_Tax",
        "MAGI", "IRMAA_Threshold", "IRMAA_Lookback_MAGI", "IRMAA_Hit", "IRMAA_Cost",
        "Net_Investment_Income", "NIIT",
        "Inflation_Rate_Realized",
        "Legacy_Target_Per_Child", "Legacy_Roth_Per_Child", "Legacy_Inheritance_Per_Child",
        "Legacy_Target_Total", "Legacy_Roth_Total", "Legacy_Inheritance_Total", "Legacy_Total",
        "Bad_Return_Year", "Cum_Gifts", "Cum_Legacy_Roth", "Cum_Legacy_Inheritance", "Cum_Lump_Sums",
        "Legacy_Pool_EOY", "Legacy_Pool_Per_Child", "Cum_Legacy_Inheritance_Per_Child",
        "Legacy_Value_To_Date", "Legacy_Value_To_Date_Per_Child",
        "Family_Net_Worth", "Estate_At_Death", "Estate_At_Death_Per_Child", "Total_Inheritance_At_Death_Per_Child",
        "Oregon_Estate_Tax", "Heir_Tax_On_PreTax", "Heir_Tax_On_HSA",
        "After_Tax_Estate_At_Death", "After_Tax_Family_Net_Worth",
        "After_Tax_Estate_Per_Child", "After_Tax_Total_Inheritance_Per_Child",
    ]
    for c, col in enumerate(ret_cols, 1):
        ws_ret.cell(row=1, column=c, value=col.replace("_", " "))
    style_header(ws_ret, len(ret_cols))

    pct_columns = {"Effective_Tax_Rate", "Withdrawal_Rate",
                   "Return_PreTax", "Return_Roth", "Return_HSA", "Return_Cash", "Return_Legacy_Pool",
                   "Return_Brokerage", "VPW_Percentage", "Inflation_Rate_Realized"}
    bool_columns = {"IRMAA_Hit", "Bad_Return_Year", "Absolute_Min_Bound_Hit", "Max_Expenses_Bound_Hit"}
    text_columns = {"Draw_Strategy", "Filing_Status", "Deduction_Type"}
    neg_fill = PatternFill("solid", fgColor="FFE0E0")

    for r, (_, row) in enumerate(retire_df.iterrows(), 2):
        for c, col in enumerate(ret_cols, 1):
            val = row.get(col, "")
            cell = ws_ret.cell(row=r, column=c, value=val)
            cell.font = num_font
            cell.border = thin_border
            if col in pct_columns:
                cell.number_format = pct_fmt
            elif col in bool_columns:
                pass
            elif col in text_columns or col in ("Age", "Year"):
                pass
            else:
                cell.number_format = money_k
            # Red highlight for negative values
            if isinstance(val, (int, float)) and val < 0:
                cell.fill = neg_fill
    auto_width(ws_ret)

    # ──────── Sheet 4: Summary ────────
    ws_s = wb.create_sheet("Summary")
    first, last = retire_df.iloc[0], retire_df.iloc[-1]
    n_kids_xl = int(cfg.get("num_children", 0))
    summary = [
        ("RETIREMENT PLAN SUMMARY", ""),
        ("", ""),
        ("Starting Retirement Assets", first["Total_Liquid_Assets"]),
        ("  PreTax 401(k)", first["PreTax_EOY"]),
        ("  Roth IRA", first["Roth_EOY"]),
        ("  HSA", first["HSA_EOY"]),
        ("  Cash", first["Cash_EOY"]),
        ("  Taxable Brokerage", first.get("Brokerage_EOY", 0.0)),
        ("", ""),
        (f"Your Own Estate at Age {cfg['planning_end_age']}", last["Total_Liquid_Assets"]),
        ("Real Value (inflation-adjusted)", last["Total_Real"]),
        ("", ""),
        ("Avg Withdrawal Rate", retire_df["Withdrawal_Rate"].mean()),
        ("Avg Effective Tax Rate", retire_df["Effective_Tax_Rate"].mean()),
        ("Total Roth Conversions", retire_df["Roth_Conversion"].sum()),
        ("Total Conversion Tax Paid", retire_df["Roth_Conversion_Tax"].sum()),
        ("", ""),
        ("LEGACY GIFTING", ""),
        ("Cumulative Annual Gifts (cash, separate from Roth legacy)", last.get("Cum_Gifts", 0.0)),
        ("Total Gifted to Kids' Roth (nominal, no growth)", last.get("Cum_Legacy_Roth", 0.0)),
        (f"Legacy Pool Today (grown, age {cfg['planning_end_age']})", last.get("Legacy_Pool_EOY", 0.0)),
        ("Cum Legacy Inheritance (bad-years-only sub-total, NOT your estate)", last.get("Cum_Legacy_Inheritance", 0.0)),
        ("Cumulative Lump Sums", last.get("Cum_Lump_Sums", 0.0)),
        ("", ""),
        ("ESTATE AT DEATH (if it occurred at the final planning age)", ""),
        ("Your Own Estate (passes to heirs)", last.get("Estate_At_Death", last["Total_Liquid_Assets"])),
        ("Already Gifted (Kids' own Roth, grown)", last.get("Legacy_Pool_EOY", 0.0)),
        ("Total to Family (both combined)", last.get("Family_Net_Worth", last["Total_Liquid_Assets"])),
    ]
    if n_kids_xl > 0:
        summary += [
            ("", ""),
            (f"PER CHILD (\u00f7 {n_kids_xl})", ""),
            ("  Estate / Child", last.get("Estate_At_Death_Per_Child", 0.0)),
            ("  Roth Pool / Child", last.get("Legacy_Pool_Per_Child", 0.0)),
            ("  Total / Child", last.get("Total_Inheritance_At_Death_Per_Child", 0.0)),
        ]
    summary += [
        ("", ""),
        ("IRMAA Years Hit", int(retire_df["IRMAA_Hit"].sum())),
    ]
    for r, (label, val) in enumerate(summary, 1):
        ws_s.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_s.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1 and val > 0:
            cell.number_format = pct_fmt
        elif isinstance(val, (int, float)) and abs(val) >= 1000:
            cell.number_format = money_k
        cell.font = num_font
    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 20

    # ──────── Sheet 5: Monte Carlo (only if MC was run) ────────
    if mc_runs:
        ws_mc = wb.create_sheet("Monte Carlo")
        surv = sum(1 for r in mc_runs if mc_outcome_row(r)["Total_Liquid_Assets"] > 0)
        mc_header = [
            ("MONTE CARLO SUMMARY", ""),
            ("Simulations", len(mc_runs)),
            ("Survived (Total_Liquid_Assets > 0 at final age)", surv),
            ("Success Rate", surv / len(mc_runs)),
            ("", ""),
        ]
        for r, (label, val) in enumerate(mc_header, 1):
            ws_mc.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
            cell = ws_mc.cell(row=r, column=2, value=val)
            if isinstance(val, float) and 0 < val < 1:
                cell.number_format = pct_fmt
            cell.font = num_font

        start_row = len(mc_header) + 2
        bands_nom = compute_percentile_bands(mc_runs, "Total_Liquid_Assets")
        bands_real = compute_percentile_bands(mc_runs, "Total_Real")
        mc_cols = ["Age", "P5 (Nominal)", "P25 (Nominal)", "P50 Median (Nominal)", "P75 (Nominal)", "P95 (Nominal)",
                   "P5 (Today's $)", "P25 (Today's $)", "P50 Median (Today's $)", "P75 (Today's $)", "P95 (Today's $)"]
        for c, col in enumerate(mc_cols, 1):
            ws_mc.cell(row=start_row, column=c, value=col)
        style_header_row = ws_mc[start_row]
        for cell in style_header_row:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ages = bands_nom["Age"]
        for i, age_val in enumerate(ages):
            row_vals = [int(age_val),
                        bands_nom["p5"][i], bands_nom["p25"][i], bands_nom["p50"][i], bands_nom["p75"][i], bands_nom["p95"][i],
                        bands_real["p5"][i], bands_real["p25"][i], bands_real["p50"][i], bands_real["p75"][i], bands_real["p95"][i]]
            for c, val in enumerate(row_vals, 1):
                cell = ws_mc.cell(row=start_row + 1 + i, column=c, value=val)
                cell.font = num_font
                cell.number_format = money_k if c > 1 else "0"
        auto_width(ws_mc)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# STREAMLIT UI
# ============================================================



def qp_int(name, default):
    try:
        raw = str(st.query_params.get(name, default)).replace(',', '').replace('$', '').strip()
        return int(float(raw))
    except (TypeError, ValueError):
        return default

def qp_float(name, default):
    try:
        return float(str(st.query_params.get(name, default)).replace('%', '').strip())
    except:
        return default

def get_balance_source(file_balances):
    qp = st.query_params
    web_keys = ["pretax", "roth", "hsa", "s5", "s10", "brokerage", "cash"]
    if any(str(qp.get(k, "")).strip() for k in web_keys):
        return "web fetch from Excel"
    if any(v is not None for v in file_balances.values()):
        return "using Starting_balances.txt"
    return "using defaults"


# ── Cached wrappers ──
# Streamlit reruns the whole script on every widget interaction, so without
# caching, moving a slider that doesn't even affect `cfg` (e.g. the "Show in
# Today's Dollars" display toggle) still re-triggers a full deterministic
# run plus a full Monte Carlo of up to mc_sims paths. These wrappers key off
# the actual arguments, so a rerun with an unchanged cfg is a cache hit.
@st.cache_data(show_spinner=False)
def cached_run_simulation(cfg):
    return run_simulation(cfg)


@st.cache_data(show_spinner="Running Monte Carlo simulation...")
def cached_run_monte_carlo(cfg, n_sims, std_dev, max_up, seed, correlation, fat_tailed, t_df):
    return run_monte_carlo(cfg, n_sims, std_dev, max_up, seed=seed,
                            correlation=correlation, fat_tailed=fat_tailed, t_df=t_df)


@st.cache_data(show_spinner=False)
def cached_run_optimizer(cfg, mc_sims, mc_std, mc_max, mc_seed, correlation, fat_tailed, t_df):
    return run_optimizer(cfg, mc_sims, mc_std, mc_max, mc_seed,
                          correlation=correlation, fat_tailed=fat_tailed, t_df=t_df)


@st.cache_data(show_spinner=False)
def cached_draw_order_comparison(cfg, mc_std, mc_max, correlation, fat_tailed, t_df):
    """Runs PreTax-First and Brokerage-First head-to-head on the SAME matched
    random paths (fixed internal seed, independent of the user's own MC seed
    choice/'random' setting) so the comparison isolates just the draw-order
    effect instead of also picking up path-to-path noise -- same approach
    used to validate this feature before building it. Uses a smaller,
    fixed sim count (not the user's mc_sims) since this runs automatically
    on every Monte Carlo view, same speed/fidelity tradeoff run_optimizer
    already makes for its own sweep. `cfg` should have brokerage_before_pretax
    already stripped by the caller so toggling that setting doesn't bust
    this cache unnecessarily -- this function overrides it both ways anyway.

    Tracks BOTH the retiree's own survival (success_rate) AND legacy value
    (median_after_tax_family_net_worth) -- verified directly that these can
    point in OPPOSITE directions: Brokerage-First spends Cash/Brokerage down
    first, which protects the retiree's own survival against sequence-of-
    returns risk, but leaves PreTax compounding longer, producing bigger
    RMDs and a bigger SECURE-Act income-tax bill for heirs later -- in one
    test scenario, +3pp survival came with -$528k in after-tax family net
    worth. A comparison that only reported survival would silently recommend
    trading legacy value away without ever saying so.
    """
    cmp_sims = 300
    cmp_seed = 42
    num_years = cfg["planning_end_age"] - cfg["retirement_age"] + 1
    accum_result = run_accumulation(cfg)
    results = {}
    for key, flag in [("pretax_first", False), ("brokerage_first", True)]:
        variant_cfg = {**cfg, "brokerage_before_pretax": flag}
        rng = np.random.default_rng(cmp_seed)
        runs = []
        for _ in range(cmp_sims):
            ov = build_mc_return_overrides(variant_cfg, num_years, rng, mc_std, mc_max, correlation, fat_tailed, t_df)
            _, df = run_simulation(variant_cfg, return_overrides=ov, accum_result=accum_result)
            runs.append(df)
        finals = np.array([r["Total_Liquid_Assets"].iloc[-1] for r in runs])
        after_tax_fnw = np.array([r["After_Tax_Family_Net_Worth"].iloc[-1] for r in runs])
        results[key] = {
            "success_rate": float((finals > 0).mean()),
            "median_final": float(np.median(finals)),
            "median_after_tax_family_net_worth": float(np.median(after_tax_fnw)),
        }
    results["sims"] = cmp_sims
    return results


def main():
    st.set_page_config(page_title="Retirement Income Planner", page_icon="\U0001F4CA",
                       layout="wide", initial_sidebar_state="expanded")
    st.title("\U0001F3E6 Retirement Income & Tax Planning Simulator")
    st.caption("Accumulation + Retirement modeling with Monte Carlo and Excel export")

    tax_staleness_warning = verify_tax_constants()
    if tax_staleness_warning:
        st.warning(tax_staleness_warning, icon="⚠️")

    with st.sidebar:
        st.header("\u2699\ufe0f Configuration")

        with st.expander("\U0001F464 Age & Timeline", expanded=True):
            current_age = st.number_input("Current Age", 45, 70, 55,
                help=f"Your age today, unless the box below is checked -- then enter "
                     f"your age as of {CURRENT_YEAR + 1} instead.")
            skip_rest_of_this_year = st.checkbox(
                f"No more contributions coming at age {current_age}", value=True,
                help=f"When checked, {CURRENT_YEAR} contributes nothing further -- your "
                     f"account balances below are treated as final for {CURRENT_YEAR}, "
                     f"and the model's first full contribution year is {CURRENT_YEAR + 1}. "
                     "Uncheck if you're using this earlier in the year and still expect "
                     f"more {CURRENT_YEAR} contributions.",
            )
            plan_base_year = CURRENT_YEAR + 1 if skip_rest_of_this_year else CURRENT_YEAR
            retirement_age = st.number_input("Retirement Age", 55, 63, 55, step=1)
            planning_end = st.slider("Plan Through Age", 85, 100, 95)

        with st.expander("\U0001F5A4 Surviving Spouse Scenario"):
            st.caption("Models the 'widow's penalty': filing status switches from MFJ to Single the year after the first spouse's death, with smaller brackets/deduction, a partial Social Security survivor benefit, and (optionally) reduced pension and living costs.")
            widow_scenario = st.checkbox("Model Surviving Spouse Scenario", value=False)
            if widow_scenario:
                first_death_age = st.slider("Age of First Spouse's Death", 55, 100, 85)
                ss_survivor_pct = st.slider("SS Survivor Benefit Retained (%)", 30, 100, 65) / 100
                jss_survivor_pct = st.slider("Pension Survivor Benefit Retained (%)", 0, 100, 100) / 100
                exp_red_widow = st.slider("Living Expense Reduction After Widowhood (%)", 0, 50, 20) / 100
            else:
                first_death_age, ss_survivor_pct, jss_survivor_pct, exp_red_widow = 999, 0.65, 1.0, 0.0

        with st.expander("\U0001F4B0 Current Balances (Today)", expanded=True):
            file_balances, file_loaded = load_starting_balances()
            balance_source = get_balance_source(file_balances)
            if balance_source == "web fetch from Excel":
                st.success("\U0001F4CA Using balances passed in via URL parameters (Excel launch)")
            elif balance_source == "using Starting_balances.txt":
                st.success("Loaded starting_balances.txt")
            else:
                st.info("No starting_balances.txt or URL parameters found; using defaults")
            balance_specs = [
                ("pretax_input", "pretax", file_balances["401k"] if file_balances["401k"] is not None else 1_475_000),
                ("roth_input", "roth", file_balances["roth"] if file_balances["roth"] is not None else 510_000),
                ("hsa_input", "hsa", file_balances["hsa"] if file_balances["hsa"] is not None else 130_000),
                ("s5_input", "s5", file_balances["s_plus_5"] if file_balances["s_plus_5"] is not None else 300_000),
                ("s10_input", "s10", file_balances["s_plus_10"] if file_balances["s_plus_10"] is not None else 400_000),
                ("brokerage_input", "brokerage", file_balances["brokerage"] if file_balances["brokerage"] is not None else 0),
                ("cash_input", "cash", file_balances["cash"] if file_balances["cash"] is not None else 745_000),
            ]
            for widget_key, param_name, fallback in balance_specs:
                if param_name in st.query_params:
                    st.session_state[widget_key] = qp_int(param_name, fallback)
                elif widget_key not in st.session_state:
                    st.session_state[widget_key] = fallback

            pretax = st.number_input("PreTax 401(k) ($)", 0, 10_000_000, step=50_000, format="%d", key="pretax_input")
            roth_bal = st.number_input("Roth IRA ($)", 0, 5_000_000, step=25_000, format="%d", key="roth_input")
            hsa_bal = st.number_input("HSA ($)", 0, 500_000, step=10_000, format="%d", key="hsa_input")
            s5_bal = st.number_input("S+ 5-Year Payout ($)", 0, 2_000_000, step=25_000, format="%d", key="s5_input")
            s10_bal = st.number_input("S+ 10-Year Payout ($)", 0, 2_000_000, step=25_000, format="%d", key="s10_input")
            brokerage_bal = st.number_input(
                "Taxable Brokerage ($)", 0, 5_000_000, step=25_000, format="%d", key="brokerage_input",
                help="A separate, non-tax-advantaged account. Contributed dollars come back "
                     "tax-free; growth is taxed as a long-term capital gain when withdrawn. "
                     "Sits between Cash and Roth in the draw order, and gets a full cost-basis "
                     "step-up at death (unlike PreTax/HSA) -- see the Tax and Legacy tabs.",
            )
            cash_bal = st.number_input("Cash ($)", 0, 1_000_000, step=10_000, format="%d", key="cash_input")

        with st.expander("\U0001F4BC Working Years Contributions"):
            st.caption(f"Annual contributions for {max(0, retirement_age - current_age)} remaining working years")
            c_401k = st.number_input("401(k) Employee ($)", 0, 50_000, 24_500, step=500, format="%d")
            c_roth401k = st.number_input("Roth 401(k) ($)", 0, 50_000, 8_000, step=500, format="%d")
            c_roth_ira = st.number_input("Roth IRA per Spouse ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_hsa = st.number_input("HSA Family ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_mega = st.number_input("Mega Backdoor Roth ($)", 0, 50_000, 29_000, step=1_000, format="%d")
            c_match = st.number_input("Employer Match ($)", 0, 50_000, 18_000, step=1_000, format="%d")
            st.divider()
            c_cash = st.number_input("Annual Cash Savings ($)", 0, 500_000, 40_000, step=5_000, format="%d")
            c_cash_lump = st.number_input("Final Year Cash Lump ($)", 0, 500_000, 0, step=10_000, format="%d")
            st.divider()
            brokerage_basis_pct = st.slider(
                "Brokerage Starting Cost Basis (% of balance)", 0, 100, 100, 5,
                help="100% = no embedded gain today (e.g. a brand-new account or one "
                     "you're funding from cash). Lower this if the account already "
                     "holds unrealized gains.",
            )
            brokerage_basis = brokerage_bal * brokerage_basis_pct / 100
            c_brokerage = st.number_input("Brokerage Annual Contribution ($)", 0, 200_000, 200_000, step=5_000, format="%d")

        with st.expander("\U0001F4C8 Performance Assumptions"):
            pretax_ret = st.slider("PreTax Target Return %", 0.0, 12.0, 6.0, 0.5) / 100
            roth_ret = st.slider("Roth Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            hsa_ret = st.slider("HSA Target Return %", 0.0, 12.0, 5.0, 0.5) / 100
            cash_ret = st.slider("Cash/MM Target Return %", 0.0, 8.0, 4.0, 0.25) / 100
            brokerage_ret = st.slider("Brokerage Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            brokerage_std_pct = st.slider("Brokerage Return Std Dev % (Monte Carlo)", 1.0, 25.0, 15.0, 0.5) / 100
            brokerage_div_yield = st.slider(
                "Brokerage Dividend Yield %", 0.0, 5.0, 1.8, 0.1,
                help="Unlike price appreciation (deferred until you sell), a taxable brokerage "
                     "account's dividends are taxed the year they're paid, whether or not you "
                     "spend them -- assumed reinvested, which raises the account's cost basis by "
                     "the same amount. Taxed at the same preferential rate as long-term capital "
                     "gains. Only applied during retirement -- the accumulation phase before "
                     "retirement isn't tax-modeled at all (a broader simplification, not specific "
                     "to this input). 0 = off (all brokerage return treated as deferred).",
            ) / 100

        with st.expander("\U0001F3B2 Monte Carlo Settings"):
            mc_enabled = st.checkbox("Enable Monte Carlo", value=True)
            mc_sims = st.slider("Simulations", 50, 10000, 5000, 50)
            mc_std = st.slider("Return Std Dev %", 1.0, 25.0, 12.0, 0.5) / 100
            mc_max = st.slider("Max Upside Cap %", 5.0, 25.0, 18.0, 0.5) / 100
            mc_seed = st.number_input("Seed (0=random)", 0, 99999, 0)
            st.caption("PreTax/Roth/HSA/Legacy Pool move together each year (same market factor) instead of being drawn fully independently -- more realistic since they typically hold similar underlying investments.")
            mc_correlation = st.slider("Cross-Account Correlation", 0.0, 1.0, 0.85, 0.05,
                help="1.0 = all equity-like accounts move in lockstep each year. 0.0 = fully independent (old behavior). Cash/MM always gets a smaller fraction of this (money-market rates track policy rates, not stocks).")
            mc_fat_tailed = st.checkbox("Fat-Tailed Returns (Student-t)", value=True,
                help="Real market returns crash harder and more often than a bell curve predicts. Enabling this uses a Student-t distribution instead of Normal, producing more realistic tail risk (more frequent extreme years) at the same target return and std dev.")
            mc_t_df = st.slider("Fat-Tail Intensity (lower = fatter tails)", 3, 15, 5, 1, disabled=not mc_fat_tailed) if mc_fat_tailed else 5
            st.divider()
            st.caption("By default every MC path uses the same fixed General Inflation rate from the Expenses section -- only returns vary. Since returns are nominal and expenses inflate deterministically, this understates real risk over a multi-decade horizon (a 1970s-style inflation path can't be drawn at all). Set a std dev above 0 to let inflation vary by path too.")
            mc_inflation_std = st.slider("Inflation Std Dev %", 0.0, 5.0, 1.5, 0.1,
                help="0 = off (every path uses the flat General Inflation rate, original behavior). Above 0, each Monte Carlo path draws its own year-by-year inflation series centered on General Inflation.") / 100
            mc_inflation_correlation = st.slider("Inflation-Equity Correlation", -0.5, 0.5, -0.15, 0.05, disabled=(mc_inflation_std == 0),
                help="Correlates inflation surprises to the SAME market factor driving this path's equity returns. Negative (default) means a bad market year is more likely to also be a high-inflation year (echoing 2022) -- a simplifying assumption, not a calibrated historical estimate. 0 = inflation drawn independently of returns.") if mc_inflation_std > 0 else -0.15
            st.divider()
            st.caption("By default every MC path is judged 'successful' by whether the portfolio survives to a single fixed age (Plan Through Age above) -- but nobody actually knows their own death date in advance. Enabling this instead draws a random age-at-death per simulation (see below) and judges each path against ITS OWN horizon.")
            mc_mortality_based = st.checkbox("Mortality-Based Planning Horizon", value=False,
                help="Off (default): success = portfolio > $0 at the fixed Plan Through Age, same for every path, the original behavior. On: each path still simulates all the way out to Plan Through Age (so the charts keep showing the full trajectory), but success/failure is judged at that path's own randomly drawn death age instead -- a path that 'fails' at 95 but whose household didn't actually live past 88 still counts as successful. Draws TWO independent lifespans per path (a household) and uses the LATER one, since a couple's plan has to survive the surviving spouse, not either person alone. Mortality is modeled with an approximate actuarial curve (Gompertz's law), not the exact published SSA table -- see MORTALITY_GOMPERTZ_* in the source for what that means.")

        with st.expander("\U0001F4E5 Income Sources"):
            ss_age = st.slider("SS Start Age", 62, 70, 65)
            ss_amount = st.number_input("SS Annual (Future $)", 0, 200_000, 106_000, step=1_000, format="%d")
            ss_cut = st.checkbox("Apply 22% benefit cut (trust fund depletion)", value=False,
                help="The SSA trustees project the OASI trust fund reserves are depleted "
                     "around 2033, at which point incoming payroll tax revenue alone "
                     "covers roughly 77-78% of scheduled benefits -- an automatic ~22% "
                     "cut unless Congress acts before then. Off by default (assumes "
                     "benefits are paid in full); check this to stress-test the plan "
                     "against that scenario instead.")
            if ss_cut:
                ss_amount = round(ss_amount * 0.78)
            ss_cola = st.slider("SS COLA %", 0.0, 5.0, 2.0, 0.25) / 100
            st.divider()
            jss_age = st.slider("JSS Pension Start Age", 60, 70, 60)
            jss_amount = st.number_input("JSS Annual ($)", 0, 100_000, 18_000, step=1_000, format="%d")
            jss_cola_pct = st.slider("JSS COLA %", 0.0, 3.0, 1.0, 0.25) / 100
            jss_recovery = st.number_input("JSS Recovery Years", 0, 10, 4)
            st.divider()
            rental = st.number_input("Rental Gross ($)", 0, 200_000, 44_000, step=1_000, format="%d")
            rental_tax_pct = st.slider("Rental Taxable %", 0, 100, 50) / 100

        with st.expander("\U0001F4B8 Expenses"):
            base_exp = st.number_input("Annual Expenses ($)", 50_000, 500_000, 168_000, step=5_000, format="%d")
            inflation = st.slider("General Inflation %", 0.0, 6.0, 2.87, 0.01) / 100
            st.subheader("Healthcare")
            hc_pre = st.number_input("Pre-Medicare ($)", 0, 100_000, 33_000, step=1_000, format="%d")
            hc_post = st.number_input("Post-Medicare ($)", 0, 50_000, 12_000, step=1_000, format="%d")
            hc_inflation = st.slider("HC Inflation %", 0.0, 10.0, 5.0, 0.5) / 100
            hdhp_cost = st.number_input("HDHP ($)", 0, 30_000, 12_000, step=1_000, format="%d")
            st.subheader("Post-80 Adjustment")
            exp_red_80 = st.slider("Expense Reduction After 80 (%)", 0, 25, 25, 5) / 100
            st.caption("Lifestyle slowdown: reduce base expenses after age 80.")
            st.subheader("Negative Return Adjustment")
            neg_ret_reduction = st.number_input("Discretionary Draw Reduction if Year Return < 0 (today's $)", 0, 200_000, 20_000, step=1000, format="%d")
            st.caption("Entered in today's dollars -- inflates every year alongside your base expenses, so a bad year in year 1 and a bad year in year 30 get an equivalent real cut.")
            st.subheader("Absolute Survival Floor")
            absolute_min_expenses = st.number_input(
                "Absolute Minimum Annual Expenses (today's $)", 0, 200_000, 145_000, step=5_000, format="%d",
                help="A hard floor on base spending that no reduction above -- post-80 slowdown "
                     "or widowhood -- can push below, no matter how many bad years compound. "
                     "0 = off (no absolute floor).",
            )
            st.subheader("Spending Ceiling")
            max_annual_expenses = st.number_input(
                "Max Annual Expenses (today's $)", 0, 500_000, int(base_exp), step=5_000, format="%d",
                help="A hard ceiling on base spending. Matters most for VPW, whose withdrawal "
                     "percentage accelerates near the end of the plan and would otherwise force-"
                     "spend far more than a realistic budget in the final years -- anything above "
                     "this ceiling simply stays invested instead of being withdrawn at all. A "
                     "no-op for Fixed Real Spending, which never exceeds its own planned amount. "
                     "Defaults to your Annual Expenses figure above. 0 = off (no ceiling).",
            )
            st.subheader("Spending Strategy")
            spending_strategy_choice = st.radio(
                "Strategy",
                ["Fixed Real Spending", "Variable Percentage Withdrawal (VPW)"],
                index=1,
            )
            if spending_strategy_choice.startswith("Variable"):
                spending_strategy = "vpw"
            else:
                spending_strategy = "fixed"

            # Default for when VPW's own input isn't shown below
            vpw_real_return = 0.04

            if spending_strategy == "vpw":
                st.caption(
                    "Spend a percentage of your CURRENT total portfolio balance each year, "
                    "recalculated fresh annually -- the percentage itself rises with age (fewer "
                    "expected remaining years), using the same amortization math as an RMD "
                    "divisor. Self-correcting by construction: a bad year automatically reduces "
                    "next year's dollar amount, a good year raises it. Popularized by the "
                    "Bogleheads community as a simpler alternative to fixed withdrawals. The "
                    "Absolute Minimum Annual Expenses floor above still applies on top as a "
                    "hard backstop."
                )
                vpw_real_return = st.slider(
                    "Assumed Real Return (%)", 0.0, 8.0, 4.0, 0.25,
                    help="Expected long-run return net of inflation for your overall portfolio "
                         "blend -- higher assumptions front-load more spending into earlier "
                         "years. Typical published Bogleheads VPW tables use roughly 3-5% for a "
                         "balanced (e.g. 60/40 stock/bond) portfolio.",
                ) / 100
            st.subheader("Lump Sum")
            lump_age = st.number_input("Lump at Age", 55, 95, 70)
            lump_amt = st.number_input("Lump Amount ($)", 0, 1_000_000, 0, step=10_000, format="%d")
            lump_sums = {lump_age: lump_amt} if lump_amt > 0 else {}

        with st.expander("\U0001F381 Gifts & Giving"):
            gifts = st.number_input("Annual Gifts ($)", 0, 100_000, 0, step=1_000, format="%d")

        with st.expander("\U0001F3AF Strategy", expanded=True):
            st.markdown("**PreTax→12% bracket, then Cash, then Roth. RMD dominates at 75+.**")
            rmd_start = st.slider("RMD Start Age", 73, 75, 75)
            st.divider()
            roth_conv = st.checkbox("Enable Roth Conversions", value=True)
            roth_bracket = st.selectbox("Conversion Target", ["12%", "22%", "24%"], index=1)
            roth_margin = st.number_input("Min PreTax Keep ($)", 0, 1_000_000, 100_000, step=25_000, format="%d")
            st.divider()
            irmaa_avoid = st.checkbox("IRMAA Avoidance", value=True)
            st.divider()
            draw_priority = st.radio(
                "PreTax vs Brokerage Draw Priority (pre-RMD years)",
                ["PreTax First (default)", "Brokerage First"],
                index=0,
                help="PreTax First: spending draws from PreTax up to the bracket target, THEN "
                     "Cash/Brokerage -- the original behavior. Brokerage First: spending draws "
                     "from Cash/Brokerage FIRST, leaving the ordinary-income bracket free for "
                     "Roth Conversions instead of the spending draw using it up. Not a free lunch, "
                     "though: leaving PreTax untouched longer means it keeps compounding, which "
                     "means bigger future RMDs and a bigger SECURE-Act income-tax bill for whoever "
                     "inherits it -- Brokerage First tends to improve the RETIREE's OWN survival "
                     "odds (protects against selling equities in an early bad market) at some cost "
                     "to what's left for HEIRS after tax. See the Monte Carlo tab for an automatic "
                     "side-by-side comparison against your specific scenario on BOTH survival and "
                     "after-tax family net worth -- they don't always agree.",
            )
            brokerage_before_pretax = draw_priority.startswith("Brokerage")

        with st.expander("\U0001F468\u200D\U0001F469\u200D\U0001F467\u200D\U0001F466 Legacy"):
            num_kids = st.number_input("Children", 0, 10, 4)
            roth_per_child = st.number_input("Roth/Child/Year ($)", 0, 50_000, 7_500, step=500, format="%d")
            legacy_years = st.slider("Legacy Duration (years)", 1, 20, 20)
            st.caption("The kids' Roth pool can use its own return/risk profile instead of automatically matching your own Roth account.")
            legacy_pool_ret = st.slider("Legacy Pool Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            legacy_pool_std_pct = st.slider("Legacy Pool Return Std Dev % (Monte Carlo)", 1.0, 25.0, 12.0, 0.5) / 100
            st.caption("Inherited pretax/HSA accounts must be distributed within 10 years (SECURE Act) and are taxed at the heir's own bracket -- unlike Roth, which passes tax-free. Used for the 'after-tax to heirs' figures on the Legacy tab.")
            heir_tax_rate = st.slider("Assumed Heir Tax Rate on PreTax/HSA (%)", 0, 45, 24, 1) / 100

        with st.expander("\U0001F3DB Tax"):
            or_resident = st.checkbox("Oregon Resident", value=True)
            std_ded = st.number_input("Std Deduction MFJ ($)", 0, 50_000, 32_500, step=100, format="%d")
            bracket_infl = st.slider("Bracket Inflation %", 0.0, 5.0, 2.67, 0.01) / 100
            tax_cash_int = st.checkbox("Tax Cash/MM Interest Annually (realistic)", value=True,
                help="Money-market interest is taxed as ordinary income every year, unlike unrealized gains in an equity account. Uncheck to model 'cash' as literal currency instead of an invested account.")

        with st.expander("\U0001F4CA Display Options"):
            show_real = st.toggle("Show in Today's Dollars", value=True)
            st.caption("When on, all dollar amounts are inflation-adjusted to present value.")

    # ── BUILD CONFIG ──
    cfg = dict(
        current_age=current_age, retirement_age=retirement_age, planning_end_age=planning_end,
        plan_base_year=plan_base_year,
        pretax_401k=pretax, roth_ira=roth_bal, hsa=hsa_bal,
        s_plus_5yr=s5_bal, s_plus_10yr=s10_bal, cash=cash_bal,
        brokerage=brokerage_bal, brokerage_basis=brokerage_basis,
        contrib_401k=c_401k, contrib_roth401k=c_roth401k, contrib_roth_ira=c_roth_ira,
        contrib_hsa=c_hsa, contrib_mega_backdoor=c_mega, contrib_employer_match=c_match,
        contrib_cash_annual=c_cash, contrib_cash_final_lump=c_cash_lump,
        contrib_brokerage_annual=c_brokerage,
        pretax_return=pretax_ret, roth_return=roth_ret, hsa_return=hsa_ret, cash_return=cash_ret,
        brokerage_return=brokerage_ret, brokerage_std=brokerage_std_pct,
        brokerage_dividend_yield=brokerage_div_yield,
        ss_start_age=ss_age, ss_annual_amount=ss_amount, ss_cola=ss_cola,
        jss_start_age=jss_age, jss_annual_amount=jss_amount, jss_cola=jss_cola_pct,
        jss_recovery_years=jss_recovery, rental_gross=rental, rental_taxable_pct=rental_tax_pct,
        base_annual_expenses=base_exp, inflation_rate=inflation,
        inflation_std=mc_inflation_std, inflation_correlation=mc_inflation_correlation,
        mortality_based_horizon=mc_mortality_based,
        healthcare_pre_medicare=hc_pre, healthcare_post_medicare=hc_post,
        healthcare_inflation=hc_inflation, hdhp_annual=hdhp_cost,
        expense_reduction_post80=exp_red_80,
        gifts_annual=gifts, lump_sums=lump_sums,
        standard_deduction=std_ded, bracket_inflation=bracket_infl, oregon_resident=or_resident,
        tax_cash_interest=tax_cash_int,
        rmd_start_age=rmd_start,
        draw_order=(["cash", "brokerage", "pretax", "roth", "hsa"] if brokerage_before_pretax
                    else ["pretax", "cash", "brokerage", "roth", "hsa"]),
        brokerage_before_pretax=brokerage_before_pretax,
        roth_conversion_enabled=roth_conv, roth_conversion_target_bracket=roth_bracket,
        roth_conversion_margin=roth_margin, irmaa_avoidance=irmaa_avoid,
        neg_ret_draw_reduction=neg_ret_reduction,
        absolute_min_annual_expenses=absolute_min_expenses,
        max_annual_expenses=max_annual_expenses,
        num_children=num_kids, roth_legacy_per_child=roth_per_child,
        legacy_years=legacy_years, legacy_inflation=inflation,
        legacy_pool_return=legacy_pool_ret, legacy_pool_std=legacy_pool_std_pct,
        heir_tax_rate=heir_tax_rate,
        spending_strategy=spending_strategy,
        vpw_real_return_pct=vpw_real_return,
        model_widow_scenario=widow_scenario, first_death_age=first_death_age,
        ss_survivor_pct=ss_survivor_pct, jss_survivor_pct=jss_survivor_pct,
        expense_reduction_widowhood=exp_red_widow,
    )

    # ── RUN ──
    accum_rows, df = cached_run_simulation(cfg)
    accum_df = pd.DataFrame(accum_rows) if accum_rows else None
    mc_runs = None
    if mc_enabled:
        if mc_seed > 0:
            # A fixed seed makes the run fully deterministic given cfg, so
            # it's safe -- and often the difference between an instant
            # rerun and a multi-second one -- to cache.
            mc_runs = cached_run_monte_carlo(cfg, mc_sims, mc_std, mc_max, mc_seed,
                                              mc_correlation, mc_fat_tailed, mc_t_df)
        else:
            # Seed 0 means "random": each rerun should draw a genuinely new
            # set of paths, same as before caching was added -- caching this
            # would freeze the "random" run to whatever it first computed.
            mc_runs = run_monte_carlo(cfg, mc_sims, mc_std, mc_max, seed=None,
                                       correlation=mc_correlation, fat_tailed=mc_fat_tailed, t_df=mc_t_df)

    # ── DEFLATOR (today's dollars conversion) ──
    # When show_real is True, deflate future dollar amounts back to present value
    dollar_label = "Today's $" if show_real else "Future $"
    def deflate(series_or_val, year_col_or_yir):
        """Deflate future dollars to today's dollars."""
        if not show_real:
            return series_or_val
        factors = (1 + inflation) ** year_col_or_yir
        return series_or_val / factors

    # Real years of growth/contributions actually modeled before retirement
    # starts -- read directly off the accumulation phase's own output
    # (len(accum_rows)) rather than re-deriving it from current_age/
    # retirement_age/plan_base_year a second time here. Two independent
    # formulas computing the same thing drift apart the moment either one
    # changes; this way there's exactly one, and it's also used below for the
    # "At Retirement" KPI so that figure and the table agree.
    accum_years = len(accum_rows) if accum_rows else 0

    # Build display-ready retirement df
    df_disp = df.copy()
    if show_real:
        df_disp["_deflate_yrs"] = accum_years + df_disp["Years_Retired"]
        money_cols = [c for c in df_disp.columns if c not in (
            "Age", "Year", "Years_Retired", "Phase", "Draw_Strategy",
            "Effective_Tax_Rate", "Withdrawal_Rate",
            "IRMAA_Hit", "Deduction_Type",
            "Return_PreTax", "Return_Roth", "Return_HSA", "Return_Cash",
            "Return_Legacy_Pool", "Return_Brokerage", "VPW_Percentage",
            # Already computed as real (today's-$) figures by the engine --
            # deflating them again here would double-discount them.
            "Total_Real", "Base_Expenses_Real",
            "RMD_Excess", "_deflate_yrs")]
        for c in money_cols:
            if c in df_disp.columns and df_disp[c].dtype in [np.float64, np.int64, float, int]:
                df_disp[c] = df_disp[c] / ((1 + inflation) ** df_disp["_deflate_yrs"])
        df_disp.drop(columns=["_deflate_yrs"], inplace=True)

    # ── KPI ROW ──
    c1, c2, c3, c4, c5 = st.columns(5)
    first, last = df_disp.iloc[0], df_disp.iloc[-1]
    # Raw entered balances, as-of today -- NOT accum_df.iloc[0], which (now
    # that accumulation can start at plan_base_year rather than CURRENT_YEAR)
    # already has a full year of contributions/growth baked in and would
    # mislabel that as "today's."
    today_assets = pretax + roth_bal + hsa_bal + cash_bal + brokerage_bal
    c1.metric("Today's Assets", f"${today_assets:,.0f}")
    # The moment retirement STARTS -- balance handed off from accumulation,
    # before that first year's retirement spending/taxes are drawn (which is
    # what df/df_disp's first row already reflects, further along than this).
    # With no accumulation years modeled (e.g. retiring immediately), that
    # handoff balance is just today's balance, unchanged -- so this and
    # Today's Assets agree exactly, as expected, instead of differing by a
    # year of spending that never actually happened.
    at_retirement_nominal = (accum_df.iloc[-1]["Total_Liquid_Assets"]
                              if accum_df is not None and len(accum_df) > 0 else today_assets)
    at_retirement = deflate(at_retirement_nominal, accum_years)
    c2.metric(f"At Retirement ({dollar_label})", f"${at_retirement:,.0f}")
    legacy_pool_last = last.get("Legacy_Pool_EOY", 0.0)
    c3.metric(f"Age {planning_end} ({dollar_label})", f"${last['Total_Liquid_Assets']:,.0f}",
              delta=(f"+${legacy_pool_last:,.0f} already gifted (see Legacy tab)" if legacy_pool_last > 0 else None),
              delta_color="off",
              help="Your own remaining accounts (pretax/Roth/HSA/cash) -- excludes money already gifted to the kids' Roth accounts, since that's no longer part of your own net worth. See the Legacy tab for the combined family total.")

    depleted = df[df["Total_Liquid_Assets"] <= 0]
    if len(depleted) > 0:
        c4.metric("\u26a0\ufe0f Depleted", f"Age {int(depleted.iloc[0]['Age'])}", delta="RISK", delta_color="inverse")
    else:
        c4.metric("\u2705 Lasts To", f"Age {planning_end}+")

    if mc_runs:
        surv = sum(1 for r in mc_runs if mc_outcome_row(r)["Total_Liquid_Assets"] > 0)
        c5.metric("MC Success", f"{surv/len(mc_runs)*100:.0f}%", delta=f"{len(mc_runs)} sims", delta_color="off")
    else:
        c5.metric("Avg WR", f"{df['Withdrawal_Rate'].mean():.1%}")

    st.divider()

    # ── TABS ──
    tab_names = ["\U0001F4BC Accumulation", "\U0001F4CA Retirement Assets", "\U0001F4B0 Income & Expenses",
                 "\U0001F381 Legacy", "\U0001F3DB Tax", "\U0001F504 Roth Conversions"]
    if mc_runs: tab_names.append("\U0001F3B2 Monte Carlo")
    if mc_runs: tab_names.append("\U0001F3AF Optimizer")
    tab_names.append("\U0001F32A Sensitivity")
    tab_names.append("\U0001F4CB Full Data")
    tabs = st.tabs(tab_names)
    ti = 0

    # ── Accumulation Tab ──
    with tabs[ti]:
        if accum_df is not None and len(accum_df) > 0:
            total_contrib = accum_df[["Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash"]].sum().sum()
            ac1, ac2, ac3 = st.columns(3)
            ac1.metric("Working Years", f"{accum_years}")
            ac2.metric("Total Contributions", f"${total_contrib:,.0f}")
            ac3.metric("Retirement Day Assets", f"${at_retirement:,.0f}")

            fig = go.Figure()
            for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"),
                                     ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                     ("Brokerage_EOY","Brokerage","rgba(230,126,34,0.6)"),
                                     ("Roth_EOY","Roth","rgba(142,68,173,0.6)"),
                                     ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
                fig.add_trace(go.Scatter(x=accum_df["Age"], y=accum_df[col], mode="lines", name=name,
                                         stackgroup="one", line=dict(width=0.5), fillcolor=color))
            fig.update_layout(title="Accumulation Phase: Account Growth",
                              xaxis_title="Age", yaxis_title="$", yaxis_tickformat="$,.0f", height=450)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(accum_df, use_container_width=True, height=300)
        else:
            st.info("No accumulation phase (already at retirement age).")
    ti += 1

    # ── Retirement Assets Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"), ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                 ("Brokerage_EOY","Brokerage","rgba(230,126,34,0.6)"),
                                 ("Roth_EOY","Roth","rgba(142,68,173,0.6)"), ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
            fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp[col], mode="lines", name=name,
                                     stackgroup="one", line=dict(width=0.5), fillcolor=color))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Real"], mode="lines", name="Real Value",
                                 line=dict(color="black", width=2, dash="dash")))
        fig.update_layout(title="Retirement Account Balances", xaxis_title="Age", yaxis_title="$",
                          yaxis_tickformat="$,.0f", hovermode="x unified", height=500)
        st.plotly_chart(fig, use_container_width=True)

        colors = ["red" if x > 0.04 else "orange" if x > 0.03 else "green" for x in df["Withdrawal_Rate"]]
        fig2 = go.Figure(go.Bar(x=df["Age"], y=df["Withdrawal_Rate"]*100, marker_color=colors))
        fig2.add_hline(y=4, line_dash="dash", line_color="red", annotation_text="4% Rule")
        fig2.update_layout(title="Withdrawal Rate", xaxis_title="Age", yaxis_title="%", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Income & Expenses Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, nm in [("SS_Income","SS"),("JSS_Income","JSS"),("Rental_Income","Rental"),
                        ("S_Plus_Income","S+"),("PreTax_Draw","PreTax Draw"),("Roth_Draw","Roth Draw")]:
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp[col], name=nm))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Expenses"], mode="lines+markers",
                                 name="Expenses", line=dict(color="red", width=3)))
        fig.update_layout(barmode="stack", title="Income vs Expenses", xaxis_title="Age",
                          yaxis_title="$", yaxis_tickformat="$,.0f", height=500)
        st.plotly_chart(fig, use_container_width=True)

        sd_c = ["green" if v >= 0 else "red" for v in df_disp["Surplus_Deficit"]]
        fig2 = go.Figure(go.Bar(x=df_disp["Age"], y=df_disp["Surplus_Deficit"], marker_color=sd_c))
        fig2.update_layout(title="Surplus / Deficit", xaxis_title="Age", yaxis_title="$",
                           yaxis_tickformat="$,.0f", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Legacy Tab ──
    with tabs[ti]:
        st.subheader("Legacy Planning")
        lg1, lg2, lg3 = st.columns(3)
        lg1.metric("Children", f"{int(cfg.get('num_children', 0))}")
        lg2.metric("Legacy Years", f"{int(cfg.get('legacy_years', 0))}")
        if mc_runs:
            avg_bad_years = np.mean([r["Bad_Return_Year"].sum() for r in mc_runs])
            lg3.metric("Avg Bad Return Years (MC)", f"{avg_bad_years:.1f}")
        else:
            bad_ct = int(df['Bad_Return_Year'].sum()) if 'Bad_Return_Year' in df.columns else 0
            lg3.metric("Bad Return Years", f"{bad_ct}")
            st.caption(
                "The baseline plan uses one constant assumed return every year, so it "
                "never has a down year on its own (0 here is expected, not a bug). "
                "Enable Monte Carlo to see how often a bad year actually skips a legacy "
                "gift across randomized market paths."
            )

        n_kids = int(cfg.get("num_children", 0))

        st.markdown("---")
        st.markdown("##### If you died at a given age, what would your kids actually receive?")
        st.caption(
            "Two very different pots make up the answer: money **already gifted** into the "
            "kids' own Roth accounts (irrevocable, compounding on its own), and whatever's "
            "**still sitting in your own accounts** at death (pretax, your own Roth, HSA, "
            "cash) -- which is what actually passes to heirs. 'Cum Legacy Inheritance' further "
            "below is neither of these; see its note."
        )
        min_age, max_age = int(df_disp["Age"].min()), int(df_disp["Age"].max())
        death_age = st.slider("Hypothetical age at death", min_age, max_age, max_age, key="legacy_death_age")
        death_row = df_disp[df_disp["Age"] == death_age].iloc[0]

        d1, d2, d3 = st.columns(3)
        d1.metric("Already Gifted (Kids' Roth, grown)", f"${death_row.get('Legacy_Pool_EOY', 0.0):,.0f}",
                   help="Money already moved into the kids' own Roth accounts by this age. Already theirs, not part of your estate.")
        d2.metric("Remaining in Your Estate", f"${death_row.get('Estate_At_Death', death_row['Total_Liquid_Assets']):,.0f}",
                   help="Whatever's left in your own pretax/Roth/HSA/cash accounts at this age -- this is what passes to heirs at death.")
        d3.metric("Total to Family", f"${death_row.get('Family_Net_Worth', 0.0):,.0f}",
                   help="The two combined.")

        oregon_estate_tax_at_death = death_row.get('Oregon_Estate_Tax', 0.0)
        st.markdown(f"**After Oregon Estate Tax{' (0% -- not an Oregon resident)' if not or_resident else ''} and heir income tax on inherited PreTax/HSA:**")
        af1, af2, af3 = st.columns(3)
        af1.metric("Oregon Estate Tax", f"${oregon_estate_tax_at_death:,.0f}",
                   help="Oregon's estate-tax exemption is $1M -- far below the ~$15M federal exemption, so this is the exposure that actually bites for most Oregon estates this size. Assumes this is the second (surviving) death, since transfers to a spouse are tax-free under the unlimited marital deduction. $0 if not an Oregon resident.")
        af2.metric("Your Estate, After-Tax", f"${death_row.get('After_Tax_Estate_At_Death', 0.0):,.0f}",
                   help="Remaining in Your Estate, minus Oregon Estate Tax and the heir's income tax on inherited PreTax/HSA (Roth, cash, and brokerage pass at full value -- see the Tax tab methodology note).")
        af3.metric("Total to Family, After-Tax", f"${death_row.get('After_Tax_Family_Net_Worth', 0.0):,.0f}",
                   help="After-tax estate plus the already-gifted Roth pool (already tax-free, unaffected).")

        if n_kids > 0:
            st.markdown(f"**Per child (÷ {n_kids}):**")
            e1, e2, e3 = st.columns(3)
            e1.metric("Roth Pool / Child", f"${death_row.get('Legacy_Pool_Per_Child', 0.0):,.0f}")
            e2.metric("Estate / Child", f"${death_row.get('Estate_At_Death_Per_Child', 0.0):,.0f}")
            e3.metric("Total / Child", f"${death_row.get('Total_Inheritance_At_Death_Per_Child', 0.0):,.0f}")
            e4, e5 = st.columns(2)
            e4.metric("Estate / Child, After-Tax", f"${death_row.get('After_Tax_Estate_Per_Child', 0.0):,.0f}")
            e5.metric("Total / Child, After-Tax", f"${death_row.get('After_Tax_Total_Inheritance_Per_Child', 0.0):,.0f}")
        else:
            st.caption("Set Number of Children > 0 to see a per-child split.")

        fig_estate = go.Figure()
        fig_estate.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Estate_At_Death"], name="Remaining Estate (your own accounts)",
                                        line=dict(color="teal", width=0.5), stackgroup="one", fillcolor="rgba(0,128,128,0.4)"))
        fig_estate.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_EOY"], name="Already Gifted (Kids' Roth)",
                                        line=dict(color="purple", width=0.5), stackgroup="one", fillcolor="rgba(128,0,128,0.4)"))
        fig_estate.add_vline(x=death_age, line_dash="dash", line_color="red")
        fig_estate.update_layout(title="What Would Pass to Family, by Age of Death", xaxis_title="Age",
                                  yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_estate, use_container_width=True)

        st.markdown("---")
        st.markdown("##### Legacy Gifting Detail")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Gifted (nominal)", f"${df_disp.iloc[-1]['Cum_Legacy_Roth']:,.0f}",
                   help="Sum of contribution amounts, at the time each was made -- does not include growth.")
        c2.metric("Legacy Pool Today", f"${df_disp.iloc[-1].get('Legacy_Pool_EOY', 0.0):,.0f}",
                   help="What that gifted money has actually grown to, compounding on its own in the kids' Roth accounts. Same number as 'Already Gifted' above.")
        c3.metric("Cum Legacy Inheritance (bad-years only)", f"${df_disp.iloc[-1].get('Cum_Legacy_Inheritance', 0.0):,.0f}",
                   help="NOT your general estate. This is only the sub-total of gift TARGETS that were skipped in down-market years and left invested rather than gifted to Roth. It's $0 whenever the plan never hits a down year (see note above) -- for what actually passes to your heirs at any age, use the 'Remaining in Your Estate' section above instead.")
        c4.metric("Total Legacy Value (Today)", f"${df_disp.iloc[-1].get('Legacy_Value_To_Date', 0.0):,.0f}",
                   help="Legacy Pool + Cum Legacy Inheritance (the bad-years sub-total, not your full estate).")

        # ── Legacy Pool balance over time (the actual compounding balance,
        # not just the annual flow into it) ──
        fig_leg_bal = go.Figure()
        fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_EOY"],
                                          name="Legacy Pool (grown)", line=dict(color="purple", width=3),
                                          fill="tozeroy", fillcolor="rgba(128,0,128,0.1)"))
        fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Cum_Legacy_Roth"],
                                          name="Cumulative Gifted (nominal, no growth)",
                                          line=dict(color="gray", width=2, dash="dot")))
        if 'Cum_Legacy_Inheritance' in df_disp.columns and df_disp['Cum_Legacy_Inheritance'].iloc[-1] > 0:
            fig_leg_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Value_To_Date"],
                                              name="Total Legacy Value (Pool + Inheritance)",
                                              line=dict(color="darkgoldenrod", width=2, dash="dash")))
        fig_leg_bal.update_layout(title="Legacy Pool: Balance vs. Nominal Contributions", xaxis_title="Age",
                                  yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_bal, use_container_width=True)
        st.caption(
            "The gap between the purple and gray lines is investment growth on money already "
            "gifted. This balance is segregated from the household's own accounts -- it "
            "compounds on its own and is never drawn back down to cover the household's "
            "living expenses."
        )

        if n_kids > 0:
            fig_leg_pc_bal = go.Figure()
            fig_leg_pc_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Legacy_Pool_Per_Child"],
                                                 name="Roth Pool / Child (grown)", line=dict(color="purple", width=2.5)))
            fig_leg_pc_bal.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Cum_Legacy_Inheritance_Per_Child"],
                                                 name="Inheritance / Child", line=dict(color="goldenrod", width=2.5)))
            fig_leg_pc_bal.update_layout(title="Per-Child Legacy Value Over Time", xaxis_title="Age",
                                         yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=380)
            st.plotly_chart(fig_leg_pc_bal, use_container_width=True)

        fig_leg_pc = go.Figure()
        fig_leg_pc.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Roth_Per_Child"], name="Roth per Child", marker_color="purple"))
        fig_leg_pc.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Inheritance_Per_Child"], name="Inheritance per Child", marker_color="goldenrod"))
        fig_leg_pc.update_layout(barmode="group", title="Annual Legacy Gift per Child by Age", xaxis_title="Age",
                                 yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_pc, use_container_width=True)

        fig_leg_total = go.Figure()
        fig_leg_total.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Roth_Total"], name="Roth Total", marker_color="mediumpurple"))
        fig_leg_total.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Legacy_Inheritance_Total"], name="Inheritance Total", marker_color="darkgoldenrod"))
        fig_leg_total.update_layout(barmode="stack", title="Annual Legacy Gift Total by Age", xaxis_title="Age",
                                    yaxis_title=dollar_label, yaxis_tickformat="$,.0f", height=420)
        st.plotly_chart(fig_leg_total, use_container_width=True)

        legacy_cols = [c for c in [
            "Age", "Year", "Bad_Return_Year",
            "Estate_At_Death", "Estate_At_Death_Per_Child",
            "Oregon_Estate_Tax", "After_Tax_Estate_At_Death", "After_Tax_Estate_Per_Child",
            "Legacy_Pool_EOY", "Legacy_Pool_Per_Child",
            "Family_Net_Worth", "Total_Inheritance_At_Death_Per_Child",
            "After_Tax_Family_Net_Worth", "After_Tax_Total_Inheritance_Per_Child",
            "Legacy_Target_Per_Child", "Legacy_Roth_Per_Child", "Legacy_Inheritance_Per_Child",
            "Legacy_Target_Total", "Legacy_Roth_Total", "Legacy_Inheritance_Total", "Legacy_Total",
            "Cum_Legacy_Roth", "Cum_Legacy_Inheritance", "Cum_Legacy_Inheritance_Per_Child",
            "Legacy_Value_To_Date", "Legacy_Value_To_Date_Per_Child",
        ] if c in df_disp.columns]
        st.dataframe(df_disp[legacy_cols], use_container_width=True, hide_index=True)
    ti += 1

    # ── Tax Tab ──
    with tabs[ti]:
        ca2, cb2 = st.columns(2)
        with ca2:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Federal_Tax"], name="Federal"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Oregon_Tax"], name="Oregon"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["IRMAA_Cost"], name="IRMAA"))
            fig.update_layout(barmode="stack", title="Tax Burden", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
        with cb2:
            fig = go.Figure(go.Scatter(x=df["Age"], y=df["Effective_Tax_Rate"]*100,
                                       mode="lines+markers", line=dict(color="darkblue", width=2)))
            fig.update_layout(title="Effective Tax Rate", yaxis_title="%", height=400)
            st.plotly_chart(fig, use_container_width=True)

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Federal_Taxable_Income"], name="Taxable", line=dict(color="blue", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_12_Ceiling"], name="12% Ceil", line=dict(color="green", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_22_Ceiling"], name="22% Ceil", line=dict(color="orange", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["MAGI"], name="MAGI (this year)", line=dict(color="red", width=1)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["IRMAA_Lookback_MAGI"], name="MAGI (IRMAA lookback, 2yr prior)", line=dict(color="firebrick", width=1, dash="dashdot")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["IRMAA_Threshold"], name="IRMAA Tier-1 Thr", line=dict(color="darkred", dash="dot")))
        fig.update_layout(title="Income vs Brackets & IRMAA", yaxis_tickformat="$,.0f", height=450)
        st.plotly_chart(fig, use_container_width=True)
        st.caption(
            "IRMAA is assessed off MAGI from 2 tax years prior (SSA/CMS "
            "lookback), compared against the CURRENT year's threshold -- not "
            "this year's own MAGI. The dashed 'IRMAA lookback' line is what "
            "actually determines IRMAA_Cost each year; a spike in this "
            "year's MAGI (solid red) shows up as an IRMAA hit two years "
            "later, not immediately. IRMAA itself is a true 5-tier "
            "staircase (crossing a tier by $1 jumps the WHOLE premium, "
            "unlike an income bracket) -- see IRMAA_Cost for the full "
            "surcharge, not just whether the tier-1 line was crossed."
        )
    ti += 1

    # ── Roth Conversions Tab ──
    with tabs[ti]:
        cdf = df[df["Roth_Conversion"] > 0]
        if len(cdf) > 0:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion"], name="Converted", marker_color="purple"))
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion_Tax"], name="Tax", marker_color="red"))
            fig.update_layout(barmode="group", title="Roth Conversions", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
            tc, tt = cdf["Roth_Conversion"].sum(), cdf["Roth_Conversion_Tax"].sum()
            if tc > 0: st.info(f"**Total:** ${tc:,.0f} | **Tax:** ${tt:,.0f} | **Rate:** {tt/tc:.1%}")
        else:
            st.info("No conversions.")
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["PreTax_EOY"], name="PreTax", line=dict(color="red", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Roth_EOY"], name="Roth", line=dict(color="purple", width=2)))
        fig.update_layout(title="PreTax vs Roth", yaxis_tickformat="$,.0f", height=400)
        st.plotly_chart(fig, use_container_width=True)
    ti += 1

    # ── Monte Carlo Tab ──
    if mc_runs:
        with tabs[ti]:
            surv = sum(1 for r in mc_runs if mc_outcome_row(r)["Total_Liquid_Assets"] > 0)
            survived_runs = [r for r in mc_runs if mc_outcome_row(r)["Total_Liquid_Assets"] > 0]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Sims", len(mc_runs)); m2.metric("Survived", surv); m3.metric("Rate", f"{surv/len(mc_runs)*100:.1f}%")
            if survived_runs:
                base0 = cfg["base_annual_expenses"]
                worst_ratios = np.array([r["Base_Expenses_Real"].min() / base0 for r in survived_runs])
                m4.metric("Worst Real Spending (median, successful paths)",
                          f"{np.percentile(worst_ratios, 50) * 100:.0f}% of Annual Expenses target",
                          help="Among the paths that 'succeeded' (portfolio > $0 at the final age), "
                               "the median of each path's own lowest real (today's $) annual spending "
                               "level, as a % of your Annual Expenses input. For Fixed Real Spending "
                               "that input is the literal target the strategy tries to sustain; VPW "
                               "has no fixed target at all (it deliberately floats with the "
                               "portfolio), so here it's just a reference point, not a goal the "
                               "strategy is failing to hit. A high success rate next to a low number "
                               "here means the strategy is buying that survival by cutting your "
                               "lifestyle, not the portfolio holding up.")

            # ── Draw-Order Recommendation (automatic) ──
            # Runs both PreTax-First and Brokerage-First on matched paths and
            # compares them on TWO dimensions, not one -- rather than asking
            # the user to manually flip the sidebar toggle and re-run to find
            # out. Survival (Monte Carlo success rate) and legacy value
            # (after-tax family net worth) can point in OPPOSITE directions:
            # Brokerage-First protects the retiree's own survival against
            # sequence-of-returns risk by spending Cash/Brokerage down first,
            # but that leaves PreTax compounding longer -- bigger RMDs, and a
            # bigger SECURE-Act income-tax bill for heirs. A recommendation
            # that only looked at survival would silently trade legacy value
            # away without ever saying so -- verified directly on one
            # scenario: +survival, but -$528k after-tax family net worth.
            cmp_cfg = {k: v for k, v in cfg.items() if k != "brokerage_before_pretax"}
            cmp = cached_draw_order_comparison(cmp_cfg, mc_std, mc_max, mc_correlation, mc_fat_tailed, mc_t_df)
            cur_key = "brokerage_first" if cfg.get("brokerage_before_pretax", False) else "pretax_first"
            cur_label = "Brokerage First" if cur_key == "brokerage_first" else "PreTax First"
            alt_key = "pretax_first" if cur_key == "brokerage_first" else "brokerage_first"
            alt_label = "PreTax First" if alt_key == "pretax_first" else "Brokerage First"
            cur, alt = cmp[cur_key], cmp[alt_key]

            surv_gap = alt["success_rate"] - cur["success_rate"]
            fnw_cur, fnw_alt = cur["median_after_tax_family_net_worth"], alt["median_after_tax_family_net_worth"]
            fnw_gap = fnw_alt - fnw_cur
            fnw_gap_pct = (fnw_gap / fnw_cur) if fnw_cur else 0.0
            SURV_TOL, FNW_TOL_PCT = 0.01, 0.02  # 1 point of success rate, 2% of family net worth
            surv_moves = abs(surv_gap) > SURV_TOL
            fnw_moves = abs(fnw_gap_pct) > FNW_TOL_PCT

            surv_line = (f"survival {cur['success_rate']*100:.1f}% -> {alt['success_rate']*100:.1f}% "
                         f"if you switch to {alt_label}")
            fnw_line = (f"after-tax family net worth ${fnw_cur:,.0f} -> ${fnw_alt:,.0f} "
                        f"({'+' if fnw_gap >= 0 else ''}{fnw_gap_pct*100:.1f}%) if you switch to {alt_label}")
            basis = f"Based on {cmp['sims']} matched-path simulations, isolated from the {mc_sims}-sim run above."

            if surv_moves and fnw_moves and (surv_gap > 0) == (fnw_gap > 0):
                # Both metrics agree -- a clean win, safe to recommend directly.
                winner = alt_label if surv_gap > 0 else cur_label
                st.success(
                    f"\U0001F4A1 **Draw-order recommendation:** **{winner}** wins on both fronts "
                    f"for this scenario -- {surv_line}, and {fnw_line}. {basis}"
                )
            elif surv_moves and fnw_moves:
                # They conflict -- a real tradeoff between the retiree's own
                # survival and what heirs end up with. Not this app's call to
                # make for you -- show both numbers, pick neither.
                better_surv = alt_label if surv_gap > 0 else cur_label
                better_fnw = alt_label if fnw_gap > 0 else cur_label
                st.warning(
                    f"⚖️ **Draw-order tradeoff, not a clean win:** {better_surv} improves "
                    f"your OWN survival odds ({surv_line}), but {better_fnw} leaves more for your "
                    f"heirs after tax ({fnw_line}). Brokerage-First tends to win on survival by "
                    f"spending Cash/Brokerage down first (protects against selling equities in an "
                    f"early bad market) while PreTax compounds longer -- but that same compounding "
                    f"means bigger RMDs and a bigger SECURE-Act income-tax bill for heirs later. "
                    f"Which matters more depends on whether you're optimizing for your own plan "
                    f"surviving or for what's left afterward. {basis}"
                )
            elif surv_moves:
                st.success(
                    f"\U0001F4A1 **Draw-order recommendation:** switching to **{alt_label}** would "
                    f"improve {surv_line}, with no meaningful change to after-tax family net worth "
                    f"({fnw_line}). {basis}"
                )
            elif fnw_moves:
                st.info(
                    f"**Legacy-value note:** switching to **{alt_label}** would change {fnw_line}, "
                    f"with no meaningful change to Monte Carlo survival ({surv_line}). {basis}"
                )
            else:
                st.caption(
                    f"Draw-order priority makes little difference for this specific scenario on "
                    f"either survival or after-tax family net worth ({cur_label}: "
                    f"{cur['success_rate']*100:.1f}% success / ${fnw_cur:,.0f} family net worth vs "
                    f"{alt_label}: {alt['success_rate']*100:.1f}% / ${fnw_alt:,.0f}, {cmp['sims']} "
                    f"matched-path simulations) -- typically means Roth Conversions are off/minimal, "
                    f"or there's little Brokerage balance for the reordering to matter."
                )

            if cfg.get("mortality_based_horizon", False):
                death_ages = np.array([r["Death_Age"].iloc[0] for r in mc_runs])
                fig_death = go.Figure(go.Histogram(x=death_ages, marker_color="slategray", nbinsx=30))
                fig_death.add_vline(x=cfg["planning_end_age"], line_dash="dash", line_color="red",
                                    annotation_text="Plan Through Age (fixed-horizon comparison)")
                fig_death.update_layout(title="Drawn Household Death Age Across Simulations (later of two lifespans)",
                                        xaxis_title="Age", height=320)
                st.plotly_chart(fig_death, use_container_width=True)
                st.caption(
                    f"Median drawn household horizon: age {np.median(death_ages):.0f} "
                    f"(vs. the fixed Plan Through Age of {cfg['planning_end_age']} used when this is "
                    "off). Success above is judged at each path's own drawn age, not always the fixed "
                    "one -- so a path that runs out of money at 94 but whose household didn't live "
                    "past 90 still counts as a success. The Optimizer tab's own sweep doesn't use this "
                    "mortality draw (it runs its own faster, reduced-fidelity simulation) -- its "
                    "success rates still reflect the fixed-horizon criterion even when this is on."
                )

            bands = compute_percentile_bands(mc_runs, "Total_Liquid_Assets")
            ages = bands["Age"]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p95"],bands["p5"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.1)", line=dict(color="rgba(255,255,255,0)"), name="5-95%"))
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p75"],bands["p25"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.25)", line=dict(color="rgba(255,255,255,0)"), name="25-75%"))
            fig.add_trace(go.Scatter(x=ages, y=bands["p50"], name="Median", line=dict(color="blue", width=2)))
            fig.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                     line=dict(color="black", width=2, dash="dash")))
            fig.update_layout(title=f"MC Total Assets ({len(mc_runs)} runs)", yaxis_tickformat="$,.0f", height=500)
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("##### What did spending actually look like behind those numbers?")
            bands_exp = compute_percentile_bands(mc_runs, "Base_Expenses_Real")
            fig_exp = go.Figure()
            fig_exp.add_trace(go.Scatter(x=np.concatenate([ages, ages[::-1]]),
                y=np.concatenate([bands_exp["p95"], bands_exp["p5"][::-1]]),
                fill="toself", fillcolor="rgba(231,76,60,0.1)", line=dict(color="rgba(255,255,255,0)"), name="5-95%"))
            fig_exp.add_trace(go.Scatter(x=np.concatenate([ages, ages[::-1]]),
                y=np.concatenate([bands_exp["p75"], bands_exp["p25"][::-1]]),
                fill="toself", fillcolor="rgba(231,76,60,0.25)", line=dict(color="rgba(255,255,255,0)"), name="25-75%"))
            fig_exp.add_trace(go.Scatter(x=ages, y=bands_exp["p50"], name="Median", line=dict(color="darkred", width=2)))
            fig_exp.add_hline(y=cfg["base_annual_expenses"], line_dash="dash", line_color="black",
                               annotation_text="Annual Expenses Target (today's $)")
            fig_exp.update_layout(title="Actual Lifestyle Spending Across Simulations (Today's $)",
                                  yaxis_tickformat="$,.0f", height=420)
            st.plotly_chart(fig_exp, use_container_width=True)
            st.caption(
                "'Success' above only means the portfolio balance stayed above $0 at the final age -- "
                "it says nothing about how much spending had to be cut to get there. This chart shows "
                "the actual lifestyle spending (base expenses, today's dollars) each simulation lived "
                "on, against your Annual Expenses input as a reference line (not a target VPW is "
                "trying to hit -- it deliberately floats with the portfolio; the line is there so you "
                "can judge the gap either way). If this band sags far below that line while the "
                "success rate still looks good, the strategy is buying survival by cutting your "
                "lifestyle, not because the portfolio actually held up -- consider setting an Absolute "
                "Minimum Annual Expenses floor (Expenses tab), or a lower Max Annual Expenses ceiling "
                "if the concern runs the other way and spending is instead being pushed too high."
            )

            if cfg.get("inflation_std", 0.0) > 0:
                bands_infl = compute_percentile_bands(mc_runs, "Inflation_Rate_Realized")
                fig_infl = go.Figure()
                fig_infl.add_trace(go.Scatter(x=ages, y=bands_infl["p95"], line=dict(width=0), showlegend=False))
                fig_infl.add_trace(go.Scatter(x=ages, y=bands_infl["p5"], fill="tonexty",
                    fillcolor="rgba(230,126,34,0.15)", line=dict(width=0), name="5-95%"))
                fig_infl.add_trace(go.Scatter(
                    x=np.concatenate([ages, ages[::-1]]),
                    y=np.concatenate([bands_infl["p75"], bands_infl["p25"][::-1]]),
                    fill="toself", fillcolor="rgba(230,126,34,0.3)", line=dict(color="rgba(255,255,255,0)"), name="25-75%"))
                fig_infl.add_trace(go.Scatter(x=ages, y=bands_infl["p50"], name="Median", line=dict(color="chocolate", width=2)))
                fig_infl.add_hline(y=cfg["inflation_rate"], line_dash="dash", line_color="black",
                                   annotation_text="General Inflation input")
                fig_infl.update_layout(title="Realized Inflation Rate Across Simulations",
                                       yaxis_tickformat=".1%", height=350)
                st.plotly_chart(fig_infl, use_container_width=True)
                st.caption(
                    "Each simulation draws its own year-by-year inflation path instead of using the "
                    "flat General Inflation rate everywhere -- correlated (negatively, by default) to "
                    "that path's own equity returns via the Inflation-Equity Correlation slider "
                    "(Monte Carlo Settings). This is what drives the spread above and below the dashed "
                    "input line; every other chart on this tab already reflects it, since expenses, "
                    "the absolute floor/ceiling, and today's-dollar conversions all compound off "
                    "this same per-path series."
                )

            n_show = min(20, len(mc_runs))
            fig3 = go.Figure()
            for i in range(n_show):
                fig3.add_trace(go.Scatter(x=mc_runs[i]["Age"], y=mc_runs[i]["Total_Liquid_Assets"],
                    mode="lines", line=dict(width=0.7, color="rgba(99,110,250,0.3)"), showlegend=False))
            fig3.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                      line=dict(color="black", width=2, dash="dash")))
            fig3.update_layout(title=f"{n_show} Sample Paths", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig3, use_container_width=True)
        ti += 1

    # ── Optimizer Tab ──
    if mc_runs:
        with tabs[ti]:
            st.subheader("\U0001F3AF Scenario Optimizer — Path to 100% Success")
            st.caption(
                "Sweeps retirement age delay, SS start age, and post-80 expense reduction "
                "to find combinations that maximize Monte Carlo success rate."
            )

            if st.button("Run Optimizer", type="primary"):
                with st.spinner("Running scenario sweep (this takes a moment)..."):
                    opt_df = cached_run_optimizer(cfg, mc_sims, mc_std, mc_max, mc_seed,
                                                   mc_correlation, mc_fat_tailed, mc_t_df)

                # ── Current scenario baseline ──
                surv = sum(1 for r in mc_runs if mc_outcome_row(r)["Total_Liquid_Assets"] > 0)
                base_rate = surv / len(mc_runs) * 100
                st.info(f"**Current settings:** Retire {cfg['retirement_age']}, "
                        f"SS @ {cfg['ss_start_age']}, "
                        f"No post-80 cut → **{base_rate:.0f}% success**")

                # ── Top recommendations ──
                perfect = opt_df[opt_df["Success_Rate"] >= 1.0]
                if len(perfect) > 0:
                    st.success(f"\u2705 Found **{len(perfect)}** scenarios with 100% success rate!")
                    st.markdown("**Top 5 — least disruption to reach 100%:**")
                    # Sort by least change needed
                    perfect_sorted = perfect.sort_values(
                        ["Extra_Work_Years", "SS_Delay_Years", "Expense_Cut_80+"]
                    ).head(5)
                    for _, row in perfect_sorted.iterrows():
                        changes = []
                        if row["Extra_Work_Years"] > 0:
                            changes.append(f"retire at **{int(row['Retire_Age'])}** (+{int(row['Extra_Work_Years'])}yr)")
                        if row["SS_Delay_Years"] > 0:
                            changes.append(f"SS at **{int(row['SS_Age'])}** (+{int(row['SS_Delay_Years'])}yr)")
                        if row["Expense_Cut_80+"] > 0:
                            changes.append(f"**{row['Expense_Cut_80+']*100:.0f}%** expense cut after 80")
                        med_val = row["Median_Final_Real"] if show_real else row["Median_Final_Assets"]
                        label = "today's $" if show_real else "future $"
                        desc = " + ".join(changes) if changes else "No changes needed!"
                        st.markdown(f"- {desc} → Median final: **${med_val:,.0f}** ({label})")
                else:
                    st.warning("No 100% success scenarios found. Best options below.")

                # ── Full results table ──
                st.subheader("All Scenarios")
                display_opt = opt_df.copy()
                display_opt["Success_Rate"] = display_opt["Success_Rate"].apply(lambda x: f"{x:.0%}")
                display_opt["Expense_Cut_80+"] = display_opt["Expense_Cut_80+"].apply(lambda x: f"{x:.0%}")
                money_opt_cols = ["Median_Final_Assets", "Median_Final_Real", "P10_Final_Assets", "Worst_Case"]
                for c in money_opt_cols:
                    display_opt[c] = display_opt[c].apply(lambda x: f"${x:,.0f}")
                st.dataframe(display_opt, use_container_width=True, height=400)
            else:
                st.markdown(
                    "Click **Run Optimizer** to sweep combinations of:\n"
                    "- Retirement age: current → +3 years\n"
                    "- SS start age: current → +4 years\n"
                    "- Post-80 expense reduction: 0% to 25%\n\n"
                    "Uses 50 MC sims per scenario for speed."
                )
        ti += 1

    # ── Sensitivity Tab ──
    with tabs[ti]:
        st.subheader("\U0001F32A Sensitivity (Tornado) Analysis")
        st.caption(
            "Varies ONE assumption at a time -- holding everything else at your current "
            "settings -- and shows how much your final-year value (today's $) moves. Sorted "
            "so the biggest driver of your outcome is at the top."
        )
        if st.button("Run Sensitivity Analysis", type="primary"):
            with st.spinner("Running sensitivity sweep..."):
                sens_df, base_final = run_sensitivity_analysis(cfg)

            fig = go.Figure()
            fig.add_trace(go.Bar(
                y=sens_df["Variable"], x=sens_df["High_Result"] - sens_df["Low_Result"],
                base=sens_df["Low_Result"], orientation="h",
                marker_color="rgba(99,110,250,0.6)",
                text=[f"${lo:,.0f} \u2192 ${hi:,.0f}" for lo, hi in zip(sens_df["Low_Result"], sens_df["High_Result"])],
                textposition="inside", insidetextanchor="middle",
                textfont=dict(color="white"),
            ))
            fig.add_vline(x=base_final, line_dash="dash", line_color="black")
            fig.update_layout(
                title=f"Impact on Final-Year Value ({dollar_label}) \u2014 Baseline: ${base_final:,.0f}",
                xaxis_title=dollar_label, xaxis_tickformat="$,.0f",
                height=120 + 45 * len(sens_df), margin=dict(l=10),
            )
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("**Range tested per variable** (\u00b1 around your current setting):")
            disp_sens = sens_df[["Variable", "Low_Value", "Base_Value", "High_Value",
                                  "Low_Result", "High_Result", "Impact_Range"]].copy()
            for c in ["Low_Result", "High_Result", "Impact_Range"]:
                disp_sens[c] = disp_sens[c].apply(lambda x: f"${x:,.0f}")
            for c in ["Low_Value", "Base_Value", "High_Value"]:
                disp_sens[c] = disp_sens[c].apply(
                    lambda x: f"{x:.1%}" if isinstance(x, float) and abs(x) < 1 else f"{x:,.0f}"
                )
            st.dataframe(disp_sens.iloc[::-1], use_container_width=True, hide_index=True)
            st.caption(
                "Note: Retirement Age and Base Annual Expenses change the shape of the whole "
                "plan (years worked, years drawing down), so their bars aren't perfectly "
                "apples-to-apples with pure return/rate assumptions -- but the ranking still "
                "tells you where a planning decision matters most."
            )
        else:
            st.info("Click **Run Sensitivity Analysis** to see which assumptions matter most to your outcome.")
        ti += 1

    # ── Full Data Tab ──
    with tabs[ti]:
        st.caption(f"Displaying in **{dollar_label}**")
        dcols = ["Age","Year","Draw_Strategy",
                 "SS_Income","JSS_Income","Rental_Income","S_Plus_Income","Passive_Income",
                 "Base_Expenses","Base_Expenses_Real","VPW_Percentage",
                 "PreTax_Draw","Roth_Draw","Cash_Draw","Brokerage_Draw","HSA_Draw",
                 "PreTax_EOY","Roth_EOY","HSA_EOY","Cash_EOY","Brokerage_EOY","Brokerage_LTCG_Gain","Brokerage_Dividend_Taxable",
                 "Total_Liquid_Assets","Family_Net_Worth","Total_Income","Total_Expenses","Total_Tax",
                 "Effective_Tax_Rate","Withdrawal_Rate","Surplus_Deficit",
                 "Roth_Conversion","RMD","RMD_Excess","IRMAA_Hit","NIIT","Bad_Return_Year","Inflation_Rate_Realized",
                 "Legacy_Target_Per_Child","Legacy_Roth_Per_Child","Legacy_Inheritance_Per_Child",
                 "Legacy_Target_Total","Legacy_Roth_Total","Legacy_Inheritance_Total","Legacy_Total",
                 "Cum_Gifts","Cum_Legacy_Roth","Legacy_Pool_EOY","Cum_Legacy_Inheritance","Cum_Lump_Sums",
                 "Estate_At_Death"]
        avail = [c for c in dcols if c in df_disp.columns]
        show = st.multiselect("Columns", df_disp.columns.tolist(), default=avail)
        disp = df_disp[show].copy()
        no_fmt = {"Age","Year","Years_Retired","Effective_Tax_Rate","Withdrawal_Rate",
                  "IRMAA_Hit","Bad_Return_Year","Draw_Strategy","Absolute_Min_Bound_Hit","Max_Expenses_Bound_Hit",
                  "Return_PreTax","Return_Roth","Return_HSA","Return_Cash","Phase"}
        pct = {"Effective_Tax_Rate","Withdrawal_Rate","Return_PreTax","Return_Roth","Return_HSA","Return_Cash","VPW_Percentage","Inflation_Rate_Realized"}
        fmt = {}
        for c in show:
            if c in pct: fmt[c] = "{:.1%}"
            elif c not in no_fmt: fmt[c] = "${:,.0f}"
        st.dataframe(disp.style.format(fmt, na_rep="").apply(
            lambda x: ["background-color:#ffcccc" if isinstance(v,(int,float)) and v<0 else "" for v in x], axis=1),
            use_container_width=True, height=600)

        # ── DOWNLOADS ──
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("\U0001F4E5 Download CSV", df.to_csv(index=False), "retirement_plan.csv", "text/csv")
        with col_dl2:
            xlsx_buf = export_to_excel(accum_df, df, cfg, mc_runs=mc_runs)
            st.download_button("\U0001F4E5 Download Excel Report", xlsx_buf,
                               "retirement_plan.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Notes ──
    st.divider()
    with st.expander("Model Notes"):
        st.markdown("""
**Accumulation Phase:** Balances grow from current age to retirement using configured returns. Contributions include 401(k), Roth 401(k), Roth IRA (x2 MFJ), HSA, Mega Backdoor Roth, employer match, cash savings, and a final-year lump sum. S+ deferred comp also grows during employment.

**Draw Strategy:** Pre-RMD: PreTax up to 12% bracket → Cash → Roth → HSA. RMD Phase (75+): mandatory RMD first → Cash → Roth → HSA. Roth conversions fill remaining bracket space.

**Monte Carlo:** Returns ~ N(target, std_dev) by default, or Student-t if Fat-Tailed is enabled, capped at max upside. PreTax/Roth/HSA/Legacy Pool share a common market factor (Cross-Account Correlation slider); Cash/MM gets a smaller fraction of it and 1/3 the std dev.

**Legacy Gifting:** In a year with positive/neutral returns, the target Roth gift is withdrawn once (via the normal draw waterfall) and credited to a segregated Legacy Pool balance that compounds on its own return/std-dev assumption and is never drawn back down for household spending. In a down-market year, the gift is skipped entirely and the target amount simply stays invested — it isn't withdrawn, and "Legacy Inheritance" is a label for that still-invested amount, not a separate pot of money.

**After-Tax to Heirs:** Inherited pretax/HSA balances must be distributed within 10 years (SECURE Act) and are taxed at an assumed heir rate; Roth (including the Legacy Pool) passes tax-free. The Legacy tab shows both nominal and after-tax figures since a dollar in each bucket is not worth the same to your heirs.

**Brokerage Dividend Tax Drag:** Unlike PreTax/HSA (fully deferred) and the brokerage account's own price appreciation (deferred until you sell), a taxable brokerage account's dividend distribution is taxed the year it's paid whether or not you spend it. Modeled as a fixed % of the account's balance each year (Performance Assumptions), taxed at the same preferential rate as long-term capital gains and assumed reinvested -- which raises the account's cost basis by the same amount, real IRS treatment for a reinvested dividend. Only applied during retirement; the accumulation phase isn't tax-modeled at all (a broader simplification, not specific to dividends).

**Surviving Spouse Scenario:** When enabled, filing status switches from MFJ to Single starting the year after the configured death age (the IRS allows MFJ in the year of death itself), pulling in single-filer federal brackets, a halved standard deduction, a halved IRMAA threshold, single Social Security taxability thresholds, and a configurable partial SS survivor benefit, pension survivor benefit, and living-expense reduction.

**Spending Strategy:** Fixed Real Spending inflates your base expenses every year. Variable Percentage Withdrawal (VPW, default) instead spends a percentage of your CURRENT total portfolio balance every year -- recalculated fresh annually using the same amortization math as an RMD divisor, with the percentage rising as you age -- so it self-corrects immediately with portfolio performance instead of carrying forward a permanent step. Popularized by the Bogleheads community. Both still respect the Absolute Minimum Annual Expenses floor as a hard backstop.

**Spending Ceiling:** A hard maximum on base spending, in today's dollars (defaults to your Annual Expenses figure). Matters most for VPW, whose withdrawal percentage accelerates toward the end of the plan -- without a ceiling, the model would draw down and "spend" whatever the raw percentage computes even in years nobody would realistically consume that much. Money above the ceiling simply stays invested rather than being withdrawn at all, so it isn't drawn-then-unaccounted-for. A no-op for Fixed Real Spending, which never exceeds its own planned amount.

**Cash Shortfalls:** Cash is allowed to go negative to represent a genuine funding gap; it is not floored to $0, so Total_Liquid_Assets and Monte Carlo success rates reflect real shortfalls rather than hiding them.

**Discretionary Reduction:** The "if Year Return < 0" cut is entered in today's dollars and inflates every year alongside base expenses, so it represents an equivalent real cut whether a bad year hits early or late in the plan.

**Excel Export:** 4-sheet workbook: Assumptions, Accumulation, Retirement, Summary -- plus a 5th Monte Carlo sheet (percentile bands + success rate) whenever Monte Carlo is enabled.
        """)


if __name__ == "__main__":
    main()
