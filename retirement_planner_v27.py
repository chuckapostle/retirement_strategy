"""
Retirement Income and Tax Planning Simulator v5
================================================
With accumulation phase, Monte Carlo, and Excel export.

Usage:
  pip install streamlit plotly pandas numpy openpyxl
  streamlit run retirement_planner.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from typing import Dict, List, Tuple
import io
import os
from pathlib import Path

# ============================================================
# CONSTANTS
# ============================================================

RMD_TABLE = {
    72: 27.4, 73: 26.5, 74: 25.5, 75: 24.6, 76: 23.7, 77: 22.9,
    78: 22.0, 79: 21.1, 80: 20.2, 81: 19.4, 82: 18.5, 83: 17.7,
    84: 16.8, 85: 16.0, 86: 15.2, 87: 14.4, 88: 13.7, 89: 12.9,
    90: 12.2, 91: 11.5, 92: 10.8, 93: 10.1, 94: 9.5, 95: 8.9,
    96: 8.4, 97: 7.8, 98: 7.3, 99: 6.8, 100: 6.4,
}
IRMAA_MAGI_THRESHOLD = 206_000
IRMAA_MONTHLY_SURCHARGE = 230.80
FEDERAL_BRACKETS = [
    (0.10, 24_800), (0.12, 100_800), (0.22, 211_400),
    (0.24, 403_550), (0.32, 512_450), (0.35, 768_700),
    (0.37, float("inf")),
]
OREGON_BRACKETS = [
    (0.0475, 8_824), (0.0675, 22_059),
    (0.0875, 250_000), (0.0990, 10_000_000),
]

# ============================================================
# TAX ENGINE
# ============================================================

def calc_tax(taxable, brackets, yrs, infl):
    if taxable <= 0: return 0.0
    tax, prev = 0.0, 0.0
    for rate, ceil in brackets:
        ac = ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
        band = max(0.0, min(taxable, ac) - prev)
        tax += band * rate; prev = ac
        if taxable <= ac: break
    return tax

def bracket_ceiling(brackets, target_rate, yrs, infl):
    for rate, ceil in brackets:
        if rate >= target_rate:
            return ceil * (1 + infl) ** yrs if ceil != float("inf") else float("inf")
    return 0.0

def ss_taxable_portion(ss, other):
    if ss <= 0: return 0.0
    prov = other + ss * 0.5
    if prov < 32_000: return 0.0
    elif prov < 44_000: return min(ss * 0.5, (prov - 32_000) * 0.5)
    else: return min(ss * 0.85, 6_000 + (prov - 44_000) * 0.85)

def get_rmd(bal, age, start):
    if age < start or bal <= 0: return 0.0
    f = RMD_TABLE.get(age, 8.9)
    return bal / f if f > 0 else 0.0

def generate_returns(target, std, mx, n, rng):
    return np.clip(rng.normal(loc=target, scale=std, size=n), -1.0, mx)


def load_starting_balances(path="starting_balances.txt"):
    defaults = {
        "401k": None,
        "roth": None,
        "hsa": None,
        "s_plus_5": None,
        "s_plus_10": None,
        "cash": None,
    }
    if not Path(path).exists():
        return defaults, False
    aliases = {
        "401k": "401k",
        "401(k)": "401k",
        "pretax": "401k",
        "roth": "roth",
        "rothira": "roth",
        "roth_ira": "roth",
        "hsa": "hsa",
        "s+5": "s_plus_5",
        "s5": "s_plus_5",
        "s_plus_5": "s_plus_5",
        "s+10": "s_plus_10",
        "s10": "s_plus_10",
        "s_plus_10": "s_plus_10",
        "cash": "cash",
    }
    loaded = False
    try:
        for raw in Path(path).read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.replace(":", " ").split()
            if len(parts) < 2:
                continue
            key = parts[0].strip().lower().replace(" ", "").replace("-", "").replace("/", "")
            key = aliases.get(key)
            if key is None:
                continue
            try:
                defaults[key] = int(float(parts[1].replace(",", "").replace("$", "")))
                loaded = True
            except ValueError:
                pass
    except Exception:
        return defaults, False
    return defaults, loaded
# ============================================================
# ACCUMULATION PHASE
# ============================================================

def run_accumulation(cfg):
    """Grow balances from current age to retirement with contributions."""
    cur_age = cfg["current_age"]
    ret_age = cfg["retirement_age"]
    base_year = 2024
    years = ret_age - cur_age
    if years <= 0:
        return [], cfg["pretax_401k"], cfg["roth_ira"], cfg["hsa"], cfg["cash"], cfg["s_plus_5yr"], cfg["s_plus_10yr"]

    pt = float(cfg["pretax_401k"])
    ro = float(cfg["roth_ira"])
    hs = float(cfg["hsa"])
    ca = float(cfg["cash"])
    s5 = float(cfg["s_plus_5yr"])
    s10 = float(cfg["s_plus_10yr"])

    # Annual contributions
    c_401k = cfg.get("contrib_401k", 24_500)
    c_roth401k = cfg.get("contrib_roth401k", 8_000)
    c_roth_ira = cfg.get("contrib_roth_ira", 8_700)  # x2 for MFJ
    c_hsa = cfg.get("contrib_hsa", 8_700)
    c_mega_backdoor = cfg.get("contrib_mega_backdoor", 29_000)
    c_employer_match = cfg.get("contrib_employer_match", 18_000)
    c_cash_annual = cfg.get("contrib_cash_annual", 150_000)
    c_cash_final_lump = cfg.get("contrib_cash_final_lump", 120_000)

    # Growth rates (same as retirement performance assumptions)
    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]

    rows = []
    for i in range(years):
        age = cur_age + i
        yr = base_year + i

        # Growth first (skip year 0 — balances are "today")
        if i > 0:
            pt *= (1 + pr_pt)
            ro *= (1 + pr_ro)
            hs *= (1 + pr_hs)
            ca *= (1 + pr_ca)
            # S+ grows in deferred comp while employed
            s5 *= (1 + pr_pt)
            s10 *= (1 + pr_pt)

        # Contributions (end of year)
        pt += c_401k + c_employer_match
        ro += c_roth401k + c_roth_ira * 2 + c_mega_backdoor  # 2x Roth IRA for MFJ
        hs += c_hsa
        ca += c_cash_annual

        # Final year lump sum
        if i == years - 1:
            ca += c_cash_final_lump

        rows.append({
            "Phase": "Accumulation",
            "Age": age, "Year": yr,
            "PreTax_EOY": pt, "Roth_EOY": ro, "HSA_EOY": hs, "Cash_EOY": ca,
            "S_Plus_5yr": s5, "S_Plus_10yr": s10,
            "Contrib_PreTax": c_401k + c_employer_match,
            "Contrib_Roth": c_roth401k + c_roth_ira * 2 + c_mega_backdoor,
            "Contrib_HSA": c_hsa,
            "Contrib_Cash": c_cash_annual + (c_cash_final_lump if i == years - 1 else 0),
            "Total_Liquid_Assets": pt + ro + hs + ca,
        })

    return rows, pt, ro, hs, ca, s5, s10


# ============================================================
# RETIREMENT SIMULATION ENGINE
# ============================================================

def run_simulation(cfg, return_overrides=None):
    results = []
    base_year = 2024
    ret_age = cfg["retirement_age"]
    cur_age = cfg["current_age"]
    end_age = cfg["planning_end_age"]
    ret_year = base_year + (ret_age - cur_age)
    infl = cfg["inflation_rate"]
    binfl = cfg["bracket_inflation"]
    num_years = end_age - ret_age + 1

    # Run accumulation phase to get retirement-day balances
    accum_rows, pt, ro, hs, ca, s5_bal, s10_bal = run_accumulation(cfg)
    cash_basis = float(accum_rows[-1].get("Cash_Basis", cfg["cash"])) if accum_rows else float(cfg["cash"])

    # S+ payout tracking
    s5_annual = s5_bal / 5.0 if s5_bal > 0 else 0.0
    s10_annual = s10_bal / 10.0 if s10_bal > 0 else 0.0
    s5_rem, s10_rem = 5, 10
    jss_rec_rem = cfg["jss_recovery_years"]

    pr_pt = cfg["pretax_return"]
    pr_ro = cfg["roth_return"]
    pr_hs = cfg["hsa_return"]
    pr_ca = cfg["cash_return"]

    cum_gifts, cum_legacy_roth, cum_lump_sums = 0.0, 0.0, 0.0
    pt_depleted = ro_depleted = hs_depleted = ca_depleted = False

    for idx in range(num_years):
        age = ret_age + idx
        yr = ret_year + idx
        yfb = yr - base_year
        yir = idx

        row = {"Phase": "Retirement", "Age": age, "Year": yr, "Years_Retired": yir}

        # ── GROWTH ──
        if yir > 0:
            if return_overrides:
                r_pt, r_ro = return_overrides["pretax"][yir], return_overrides["roth"][yir]
                r_hs, r_ca = return_overrides["hsa"][yir], return_overrides["cash"][yir]
            else:
                r_pt, r_ro, r_hs, r_ca = pr_pt, pr_ro, pr_hs, pr_ca
            if not pt_depleted: pt *= (1 + r_pt)
            if not ro_depleted: ro *= (1 + r_ro)
            if not hs_depleted: hs *= (1 + r_hs)
            if not ca_depleted: ca *= (1 + r_ca)
            row["Return_PreTax"], row["Return_Roth"] = r_pt, r_ro
            row["Return_HSA"], row["Return_Cash"] = r_hs, r_ca
        else:
            row["Return_PreTax"] = row["Return_Roth"] = row["Return_HSA"] = row["Return_Cash"] = 0.0

        # ── INCOME SOURCES ──
        sp_inc = 0.0
        if s5_rem > 0: sp_inc += s5_annual; s5_rem -= 1
        if s10_rem > 0: sp_inc += s10_annual; s10_rem -= 1
        row["S_Plus_Income"] = sp_inc

        jss_inc = jss_tax = 0.0
        if age >= cfg["jss_start_age"]:
            jss_inc = cfg["jss_annual_amount"] * (1 + cfg["jss_cola"]) ** (age - cfg["jss_start_age"])
            if jss_rec_rem > 0: jss_rec_rem -= 1
            else: jss_tax = jss_inc
        row["JSS_Income"], row["JSS_Taxable"] = jss_inc, jss_tax

        ss_inc = 0.0
        if age >= cfg["ss_start_age"]:
            ss_inc = cfg["ss_annual_amount"] * (1 + cfg["ss_cola"]) ** (age - cfg["ss_start_age"])
        row["SS_Income"] = ss_inc

        rent_inc = cfg["rental_gross"] * (1 + infl) ** yir
        rent_tax = rent_inc * cfg["rental_taxable_pct"]
        row["Rental_Income"], row["Rental_Taxable"] = rent_inc, rent_tax

        # ── EXPENSES ──
        base_exp = cfg["base_annual_expenses"] * (1 + infl) ** yir
        # Post-80 expense reduction (lifestyle slowdown)
        if age >= 80:
            base_exp *= (1 - cfg.get("expense_reduction_post80", 0.0))
        healthcare = (cfg["healthcare_pre_medicare"] if age < 65 else cfg["healthcare_post_medicare"]) * (1 + cfg["healthcare_inflation"]) ** yir
        hdhp = cfg["hdhp_annual"] * (1 + infl) ** yir if age < 65 else 0.0
        gifts = cfg["gifts_annual"] * (1 + infl) ** yir
        lump = cfg.get("lump_sums", {}).get(age, 0.0)
        roth_leg = 0.0
        if yir < cfg.get("legacy_years", 10):
            roth_leg = cfg["num_children"] * cfg["roth_legacy_per_child"] * (1 + cfg["legacy_inflation"]) ** yir
        total_exp = base_exp + healthcare + hdhp + gifts + lump + roth_leg

        # ── DISCRETIONARY REDUCTION ON NEGATIVE RETURN ──
        # Use pretax return as proxy for overall market return in that year
        is_negative_return = False
        if return_overrides and yir < len(return_overrides["pretax"]):
            if return_overrides["pretax"][yir] < 0:
                is_negative_return = True
        elif cfg["pretax_return"] < 0:
            is_negative_return = True

        if is_negative_return:
            reduction = cfg.get("neg_ret_draw_reduction", 0.0)
            # Cannot reduce expenses below 0, but total_exp should be large enough
            reduction_applied = min(reduction, total_exp)
            total_exp -= reduction_applied
            row["Discretionary_Reduction"] = reduction_applied
        else:
            row["Discretionary_Reduction"] = 0.0

        row["Base_Expenses"], row["Healthcare_Cost"], row["HDHP"] = base_exp, healthcare, hdhp
        row["Gifts"], row["Lump_Sum"], row["Legacy_Roth"] = gifts, lump, roth_leg
        row["Total_Expenses"] = total_exp

        passive = sp_inc + jss_inc + ss_inc + rent_inc
        row["Passive_Income"] = passive

        rmd = get_rmd(pt, age, cfg["rmd_start_age"])
        row["RMD"] = rmd

        # ── BRACKET-OPTIMIZED DRAW STRATEGY ──
        ptd = rod = hsd = cad = 0.0
        need = max(0.0, total_exp - passive)
        std_ded_est = cfg["standard_deduction"] * (1 + binfl) ** yfb
        base_taxable = jss_tax + rent_tax + sp_inc
        sst_est = ss_taxable_portion(ss_inc, base_taxable)
        existing_taxable = base_taxable + sst_est
        br12_gross = bracket_ceiling(FEDERAL_BRACKETS, 0.12, yfb, binfl)
        pretax_room_12 = max(0.0, br12_gross + std_ded_est - existing_taxable)
        is_rmd_phase = age >= cfg["rmd_start_age"]

        if is_rmd_phase:
            if not pt_depleted: ptd = min(rmd, pt); need = max(0.0, need - ptd)
            if need > 0 and not ca_depleted and ca > 0: d = min(need, ca); cad += d; need -= d
            if need > 0 and not ro_depleted and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and not hs_depleted and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "RMD-Dominated"
        else:
            if need > 0 and not pt_depleted and pt > 0:
                d = min(need, pretax_room_12, pt); ptd += d; need -= d
            if need > 0 and not ca_depleted and ca > 0:
                mx = ca * pr_ca if cfg["performance_draw_only"] else ca
                d = min(need, mx); cad += d; need -= d
            if need > 0 and not ro_depleted and ro > 0: d = min(need, ro); rod += d; need -= d
            if need > 0 and not hs_depleted and hs > 0: d = min(need, hs); hsd += d; need -= d
            row["Draw_Strategy"] = "Bracket-Optimized"

        if need > 0:
            for nm, avail, dep in [("pretax", pt-ptd, pt_depleted), ("cash", ca-cad, ca_depleted),
                                   ("roth", ro-rod, ro_depleted), ("hsa", hs-hsd, hs_depleted)]:
                if need <= 0 or dep: continue
                d = min(need, max(0.0, avail))
                if nm == "pretax": ptd += d
                elif nm == "cash": cad += d
                elif nm == "roth": rod += d
                elif nm == "hsa": hsd += d
                need -= d

        row["PreTax_Draw"], row["Roth_Draw"] = ptd, rod
        row["HSA_Draw"], row["Cash_Draw"] = hsd, cad
        row["RMD_Excess"] = max(0.0, rmd - (total_exp - passive)) if is_rmd_phase else 0.0

        # ── TAX ──
        other_taxable = ptd + jss_tax + rent_tax + sp_inc
        sst = ss_taxable_portion(ss_inc, other_taxable)
        gross_taxable = other_taxable + sst
        std_ded = cfg["standard_deduction"] * (1 + binfl) ** yfb
        med_ded = max(0.0, healthcare - 0.075 * gross_taxable)
        item_ded = med_ded + hdhp + (jss_inc - jss_tax)
        best_ded = max(std_ded, item_ded)
        fed_taxable = max(0.0, gross_taxable - best_ded)
        fed_tax = calc_tax(fed_taxable, FEDERAL_BRACKETS, yfb, binfl)
        or_taxable = max(0.0, gross_taxable - sst - ss_inc * 0.15 - best_ded)
        or_tax = calc_tax(or_taxable, OREGON_BRACKETS, yfb, binfl) if cfg["oregon_resident"] else 0.0
        total_tax = fed_tax + or_tax

        row["Gross_Taxable_Income"], row["Federal_Taxable_Income"] = gross_taxable, fed_taxable
        row["Standard_Deduction"], row["Itemized_Deduction"], row["Best_Deduction"] = std_ded, item_ded, best_ded
        row["Deduction_Type"] = "Itemized" if item_ded > std_ded else "Standard"
        row["Federal_Tax"], row["Oregon_Tax"], row["Total_Tax"] = fed_tax, or_tax, total_tax
        row["Effective_Tax_Rate"] = total_tax / gross_taxable if gross_taxable > 0 else 0.0

        magi = gross_taxable + (ss_inc - sst)
        irmaa_thr = IRMAA_MAGI_THRESHOLD * (1 + binfl) ** yfb
        irmaa_hit = age >= 65 and magi > irmaa_thr
        irmaa_cost = (IRMAA_MONTHLY_SURCHARGE * 12 * 2 * (1 + infl) ** yfb) if irmaa_hit else 0.0
        row["MAGI"], row["IRMAA_Threshold"] = magi, irmaa_thr
        row["IRMAA_Hit"], row["IRMAA_Cost"] = irmaa_hit, irmaa_cost
        fpl_700 = 7 * 20_440 * (1 + infl) ** yfb
        row["FPL_700"], row["Under_700_FPL"] = fpl_700, magi < fpl_700

        # ── ROTH CONVERSION ──
        roth_conv_amt = roth_conv_tax = 0.0
        if cfg["roth_conversion_enabled"] and pt - ptd > cfg["roth_conversion_margin"]:
            tgt = 0.12 if cfg["roth_conversion_target_bracket"] == "12%" else 0.22
            conv_room = max(0.0, bracket_ceiling(FEDERAL_BRACKETS, tgt, yfb, binfl) - fed_taxable)
            if cfg["irmaa_avoidance"] and age >= 63:
                conv_room = min(conv_room, max(0.0, irmaa_thr - magi))
            roth_conv_amt = min(conv_room, max(0.0, pt - ptd - cfg["roth_conversion_margin"]))
            if roth_conv_amt > 0:
                fc = calc_tax(fed_taxable + roth_conv_amt, FEDERAL_BRACKETS, yfb, binfl) - fed_tax
                oc = roth_conv_amt * 0.09 if cfg["oregon_resident"] else 0.0
                roth_conv_tax = fc + oc
        row["Roth_Conversion"], row["Roth_Conversion_Tax"] = roth_conv_amt, roth_conv_tax

        # ── UPDATE BALANCES ──
        pt -= (ptd + roth_conv_amt); ro -= rod; ro += roth_conv_amt; hs -= hsd; ca -= cad
        total_income = passive + ptd + rod + hsd + cad
        surplus = total_income - total_exp - total_tax - roth_conv_tax - irmaa_cost
        ca += surplus

        legacy_actual = 0.0
        if roth_leg > 0 and cfg["num_children"] > 0:
            ro += roth_leg; ca -= roth_leg; legacy_actual = roth_leg

        cum_gifts += gifts; cum_legacy_roth += legacy_actual; cum_lump_sums += lump
        row["Surplus_Deficit"], row["Total_Income"] = surplus, total_income
        row["Legacy_Roth_Actual"] = legacy_actual
        row["Cum_Gifts"], row["Cum_Legacy_Roth"], row["Cum_Lump_Sums"] = cum_gifts, cum_legacy_roth, cum_lump_sums

        if pt <= 0: pt = 0.0; pt_depleted = True
        if ro <= 0: ro = 0.0; ro_depleted = True
        if hs <= 0: hs = 0.0; hs_depleted = True
        if ca < -50_000: ca_depleted = True

        row["PreTax_EOY"], row["Roth_EOY"], row["HSA_EOY"] = pt, ro, hs
        row["Cash_EOY"] = max(0.0, ca) if ca_depleted else ca
        total_liquid = pt + ro + hs + max(0.0, ca)
        row["Total_Liquid_Assets"] = total_liquid
        row["Total_Real"] = (total_liquid / (1 + infl) ** yir) if yir > 0 else total_liquid

        total_draws = ptd + rod + hsd + cad
        row["Total_Draws"] = total_draws
        row["Withdrawal_Rate"] = total_draws / total_liquid if total_liquid > 0 else 0.0

        br12 = bracket_ceiling(FEDERAL_BRACKETS, 0.12, yfb, binfl)
        br22 = bracket_ceiling(FEDERAL_BRACKETS, 0.22, yfb, binfl)
        row["Bracket_12_Ceiling"], row["Bracket_22_Ceiling"] = br12, br22

        results.append(row)

    return accum_rows, pd.DataFrame(results)


# ============================================================
# MONTE CARLO
# ============================================================

def run_monte_carlo(cfg, n_sims, std_dev, max_up, seed=None):
    rng = np.random.default_rng(seed)
    num_years = cfg["planning_end_age"] - cfg["retirement_age"] + 1
    all_runs = []
    for _ in range(n_sims):
        ov = {
            "pretax": generate_returns(cfg["pretax_return"], std_dev, max_up, num_years, rng),
            "roth": generate_returns(cfg["roth_return"], std_dev, max_up, num_years, rng),
            "hsa": generate_returns(cfg["hsa_return"], std_dev, max_up, num_years, rng),
            "cash": generate_returns(cfg["cash_return"], std_dev / 3, max_up / 2, num_years, rng),
        }
        _, df = run_simulation(cfg, return_overrides=ov)
        all_runs.append(df)
    return all_runs

def compute_percentile_bands(runs, col, pcts=(5, 25, 50, 75, 95)):
    ages = runs[0]["Age"].values
    mx = np.array([r[col].values for r in runs])
    bands = {f"p{p}": np.percentile(mx, p, axis=0) for p in pcts}
    bands["mean"] = np.mean(mx, axis=0); bands["Age"] = ages
    return bands


# ============================================================
# SCENARIO OPTIMIZER
# ============================================================

def run_optimizer(base_cfg, mc_sims, mc_std, mc_max, mc_seed):
    """
    Sweep retirement_age × ss_start_age × expense_reduction_post80
    and find combinations that achieve 100% MC success rate.
    Returns a DataFrame of all scenarios sorted by success rate.
    """
    rng_base = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)

    # Define search space
    ret_ages = list(range(base_cfg["retirement_age"], min(base_cfg["retirement_age"] + 4, 64)))
    ss_ages = list(range(max(62, base_cfg["ss_start_age"]), min(base_cfg["ss_start_age"] + 5, 71)))
    reductions = [0.0, 0.05, 0.10, 0.15, 0.20, 0.25]

    # Use fewer sims for optimizer (speed) — 50 is enough for ranking
    opt_sims = min(mc_sims, 50)
    num_years_max = base_cfg["planning_end_age"] - min(ret_ages) + 1

    results = []
    total_combos = len(ret_ages) * len(ss_ages) * len(reductions)

    for ret_age in ret_ages:
        for ss_age in ss_ages:
            if ss_age < ret_age:
                continue  # Can't claim SS before retirement in this model
            for red in reductions:
                test_cfg = {**base_cfg,
                            "retirement_age": ret_age,
                            "ss_start_age": ss_age,
                            "expense_reduction_post80": red}

                # Generate MC return paths (reuse same seed for fair comparison)
                rng = np.random.default_rng(mc_seed if mc_seed and mc_seed > 0 else 42)
                num_years = test_cfg["planning_end_age"] - ret_age + 1
                survived = 0
                final_assets = []
                final_real = []

                for _ in range(opt_sims):
                    ov = {
                        "pretax": generate_returns(test_cfg["pretax_return"], mc_std, mc_max, num_years, rng),
                        "roth": generate_returns(test_cfg["roth_return"], mc_std, mc_max, num_years, rng),
                        "hsa": generate_returns(test_cfg["hsa_return"], mc_std, mc_max, num_years, rng),
                        "cash": generate_returns(test_cfg["cash_return"], mc_std / 3, mc_max / 2, num_years, rng),
                    }
                    _, sim_df = run_simulation(test_cfg, return_overrides=ov)
                    last = sim_df.iloc[-1]
                    if last["Total_Liquid_Assets"] > 0:
                        survived += 1
                    final_assets.append(last["Total_Liquid_Assets"])
                    final_real.append(last["Total_Real"])

                success_rate = survived / opt_sims
                results.append({
                    "Retire_Age": ret_age,
                    "SS_Age": ss_age,
                    "Expense_Cut_80+": red,
                    "Success_Rate": success_rate,
                    "Median_Final_Assets": np.median(final_assets),
                    "Median_Final_Real": np.median(final_real),
                    "P10_Final_Assets": np.percentile(final_assets, 10),
                    "Worst_Case": min(final_assets),
                    "Extra_Work_Years": ret_age - base_cfg["retirement_age"],
                    "SS_Delay_Years": ss_age - base_cfg["ss_start_age"],
                })

    opt_df = pd.DataFrame(results).sort_values(
        ["Success_Rate", "Median_Final_Real"], ascending=[False, False]
    ).reset_index(drop=True)
    return opt_df


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_to_excel(accum_df, retire_df, cfg):
    """Create a professional Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    accum_fill = PatternFill("solid", fgColor="4472C4")
    input_font = Font(name="Arial", color="0000FF", size=10)  # Blue = inputs
    num_font = Font(name="Arial", size=10)
    pct_fmt = '0.0%'
    money_fmt = '$#,##0;($#,##0);"-"'
    money_k = '$#,##0'
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    def style_header(ws, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            mx = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(mx + 2, 10), 22)

    # ──────── Sheet 1: Assumptions ────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    assumptions = [
        ("RETIREMENT PLANNING MODEL", ""),
        ("", ""),
        ("TIMELINE", ""),
        ("Current Age", cfg["current_age"]),
        ("Retirement Age", cfg["retirement_age"]),
        ("Plan Through Age", cfg["planning_end_age"]),
        ("", ""),
        ("STARTING BALANCES (Today)", ""),
        ("PreTax 401(k)", cfg["pretax_401k"]),
        ("Roth IRA", cfg["roth_ira"]),
        ("HSA", cfg["hsa"]),
        ("S+ 5-Year", cfg["s_plus_5yr"]),
        ("S+ 10-Year", cfg["s_plus_10yr"]),
        ("Cash", cfg["cash"]),
        ("", ""),
        ("ANNUAL CONTRIBUTIONS (Working Years)", ""),
        ("401(k) Employee", cfg.get("contrib_401k", 24_500)),
        ("Roth 401(k)", cfg.get("contrib_roth401k", 8_000)),
        ("Roth IRA (per spouse)", cfg.get("contrib_roth_ira", 8_700)),
        ("HSA (family)", cfg.get("contrib_hsa", 8_700)),
        ("Mega Backdoor Roth", cfg.get("contrib_mega_backdoor", 29_000)),
        ("Employer Match", cfg.get("contrib_employer_match", 18_000)),
        ("Cash Savings", cfg.get("contrib_cash_annual", 150_000)),
        ("Final Year Cash Lump", cfg.get("contrib_cash_final_lump", 120_000)),
        ("", ""),
        ("PERFORMANCE", ""),
        ("PreTax Return", cfg["pretax_return"]),
        ("Roth Return", cfg["roth_return"]),
        ("HSA Return", cfg["hsa_return"]),
        ("Cash Return", cfg["cash_return"]),
        ("", ""),
        ("INCOME SOURCES", ""),
        ("SS Start Age", cfg["ss_start_age"]),
        ("SS Annual (future $)", cfg["ss_annual_amount"]),
        ("SS COLA", cfg["ss_cola"]),
        ("JSS Start Age", cfg["jss_start_age"]),
        ("JSS Annual", cfg["jss_annual_amount"]),
        ("JSS COLA", cfg["jss_cola"]),
        ("Rental Gross", cfg["rental_gross"]),
        ("", ""),
        ("EXPENSES & INFLATION", ""),
        ("Annual Expenses", cfg["base_annual_expenses"]),
        ("General Inflation", cfg["inflation_rate"]),
        ("Healthcare Pre-Medicare", cfg["healthcare_pre_medicare"]),
        ("Healthcare Post-Medicare", cfg["healthcare_post_medicare"]),
        ("Healthcare Inflation", cfg["healthcare_inflation"]),
        ("", ""),
        ("STRATEGY", ""),
        ("RMD Start Age", cfg["rmd_start_age"]),
        ("Roth Conversion Target", cfg["roth_conversion_target_bracket"]),
        ("IRMAA Avoidance", cfg["irmaa_avoidance"]),
    ]
    for r, (label, val) in enumerate(assumptions, 1):
        ws_a.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_a.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1:
            cell.number_format = pct_fmt
            cell.font = input_font
        elif isinstance(val, (int, float)) and val >= 1000:
            cell.number_format = money_k
            cell.font = input_font
        else:
            cell.font = input_font
    ws_a.column_dimensions["A"].width = 30
    ws_a.column_dimensions["B"].width = 18

    # ──────── Sheet 2: Accumulation Phase ────────
    if accum_df is not None and len(accum_df) > 0:
        ws_acc = wb.create_sheet("Accumulation")
        acc_cols = ["Age", "Year", "Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash",
                    "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "S_Plus_5yr", "S_Plus_10yr",
                    "Total_Liquid_Assets"]
        for c, col in enumerate(acc_cols, 1):
            ws_acc.cell(row=1, column=c, value=col.replace("_", " "))
        style_header(ws_acc, len(acc_cols))
        for r, row in enumerate(accum_df.to_dict("records") if isinstance(accum_df, pd.DataFrame) else accum_df, 2):
            for c, col in enumerate(acc_cols, 1):
                val = row.get(col, "")
                cell = ws_acc.cell(row=r, column=c, value=val)
                cell.font = num_font
                if col not in ("Age", "Year", "Phase"):
                    cell.number_format = money_k
        auto_width(ws_acc)

    # ──────── Sheet 3: Retirement Phase ────────
    ws_ret = wb.create_sheet("Retirement")
    ret_cols = [
        "Age", "Year", "Draw_Strategy",
        "SS_Income", "JSS_Income", "Rental_Income", "S_Plus_Income", "Passive_Income",
        "PreTax_Draw", "Roth_Draw", "Cash_Draw", "HSA_Draw",
        "Total_Income", "Total_Expenses", "Total_Tax", "Surplus_Deficit",
        "PreTax_EOY", "Roth_EOY", "HSA_EOY", "Cash_EOY", "Total_Liquid_Assets", "Total_Real",
        "Effective_Tax_Rate", "Withdrawal_Rate",
        "RMD", "Roth_Conversion", "Roth_Conversion_Tax",
        "MAGI", "IRMAA_Hit", "IRMAA_Cost",
        "Cum_Gifts", "Cum_Legacy_Roth", "Cum_Lump_Sums",
    ]
    for c, col in enumerate(ret_cols, 1):
        ws_ret.cell(row=1, column=c, value=col.replace("_", " "))
    style_header(ws_ret, len(ret_cols))

    pct_columns = {"Effective_Tax_Rate", "Withdrawal_Rate"}
    bool_columns = {"IRMAA_Hit"}
    text_columns = {"Draw_Strategy"}
    neg_fill = PatternFill("solid", fgColor="FFE0E0")

    for r, (_, row) in enumerate(retire_df.iterrows(), 2):
        for c, col in enumerate(ret_cols, 1):
            val = row.get(col, "")
            cell = ws_ret.cell(row=r, column=c, value=val)
            cell.font = num_font
            cell.border = thin_border
            if col in pct_columns:
                cell.number_format = pct_fmt
            elif col in bool_columns:
                pass
            elif col in text_columns or col in ("Age", "Year"):
                pass
            else:
                cell.number_format = money_k
            # Red highlight for negative values
            if isinstance(val, (int, float)) and val < 0:
                cell.fill = neg_fill
    auto_width(ws_ret)

    # ──────── Sheet 4: Summary ────────
    ws_s = wb.create_sheet("Summary")
    first, last = retire_df.iloc[0], retire_df.iloc[-1]
    summary = [
        ("RETIREMENT PLAN SUMMARY", ""),
        ("", ""),
        ("Starting Retirement Assets", first["Total_Liquid_Assets"]),
        ("  PreTax 401(k)", first["PreTax_EOY"]),
        ("  Roth IRA", first["Roth_EOY"]),
        ("  HSA", first["HSA_EOY"]),
        ("  Cash", first["Cash_EOY"]),
        ("", ""),
        (f"Assets at Age {cfg['planning_end_age']}", last["Total_Liquid_Assets"]),
        ("Real Value (inflation-adjusted)", last["Total_Real"]),
        ("", ""),
        ("Avg Withdrawal Rate", retire_df["Withdrawal_Rate"].mean()),
        ("Avg Effective Tax Rate", retire_df["Effective_Tax_Rate"].mean()),
        ("Total Roth Conversions", retire_df["Roth_Conversion"].sum()),
        ("Total Conversion Tax Paid", retire_df["Roth_Conversion_Tax"].sum()),
        ("", ""),
        ("Cumulative Gifts", last["Cum_Gifts"]),
        ("Cumulative Legacy Roth", last["Cum_Legacy_Roth"]),
        ("Cumulative Lump Sums", last["Cum_Lump_Sums"]),
        ("", ""),
        ("IRMAA Years Hit", int(retire_df["IRMAA_Hit"].sum())),
    ]
    for r, (label, val) in enumerate(summary, 1):
        ws_s.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=label.isupper(), size=11 if label.isupper() else 10)
        cell = ws_s.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val < 1 and val > 0:
            cell.number_format = pct_fmt
        elif isinstance(val, (int, float)) and abs(val) >= 1000:
            cell.number_format = money_k
        cell.font = num_font
    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 20

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# STREAMLIT UI
# ============================================================



def qp_int(name, default):
    try:
        return int(str(st.query_params.get(name, default)).replace(',', '').replace('$', '').strip())
    except:
        return default

def qp_float(name, default):
    try:
        return float(str(st.query_params.get(name, default)).replace('%', '').strip())
    except:
        return default

def main():
    st.set_page_config(page_title="Retirement Income Planner", page_icon="\U0001F4CA",
                       layout="wide", initial_sidebar_state="expanded")
    st.title("\U0001F3E6 Retirement Income & Tax Planning Simulator")
    st.caption("Accumulation + Retirement modeling with Monte Carlo and Excel export")

    with st.sidebar:
        st.header("\u2699\ufe0f Configuration")

        with st.expander("\U0001F464 Age & Timeline", expanded=True):
            current_age = st.number_input("Current Age", 45, 70, 55)
            retirement_age = st.slider("Retirement Age", 55, 63, 55)
            planning_end = st.slider("Plan Through Age", 85, 100, 89)

        with st.expander("\U0001F4B0 Current Balances (Today)", expanded=True):
            file_balances, file_loaded = load_starting_balances()
            if file_loaded:
                st.success("Loaded starting_balances.txt")
            else:
                st.info("No starting_balances.txt found; using defaults")
            pretax_default = qp_int("pretax", file_balances["401k"] if file_balances["401k"] is not None else 1_475_000)
            roth_default = qp_int("roth", file_balances["roth"] if file_balances["roth"] is not None else 510_000)
            hsa_default = qp_int("hsa", file_balances["hsa"] if file_balances["hsa"] is not None else 130_000)
            s5_default = qp_int("s5", file_balances["s_plus_5"] if file_balances["s_plus_5"] is not None else 300_000)
            s10_default = qp_int("s10", file_balances["s_plus_10"] if file_balances["s_plus_10"] is not None else 400_000)
            cash_default = qp_int("cash", file_balances["cash"] if file_balances["cash"] is not None else 745_000)
            pretax = st.number_input("PreTax 401(k) ($)", 0, 10_000_000, pretax_default, step=50_000, format="%d")
            roth_bal = st.number_input("Roth IRA ($)", 0, 5_000_000, roth_default, step=25_000, format="%d")
            hsa_bal = st.number_input("HSA ($)", 0, 500_000, hsa_default, step=10_000, format="%d")
            s5_bal = st.number_input("S+ 5-Year Payout ($)", 0, 2_000_000, s5_default, step=25_000, format="%d")
            s10_bal = st.number_input("S+ 10-Year Payout ($)", 0, 2_000_000, s10_default, step=25_000, format="%d")
            cash_bal = st.number_input("Cash ($)", 0, 1_000_000, cash_default, step=10_000, format="%d")

        with st.expander("\U0001F4BC Working Years Contributions"):
            st.caption(f"Annual contributions for {max(0, retirement_age - current_age)} remaining working years")
            c_401k = st.number_input("401(k) Employee ($)", 0, 50_000, 24_500, step=500, format="%d")
            c_roth401k = st.number_input("Roth 401(k) ($)", 0, 50_000, 8_000, step=500, format="%d")
            c_roth_ira = st.number_input("Roth IRA per Spouse ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_hsa = st.number_input("HSA Family ($)", 0, 15_000, 8_700, step=100, format="%d")
            c_mega = st.number_input("Mega Backdoor Roth ($)", 0, 50_000, 29_000, step=1_000, format="%d")
            c_match = st.number_input("Employer Match ($)", 0, 50_000, 18_000, step=1_000, format="%d")
            st.divider()
            c_cash = st.number_input("Annual Cash Savings ($)", 0, 500_000, 150_000, step=5_000, format="%d")
            c_cash_lump = st.number_input("Final Year Cash Lump ($)", 0, 500_000, 120_000, step=10_000, format="%d")

        with st.expander("\U0001F4C8 Performance Assumptions"):
            pretax_ret = st.slider("PreTax Target Return %", 0.0, 12.0, 6.0, 0.5) / 100
            roth_ret = st.slider("Roth Target Return %", 0.0, 12.0, 7.0, 0.5) / 100
            hsa_ret = st.slider("HSA Target Return %", 0.0, 12.0, 5.0, 0.5) / 100
            cash_ret = st.slider("Cash/MM Target Return %", 0.0, 8.0, 4.0, 0.25) / 100

        with st.expander("\U0001F3B2 Monte Carlo Settings"):
            mc_enabled = st.checkbox("Enable Monte Carlo", value=True)
            mc_sims = st.slider("Simulations", 50, 10000, 5000, 50)
            mc_std = st.slider("Return Std Dev %", 1.0, 25.0, 12.0, 0.5) / 100
            mc_max = st.slider("Max Upside Cap %", 5.0, 25.0, 18.0, 0.5) / 100
            mc_seed = st.number_input("Seed (0=random)", 0, 99999, 0)

        with st.expander("\U0001F4E5 Income Sources"):
            ss_age = st.slider("SS Start Age", 62, 70, 65)
            ss_amount = st.number_input("SS Annual (Future $)", 0, 200_000, 78_000, step=1_000, format="%d")
            ss_cola = st.slider("SS COLA %", 0.0, 5.0, 2.0, 0.25) / 100
            st.divider()
            jss_age = st.slider("JSS Pension Start Age", 60, 70, 60)
            jss_amount = st.number_input("JSS Annual ($)", 0, 100_000, 18_000, step=1_000, format="%d")
            jss_cola_pct = st.slider("JSS COLA %", 0.0, 3.0, 1.0, 0.25) / 100
            jss_recovery = st.number_input("JSS Recovery Years", 0, 10, 4)
            st.divider()
            rental = st.number_input("Rental Gross ($)", 0, 200_000, 44_000, step=1_000, format="%d")
            rental_tax_pct = st.slider("Rental Taxable %", 0, 100, 50) / 100

        with st.expander("\U0001F4B8 Expenses"):
            base_exp = st.number_input("Annual Expenses ($)", 50_000, 500_000, 168_000, step=5_000, format="%d")
            inflation = st.slider("General Inflation %", 0.0, 6.0, 2.87, 0.01) / 100
            st.subheader("Healthcare")
            hc_pre = st.number_input("Pre-Medicare ($)", 0, 100_000, 33_000, step=1_000, format="%d")
            hc_post = st.number_input("Post-Medicare ($)", 0, 50_000, 12_000, step=1_000, format="%d")
            hc_inflation = st.slider("HC Inflation %", 0.0, 10.0, 5.0, 0.5) / 100
            hdhp_cost = st.number_input("HDHP ($)", 0, 30_000, 12_000, step=1_000, format="%d")
            st.subheader("Post-80 Adjustment")
            exp_red_80 = st.slider("Expense Reduction After 80 (%)", 0, 25, 25, 5) / 100
            st.caption("Lifestyle slowdown: reduce base expenses after age 80.")
            st.subheader("Negative Return Adjustment")
            neg_ret_reduction = st.number_input("Discretionary Draw Reduction if Year Return < 0 ($)", 0, 200_000, 50_000, step=1000, format="%d")
            st.subheader("Lump Sum")
            lump_age = st.number_input("Lump at Age", 55, 95, 70)
            lump_amt = st.number_input("Lump Amount ($)", 0, 1_000_000, 0, step=10_000, format="%d")
            lump_sums = {lump_age: lump_amt} if lump_amt > 0 else {}

        with st.expander("\U0001F381 Gifts & Giving"):
            gifts = st.number_input("Annual Gifts ($)", 0, 100_000, 0, step=1_000, format="%d")

        with st.expander("\U0001F3AF Strategy", expanded=True):
            st.markdown("**PreTax→12% bracket, then Cash, then Roth. RMD dominates at 75+.**")
            rmd_start = st.slider("RMD Start Age", 73, 75, 75)
            perf_only = st.checkbox("Performance Draw Only", value=False)
            st.divider()
            roth_conv = st.checkbox("Enable Roth Conversions", value=False)
            roth_bracket = st.selectbox("Conversion Target", ["12%", "22%"])
            roth_margin = st.number_input("Min PreTax Keep ($)", 0, 1_000_000, 100_000, step=25_000, format="%d")
            st.divider()
            irmaa_avoid = st.checkbox("IRMAA Avoidance", value=True)

        with st.expander("\U0001F468\u200D\U0001F469\u200D\U0001F467\u200D\U0001F466 Legacy"):
            num_kids = st.number_input("Children", 0, 10, 4)
            roth_per_child = st.number_input("Roth/Child/Year ($)", 0, 50_000, 7_500, step=500, format="%d")
            legacy_years = st.slider("Legacy Duration (years)", 1, 20, 10)

        with st.expander("\U0001F3DB Tax"):
            or_resident = st.checkbox("Oregon Resident", value=True)
            std_ded = st.number_input("Std Deduction MFJ ($)", 0, 50_000, 32_500, step=100, format="%d")
            bracket_infl = st.slider("Bracket Inflation %", 0.0, 5.0, 2.67, 0.01) / 100

        with st.expander("\U0001F4CA Display Options"):
            show_real = st.toggle("Show in Today's Dollars", value=True)
            st.caption("When on, all dollar amounts are inflation-adjusted to present value.")

    # ── BUILD CONFIG ──
    cfg = dict(
        current_age=current_age, retirement_age=retirement_age, planning_end_age=planning_end,
        pretax_401k=pretax, roth_ira=roth_bal, hsa=hsa_bal,
        s_plus_5yr=s5_bal, s_plus_10yr=s10_bal, cash=cash_bal,
        contrib_401k=c_401k, contrib_roth401k=c_roth401k, contrib_roth_ira=c_roth_ira,
        contrib_hsa=c_hsa, contrib_mega_backdoor=c_mega, contrib_employer_match=c_match,
        contrib_cash_annual=c_cash, contrib_cash_final_lump=c_cash_lump,
        pretax_return=pretax_ret, roth_return=roth_ret, hsa_return=hsa_ret, cash_return=cash_ret,
        ss_start_age=ss_age, ss_annual_amount=ss_amount, ss_cola=ss_cola,
        jss_start_age=jss_age, jss_annual_amount=jss_amount, jss_cola=jss_cola_pct,
        jss_recovery_years=jss_recovery, rental_gross=rental, rental_taxable_pct=rental_tax_pct,
        base_annual_expenses=base_exp, inflation_rate=inflation,
        healthcare_pre_medicare=hc_pre, healthcare_post_medicare=hc_post,
        healthcare_inflation=hc_inflation, hdhp_annual=hdhp_cost,
        expense_reduction_post80=exp_red_80,
        gifts_annual=gifts, lump_sums=lump_sums,
        standard_deduction=std_ded, bracket_inflation=bracket_infl, oregon_resident=or_resident,
        rmd_start_age=rmd_start, draw_order=["pretax", "cash", "roth", "hsa"],
        roth_conversion_enabled=roth_conv, roth_conversion_target_bracket=roth_bracket,
        roth_conversion_margin=roth_margin, irmaa_avoidance=irmaa_avoid,
        performance_draw_only=perf_only,
        num_children=num_kids, roth_legacy_per_child=roth_per_child,
        legacy_years=legacy_years, legacy_inflation=inflation,
    )

    # ── RUN ──
    accum_rows, df = run_simulation(cfg)
    accum_df = pd.DataFrame(accum_rows) if accum_rows else None
    mc_runs = None
    if mc_enabled:
        mc_runs = run_monte_carlo(cfg, mc_sims, mc_std, mc_max, seed=mc_seed if mc_seed > 0 else None)

    # ── DEFLATOR (today's dollars conversion) ──
    # When show_real is True, deflate future dollar amounts back to present value
    dollar_label = "Today's $" if show_real else "Future $"
    def deflate(series_or_val, year_col_or_yir):
        """Deflate future dollars to today's dollars."""
        if not show_real:
            return series_or_val
        if isinstance(year_col_or_yir, (pd.Series, np.ndarray)):
            factors = (1 + inflation) ** year_col_or_yir
        else:
            factors = (1 + inflation) ** year_col_or_yir
        return series_or_val / factors

    # Build display-ready retirement df
    df_disp = df.copy()
    if show_real:
        # Years from today (accumulation years + retirement years)
        accum_years = max(0, retirement_age - current_age)
        df_disp["_deflate_yrs"] = accum_years + df_disp["Years_Retired"]
        money_cols = [c for c in df_disp.columns if c not in (
            "Age", "Year", "Years_Retired", "Phase", "Draw_Strategy",
            "Effective_Tax_Rate", "Withdrawal_Rate", "PreTax_WR",
            "IRMAA_Hit", "Under_700_FPL", "In_12_Bracket", "In_22_Bracket",
            "Deduction_Type", "PreTax_Depleted", "Roth_Depleted", "HSA_Depleted",
            "Return_PreTax", "Return_Roth", "Return_HSA", "Return_Cash",
            "RMD_Excess", "_deflate_yrs")]
        for c in money_cols:
            if c in df_disp.columns and df_disp[c].dtype in [np.float64, np.int64, float, int]:
                df_disp[c] = df_disp[c] / ((1 + inflation) ** df_disp["_deflate_yrs"])
        df_disp.drop(columns=["_deflate_yrs"], inplace=True)

    # ── KPI ROW ──
    c1, c2, c3, c4, c5 = st.columns(5)
    first, last = df_disp.iloc[0], df_disp.iloc[-1]
    if accum_df is not None and len(accum_df) > 0:
        c1.metric(f"Today's Assets", f"${accum_df.iloc[0]['Total_Liquid_Assets']:,.0f}")
    else:
        c1.metric("Starting Assets", f"${first['Total_Liquid_Assets']:,.0f}")
    c2.metric(f"At Retirement ({dollar_label})", f"${first['Total_Liquid_Assets']:,.0f}")
    c3.metric(f"Age {planning_end} ({dollar_label})", f"${last['Total_Liquid_Assets']:,.0f}")

    depleted = df[df["Total_Liquid_Assets"] <= 0]
    if len(depleted) > 0:
        c4.metric("\u26a0\ufe0f Depleted", f"Age {int(depleted.iloc[0]['Age'])}", delta="RISK", delta_color="inverse")
    else:
        c4.metric("\u2705 Lasts To", f"Age {planning_end}+")

    if mc_runs:
        surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
        c5.metric("MC Success", f"{surv/len(mc_runs)*100:.0f}%", delta=f"{len(mc_runs)} sims", delta_color="off")
    else:
        c5.metric("Avg WR", f"{df['Withdrawal_Rate'].mean():.1%}")

    st.divider()

    # ── TABS ──
    tab_names = ["\U0001F4BC Accumulation", "\U0001F4CA Retirement Assets", "\U0001F4B0 Income & Expenses",
                 "\U0001F3DB Tax", "\U0001F504 Roth Conversions"]
    if mc_runs: tab_names.append("\U0001F3B2 Monte Carlo")
    if mc_runs: tab_names.append("\U0001F3AF Optimizer")
    tab_names.append("\U0001F4CB Full Data")
    tabs = st.tabs(tab_names)
    ti = 0

    # ── Accumulation Tab ──
    with tabs[ti]:
        if accum_df is not None and len(accum_df) > 0:
            yrs = retirement_age - current_age
            total_contrib = accum_df[["Contrib_PreTax", "Contrib_Roth", "Contrib_HSA", "Contrib_Cash"]].sum().sum()
            ac1, ac2, ac3 = st.columns(3)
            ac1.metric("Working Years", f"{yrs}")
            ac2.metric("Total Contributions", f"${total_contrib:,.0f}")
            ac3.metric("Retirement Day Assets", f"${first['Total_Liquid_Assets']:,.0f}")

            fig = go.Figure()
            for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"),
                                     ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                     ("Roth_EOY","Roth","rgba(142,68,173,0.6)"),
                                     ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
                fig.add_trace(go.Scatter(x=accum_df["Age"], y=accum_df[col], mode="lines", name=name,
                                         stackgroup="one", line=dict(width=0.5), fillcolor=color))
            fig.update_layout(title="Accumulation Phase: Account Growth",
                              xaxis_title="Age", yaxis_title="$", yaxis_tickformat="$,.0f", height=450)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(accum_df, use_container_width=True, height=300)
        else:
            st.info("No accumulation phase (already at retirement age).")
    ti += 1

    # ── Retirement Assets Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, name, color in [("Cash_EOY","Cash","rgba(46,134,193,0.6)"), ("HSA_EOY","HSA","rgba(39,174,96,0.6)"),
                                 ("Roth_EOY","Roth","rgba(142,68,173,0.6)"), ("PreTax_EOY","PreTax","rgba(231,76,60,0.6)")]:
            fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp[col], mode="lines", name=name,
                                     stackgroup="one", line=dict(width=0.5), fillcolor=color))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Real"], mode="lines", name="Real Value",
                                 line=dict(color="black", width=2, dash="dash")))
        fig.update_layout(title="Retirement Account Balances", xaxis_title="Age", yaxis_title="$",
                          yaxis_tickformat="$,.0f", hovermode="x unified", height=500)
        st.plotly_chart(fig, use_container_width=True)

        colors = ["red" if x > 0.04 else "orange" if x > 0.03 else "green" for x in df["Withdrawal_Rate"]]
        fig2 = go.Figure(go.Bar(x=df["Age"], y=df["Withdrawal_Rate"]*100, marker_color=colors))
        fig2.add_hline(y=4, line_dash="dash", line_color="red", annotation_text="4% Rule")
        fig2.update_layout(title="Withdrawal Rate", xaxis_title="Age", yaxis_title="%", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Income & Expenses Tab ──
    with tabs[ti]:
        fig = go.Figure()
        for col, nm in [("SS_Income","SS"),("JSS_Income","JSS"),("Rental_Income","Rental"),
                        ("S_Plus_Income","S+"),("PreTax_Draw","PreTax Draw"),("Roth_Draw","Roth Draw")]:
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp[col], name=nm))
        fig.add_trace(go.Scatter(x=df_disp["Age"], y=df_disp["Total_Expenses"], mode="lines+markers",
                                 name="Expenses", line=dict(color="red", width=3)))
        fig.update_layout(barmode="stack", title="Income vs Expenses", xaxis_title="Age",
                          yaxis_title="$", yaxis_tickformat="$,.0f", height=500)
        st.plotly_chart(fig, use_container_width=True)

        sd_c = ["green" if v >= 0 else "red" for v in df_disp["Surplus_Deficit"]]
        fig2 = go.Figure(go.Bar(x=df_disp["Age"], y=df_disp["Surplus_Deficit"], marker_color=sd_c))
        fig2.update_layout(title="Surplus / Deficit", xaxis_title="Age", yaxis_title="$",
                           yaxis_tickformat="$,.0f", height=350)
        st.plotly_chart(fig2, use_container_width=True)
    ti += 1

    # ── Tax Tab ──
    with tabs[ti]:
        ca2, cb2 = st.columns(2)
        with ca2:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Federal_Tax"], name="Federal"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["Oregon_Tax"], name="Oregon"))
            fig.add_trace(go.Bar(x=df_disp["Age"], y=df_disp["IRMAA_Cost"], name="IRMAA"))
            fig.update_layout(barmode="stack", title="Tax Burden", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
        with cb2:
            fig = go.Figure(go.Scatter(x=df["Age"], y=df["Effective_Tax_Rate"]*100,
                                       mode="lines+markers", line=dict(color="darkblue", width=2)))
            fig.update_layout(title="Effective Tax Rate", yaxis_title="%", height=400)
            st.plotly_chart(fig, use_container_width=True)

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Federal_Taxable_Income"], name="Taxable", line=dict(color="blue", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_12_Ceiling"], name="12% Ceil", line=dict(color="green", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Bracket_22_Ceiling"], name="22% Ceil", line=dict(color="orange", dash="dash")))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["MAGI"], name="MAGI", line=dict(color="red", width=1)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["IRMAA_Threshold"], name="IRMAA Thr", line=dict(color="darkred", dash="dot")))
        fig.update_layout(title="Income vs Brackets & IRMAA", yaxis_tickformat="$,.0f", height=450)
        st.plotly_chart(fig, use_container_width=True)
    ti += 1

    # ── Roth Conversions Tab ──
    with tabs[ti]:
        cdf = df[df["Roth_Conversion"] > 0]
        if len(cdf) > 0:
            fig = go.Figure()
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion"], name="Converted", marker_color="purple"))
            fig.add_trace(go.Bar(x=cdf["Age"], y=cdf["Roth_Conversion_Tax"], name="Tax", marker_color="red"))
            fig.update_layout(barmode="group", title="Roth Conversions", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig, use_container_width=True)
            tc, tt = cdf["Roth_Conversion"].sum(), cdf["Roth_Conversion_Tax"].sum()
            if tc > 0: st.info(f"**Total:** ${tc:,.0f} | **Tax:** ${tt:,.0f} | **Rate:** {tt/tc:.1%}")
        else:
            st.info("No conversions.")
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df["Age"], y=df["PreTax_EOY"], name="PreTax", line=dict(color="red", width=2)))
        fig.add_trace(go.Scatter(x=df["Age"], y=df["Roth_EOY"], name="Roth", line=dict(color="purple", width=2)))
        fig.update_layout(title="PreTax vs Roth", yaxis_tickformat="$,.0f", height=400)
        st.plotly_chart(fig, use_container_width=True)
    ti += 1

    # ── Monte Carlo Tab ──
    if mc_runs:
        with tabs[ti]:
            surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
            m1, m2, m3 = st.columns(3)
            m1.metric("Sims", len(mc_runs)); m2.metric("Survived", surv); m3.metric("Rate", f"{surv/len(mc_runs)*100:.1f}%")

            bands = compute_percentile_bands(mc_runs, "Total_Liquid_Assets")
            ages = bands["Age"]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p95"],bands["p5"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.1)", line=dict(color="rgba(255,255,255,0)"), name="5-95%"))
            fig.add_trace(go.Scatter(x=np.concatenate([ages,ages[::-1]]),
                y=np.concatenate([bands["p75"],bands["p25"][::-1]]),
                fill="toself", fillcolor="rgba(99,110,250,0.25)", line=dict(color="rgba(255,255,255,0)"), name="25-75%"))
            fig.add_trace(go.Scatter(x=ages, y=bands["p50"], name="Median", line=dict(color="blue", width=2)))
            fig.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                     line=dict(color="black", width=2, dash="dash")))
            fig.update_layout(title=f"MC Total Assets ({len(mc_runs)} runs)", yaxis_tickformat="$,.0f", height=500)
            st.plotly_chart(fig, use_container_width=True)

            n_show = min(20, len(mc_runs))
            fig3 = go.Figure()
            for i in range(n_show):
                fig3.add_trace(go.Scatter(x=mc_runs[i]["Age"], y=mc_runs[i]["Total_Liquid_Assets"],
                    mode="lines", line=dict(width=0.7, color="rgba(99,110,250,0.3)"), showlegend=False))
            fig3.add_trace(go.Scatter(x=df["Age"], y=df["Total_Liquid_Assets"], name="Baseline",
                                      line=dict(color="black", width=2, dash="dash")))
            fig3.update_layout(title=f"{n_show} Sample Paths", yaxis_tickformat="$,.0f", height=400)
            st.plotly_chart(fig3, use_container_width=True)
        ti += 1

    # ── Optimizer Tab ──
    if mc_runs:
        with tabs[ti]:
            st.subheader("\U0001F3AF Scenario Optimizer — Path to 100% Success")
            st.caption(
                "Sweeps retirement age delay, SS start age, and post-80 expense reduction "
                "to find combinations that maximize Monte Carlo success rate."
            )

            if st.button("Run Optimizer", type="primary"):
                with st.spinner("Running scenario sweep (this takes a moment)..."):
                    opt_df = run_optimizer(cfg, mc_sims, mc_std, mc_max, mc_seed)

                # ── Current scenario baseline ──
                surv = sum(1 for r in mc_runs if r.iloc[-1]["Total_Liquid_Assets"] > 0)
                base_rate = surv / len(mc_runs) * 100
                st.info(f"**Current settings:** Retire {cfg['retirement_age']}, "
                        f"SS @ {cfg['ss_start_age']}, "
                        f"No post-80 cut → **{base_rate:.0f}% success**")

                # ── Top recommendations ──
                perfect = opt_df[opt_df["Success_Rate"] >= 1.0]
                if len(perfect) > 0:
                    st.success(f"\u2705 Found **{len(perfect)}** scenarios with 100% success rate!")
                    st.markdown("**Top 5 — least disruption to reach 100%:**")
                    # Sort by least change needed
                    perfect_sorted = perfect.sort_values(
                        ["Extra_Work_Years", "SS_Delay_Years", "Expense_Cut_80+"]
                    ).head(5)
                    for _, row in perfect_sorted.iterrows():
                        changes = []
                        if row["Extra_Work_Years"] > 0:
                            changes.append(f"retire at **{int(row['Retire_Age'])}** (+{int(row['Extra_Work_Years'])}yr)")
                        if row["SS_Delay_Years"] > 0:
                            changes.append(f"SS at **{int(row['SS_Age'])}** (+{int(row['SS_Delay_Years'])}yr)")
                        if row["Expense_Cut_80+"] > 0:
                            changes.append(f"**{row['Expense_Cut_80+']*100:.0f}%** expense cut after 80")
                        med_val = row["Median_Final_Real"] if show_real else row["Median_Final_Assets"]
                        label = "today's $" if show_real else "future $"
                        desc = " + ".join(changes) if changes else "No changes needed!"
                        st.markdown(f"- {desc} → Median final: **${med_val:,.0f}** ({label})")
                else:
                    st.warning("No 100% success scenarios found. Best options below.")

                # ── Full results table ──
                st.subheader("All Scenarios")
                display_opt = opt_df.copy()
                display_opt["Success_Rate"] = display_opt["Success_Rate"].apply(lambda x: f"{x:.0%}")
                display_opt["Expense_Cut_80+"] = display_opt["Expense_Cut_80+"].apply(lambda x: f"{x:.0%}")
                money_opt_cols = ["Median_Final_Assets", "Median_Final_Real", "P10_Final_Assets", "Worst_Case"]
                for c in money_opt_cols:
                    display_opt[c] = display_opt[c].apply(lambda x: f"${x:,.0f}")
                st.dataframe(display_opt, use_container_width=True, height=400)
            else:
                st.markdown(
                    "Click **Run Optimizer** to sweep combinations of:\n"
                    "- Retirement age: current → +3 years\n"
                    "- SS start age: current → +4 years\n"
                    "- Post-80 expense reduction: 0% to 25%\n\n"
                    "Uses 50 MC sims per scenario for speed."
                )
        ti += 1

    # ── Full Data Tab ──
    with tabs[ti]:
        st.caption(f"Displaying in **{dollar_label}**")
        dcols = ["Age","Year","Draw_Strategy",
                 "SS_Income","JSS_Income","Rental_Income","S_Plus_Income","Passive_Income",
                 "PreTax_Draw","Roth_Draw","Cash_Draw","HSA_Draw",
                 "PreTax_EOY","Roth_EOY","HSA_EOY","Cash_EOY",
                 "Total_Liquid_Assets","Total_Income","Total_Expenses","Total_Tax",
                 "Effective_Tax_Rate","Withdrawal_Rate","Surplus_Deficit",
                 "Roth_Conversion","RMD","RMD_Excess","IRMAA_Hit",
                 "Cum_Gifts","Cum_Legacy_Roth","Cum_Lump_Sums"]
        avail = [c for c in dcols if c in df_disp.columns]
        show = st.multiselect("Columns", df_disp.columns.tolist(), default=avail)
        disp = df_disp[show].copy()
        no_fmt = {"Age","Year","Years_Retired","Effective_Tax_Rate","Withdrawal_Rate","PreTax_WR",
                  "IRMAA_Hit","Under_700_FPL","In_12_Bracket","In_22_Bracket","Draw_Strategy",
                  "PreTax_Depleted","Roth_Depleted","HSA_Depleted","Return_PreTax","Return_Roth","Return_HSA","Return_Cash","Phase"}
        pct = {"Effective_Tax_Rate","Withdrawal_Rate","PreTax_WR","Return_PreTax","Return_Roth","Return_HSA","Return_Cash"}
        fmt = {}
        for c in show:
            if c in pct: fmt[c] = "{:.1%}"
            elif c not in no_fmt: fmt[c] = "${:,.0f}"
        st.dataframe(disp.style.format(fmt, na_rep="").apply(
            lambda x: ["background-color:#ffcccc" if isinstance(v,(int,float)) and v<0 else "" for v in x], axis=1),
            use_container_width=True, height=600)

        # ── DOWNLOADS ──
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("\U0001F4E5 Download CSV", df.to_csv(index=False), "retirement_plan.csv", "text/csv")
        with col_dl2:
            xlsx_buf = export_to_excel(accum_df, df, cfg)
            st.download_button("\U0001F4E5 Download Excel Report", xlsx_buf,
                               "retirement_plan.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── Notes ──
    st.divider()
    with st.expander("Model Notes"):
        st.markdown("""
**Accumulation Phase:** Balances grow from current age to retirement using configured returns. Contributions include 401(k), Roth 401(k), Roth IRA (x2 MFJ), HSA, Mega Backdoor Roth, employer match, cash savings, and a final-year lump sum. S+ deferred comp also grows during employment.

**Draw Strategy:** Pre-RMD: PreTax up to 12% bracket → Cash → Roth → HSA. RMD Phase (75+): mandatory RMD first → Cash → Roth → HSA. Roth conversions fill remaining bracket space.

**Monte Carlo:** Returns ~ N(target, std_dev), capped at max upside. Cash uses 1/3 std dev.

**Excel Export:** 4-sheet workbook: Assumptions, Accumulation, Retirement, Summary.
        """)


if __name__ == "__main__":
    main()
